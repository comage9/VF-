# -*- coding: utf-8 -*-
"""Compare warehouse zip inventory vs enhanced inventory_stock formula."""
from __future__ import annotations

import os
import sys
from collections import defaultdict
from pathlib import Path

import django

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

import openpyxl
from django.db.models import Sum
from django.db.models.functions import Coalesce

from sales_api.inventory_stock import (
    aggregate_movements_after_baseline,
    compute_current_stock,
)
from sales_api.models import (
    InventoryBaselineItem,
    InventoryBaselineUpload,
    InventoryReceiptItem,
    OutboundRecord,
)


def load_warehouse_xlsx_folder(folder: Path) -> dict[str, dict]:
    """barcode -> {name, qty, pages}"""
    by_bc: dict[str, dict] = {}
    files = sorted(folder.glob("*.xlsx"))
    for fp in files:
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # find header
        header_idx = 0
        headers = [str(c or "").strip() for c in rows[0]]
        for i, r in enumerate(rows[:5]):
            cells = [str(c or "").strip() for c in r]
            joined = " ".join(cells)
            if any(k in joined for k in ("바코드", "barcode", "Barcode", "SKU", "재고")):
                headers = cells
                header_idx = i
                break

        def col(*names):
            for n in names:
                for j, h in enumerate(headers):
                    if n.lower() in h.lower() or h == n:
                        return j
            return None

        # also try common Korean headers by scanning first data row heuristics
        # 상품 바코드 / 수량 (페이지별재고리스트 고정 컬럼)
        i_bc = col("상품 바코드", "상품바코드", "바코드", "barcode")
        i_name = col("상품명", "품명", "상품이름")
        i_qty = col("수량", "재고", "현재고")
        # 할당수량 제외 — '수량' 컬럼만
        if i_qty is not None and headers[i_qty] == "할당수량":
            for j, h in enumerate(headers):
                if h == "수량":
                    i_qty = j
                    break

        if fp == files[0]:
            print("HEADER_MAP", {"bc": i_bc, "name": i_name, "qty": i_qty, "headers": headers})

        for r in rows[header_idx + 1 :]:
            if not r or all(c is None or str(c).strip() == "" for c in r):
                continue
            bc = ""
            name = ""
            qty = None
            if i_bc is not None and i_bc < len(r):
                bc = str(r[i_bc] or "").strip().replace(" ", "")
                if bc.endswith(".0"):
                    bc = bc[:-2]
            if i_name is not None and i_name < len(r):
                name = str(r[i_name] or "").strip()
            if i_qty is not None and i_qty < len(r):
                try:
                    qty = int(float(str(r[i_qty]).replace(",", "")))
                except Exception:
                    qty = None
            if not bc or not bc.startswith("R") or qty is None:
                continue
            if bc in by_bc:
                by_bc[bc]["qty"] += qty
                by_bc[bc]["pages"] += 1
                if name and not by_bc[bc]["name"]:
                    by_bc[bc]["name"] = name
            else:
                by_bc[bc] = {"name": name, "qty": qty, "pages": 1}
    return by_bc


def system_current_stock() -> tuple[dict[str, dict], object]:
    latest = InventoryBaselineUpload.objects.order_by("-as_of_date", "-uploaded_at").first()
    if not latest:
        raise SystemExit("No baseline upload")
    as_of = latest.as_of_date
    items = list(
        InventoryBaselineItem.objects.filter(upload=latest).values(
            "barcode", "product_name", "quantity_box"
        )
    )
    # barcode 별 baseline 합산 (로케이션 분할 행 가능)
    base_map = {}
    for it in items:
        bc = (it["barcode"] or "").strip()
        if not bc:
            continue
        q = int(it.get("quantity_box") or 0)
        if bc in base_map:
            base_map[bc]["base"] += q
            if it.get("product_name") and not base_map[bc]["name"]:
                base_map[bc]["name"] = it["product_name"]
        else:
            base_map[bc] = {
                "name": it["product_name"] or "",
                "base": q,
            }

    # barcodes with only post-baseline receipts not in baseline
    extra_rcv = (
        InventoryReceiptItem.objects.filter(receipt_date__gt=as_of)
        .values("barcode")
        .annotate(q=Coalesce(Sum("quantity_box"), 0))
    )
    for row in extra_rcv:
        bc = (row["barcode"] or "").strip()
        if bc and bc not in base_map:
            base_map[bc] = {"name": "", "base": 0}

    barcodes = list(base_map.keys())
    out_agg, rcv_agg = aggregate_movements_after_baseline(
        as_of=as_of,
        barcodes=barcodes,
        outbound_model=OutboundRecord,
        receipt_model=InventoryReceiptItem,
    )

    result = {}
    for bc, meta in base_map.items():
        out_q = int(out_agg.get(bc, 0) or 0)
        rcv_q = int(rcv_agg.get(bc, 0) or 0)
        cur = compute_current_stock(meta["base"], rcv_q, out_q)
        result[bc] = {
            "name": meta["name"],
            "base": meta["base"],
            "rcv": rcv_q,
            "out": out_q,
            "current": cur,
        }
    return result, latest


def main():
    folder = Path(r"C:\Users\kis\AppData\Local\Temp\vf_inv_compare_20260720")
    # peek headers from first file
    files = sorted(folder.glob("*.xlsx"))
    wb = openpyxl.load_workbook(files[0], data_only=True)
    ws = wb.active
    print("=== SAMPLE SHEET ===")
    for i, row in enumerate(ws.iter_rows(max_row=12, values_only=True), 1):
        print(i, row)

    wh = load_warehouse_xlsx_folder(folder)
    sys_map, latest = system_current_stock()
    print("\n=== BASELINE ===")
    print("as_of", latest.as_of_date, "upload", latest.uploaded_at, "id", latest.id)
    print("warehouse barcodes", len(wh), "system barcodes", len(sys_map))

    all_bc = set(wh) | set(sys_map)
    mismatches = []
    only_wh = []
    only_sys = []
    match = 0
    for bc in sorted(all_bc):
        w = wh.get(bc)
        s = sys_map.get(bc)
        if w and not s:
            if w["qty"] != 0:
                only_wh.append((bc, w))
            continue
        if s and not w:
            if s["current"] != 0:
                only_sys.append((bc, s))
            continue
        if not w or not s:
            continue
        if int(w["qty"]) != int(s["current"]):
            mismatches.append(
                {
                    "barcode": bc,
                    "name": w.get("name") or s.get("name"),
                    "warehouse": w["qty"],
                    "system": s["current"],
                    "base": s["base"],
                    "rcv": s["rcv"],
                    "out": s["out"],
                    "diff": s["current"] - w["qty"],  # system - warehouse
                }
            )
        else:
            match += 1

    mismatches.sort(key=lambda x: abs(x["diff"]), reverse=True)
    print("\n=== SUMMARY ===")
    print("exact_match", match)
    print("mismatch", len(mismatches))
    print("only_warehouse_nonzero", len(only_wh))
    print("only_system_nonzero", len(only_sys))
    total_wh = sum(v["qty"] for v in wh.values())
    total_sys = sum(v["current"] for v in sys_map.values())
    print("total_qty warehouse", total_wh, "system", total_sys, "diff", total_sys - total_wh)

    print("\n=== TOP MISMATCH (system - warehouse) ===")
    for m in mismatches[:40]:
        print(
            f"{m['barcode']}\t wh={m['warehouse']}\t sys={m['system']}\t "
            f"diff={m['diff']}\t base={m['base']} rcv={m['rcv']} out={m['out']}\t {m['name'][:40]}"
        )

    if only_wh[:15]:
        print("\n=== ONLY IN WAREHOUSE (sample) ===")
        for bc, w in only_wh[:15]:
            print(bc, w["qty"], w["name"][:40])
    if only_sys[:15]:
        print("\n=== ONLY IN SYSTEM nonzero (sample) ===")
        for bc, s in only_sys[:15]:
            print(bc, s["current"], "base", s["base"], "rcv", s["rcv"], "out", s["out"], s["name"][:40])

    # write report
    out_path = Path(__file__).resolve().parents[1] / "data" / "inventory_compare_report_20260720.txt"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"as_of={latest.as_of_date}\n")
        f.write(f"match={match} mismatch={len(mismatches)}\n")
        f.write(f"only_wh={len(only_wh)} only_sys={len(only_sys)}\n")
        f.write(f"total_wh={total_wh} total_sys={total_sys}\n\n")
        for m in mismatches:
            f.write(
                f"{m['barcode']}\t{m['warehouse']}\t{m['system']}\t{m['diff']}\t"
                f"base={m['base']}\trcv={m['rcv']}\tout={m['out']}\t{m['name']}\n"
            )
    print("\nreport", out_path)


if __name__ == "__main__":
    main()
