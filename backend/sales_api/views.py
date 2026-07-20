from rest_framework import status, viewsets, generics
from rest_framework.decorators import api_view
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.response import Response
from django.db import models
from django.db.models import Sum, Count, Min, Max, Value, DecimalField, Q
from django.db.models.functions import Coalesce
from django.db.models.functions import TruncDay, TruncMonth, TruncWeek, TruncYear
from django.http import HttpResponse
from django.http import JsonResponse
from django.utils import timezone
from django.db import transaction
from django.db import connection
from .models import (
    OutboundRecord,
    InventoryItem,
    DataSource,
    DeliveryDailyRecord,
    DeliverySpecialNote,
    BarcodeTransferRecord,
    BarcodeMaster,
    InventoryBaselineUpload,
    InventoryBaselineItem,
    InventoryReceiptUpload,
    InventoryReceiptItem,
    MasterSpec,
    ProductionLog,
    MachineUser,
    MachinePlan,
    InboundOrderUpload,
    InboundOrderLine,
    InboundPolicy,
    FCInboundRecord,
    FCInboundFileUpload,
    OutboundAnalysis,
)
from .serializers import (
    FCInboundRecordSerializer,
    FCInboundFileUploadSerializer,
    OutboundRecordSerializer,
    InventoryItemSerializer,
    DataSourceSerializer,
    DeliverySpecialNoteSerializer,
    InboundOrderUploadSerializer,
    InboundOrderLineSerializer,
    InboundPolicySerializer,
    ProductionLogSerializer,
)
from datetime import datetime, timedelta
from decimal import Decimal
import json
import uuid
import hashlib
import os
import logging
import traceback
import pandas as pd
import csv
import io
import urllib.request
import urllib.error
import csv
import re


def _extract_korean_only(text: str) -> str:
    """영어 thinking 부분을 제거하고 한국어 분석 결과만 추출"""
    if not text:
        return text

    # thinking/thought 키워드 패턴 제거
    thinking_patterns = [
        r"^\s*\[.*?\]\s*",
        r"^\s*Thought[\s:].*$",
        r"^\s*Thinking[\s:].*$",
        r"^\s*##?\s*(thoughts?|thinking|reasoning)",
        r"^\s*<!--.*?-->\s*$",
        r"^\s*###\s+Thought\s*$",
    ]

    lines = text.split("\n")
    result = []
    in_code_block = False

    for line in lines:
        if line.strip().startswith("```"):
            in_code_block = not in_code_block
            result.append(line)
            continue

        if in_code_block:
            result.append(line)
            continue

        stripped = line.strip()
        skip_line = False
        for pattern in thinking_patterns:
            if re.match(pattern, stripped, re.IGNORECASE):
                skip_line = True
                break

        if skip_line:
            continue

        korean_count = len(re.findall(r"[가-힣]", line))
        english_count = len(re.findall(r"[a-zA-Z]", line))
        if korean_count >= 3 and korean_count > english_count:
            line = re.sub(r"^Line\d+\s*:\s*", "", line)
            line = re.sub(r"^\d+[\)\.\s]*", "", line)
            if line.strip():
                result.append(line.strip())

    return "\n".join(result)


logger = logging.getLogger("sales_api.inventory")

_MASTER_SPECS = []
_MASTER_SPEC_NEXT_ID = 1

_PRODUCTION_LOG = []
_PRODUCTION_NEXT_ID = 1


_PRODUCTION_STATUS_VALUES = {"pending", "started", "ended", "stopped"}


def _production_calc_total(quantity, unit_quantity, current_total=None):
    try:
        q = int(float(quantity))
    except (ValueError, TypeError):
        q = 0
    try:
        uq = int(float(unit_quantity))
    except (ValueError, TypeError):
        uq = 0
    if q < 0:
        q = 0
    if uq < 0:
        uq = 0
    return q * uq


def _production_normalize_status(value: str):
    s = (value or "").strip().lower()
    if s in _PRODUCTION_STATUS_VALUES:
        return s
    # legacy mapping
    if s in ("in-progress", "inprogress", "progress"):
        return "started"
    if s in ("completed", "complete", "done"):
        return "ended"
    return "pending"


def _production_apply_status(item: dict, status_value: str):
    now_iso = timezone.now().isoformat()
    s = _production_normalize_status(status_value)
    item["status"] = s
    if s == "pending":
        item["startTime"] = None
        item["endTime"] = None
    elif s == "started":
        if not item.get("startTime"):
            item["startTime"] = now_iso
        item["endTime"] = None
    elif s == "ended":
        if not item.get("startTime"):
            item["startTime"] = now_iso
        if not item.get("endTime"):
            item["endTime"] = now_iso
    elif s == "stopped":
        if not item.get("startTime"):
            item["startTime"] = now_iso
        if not item.get("endTime"):
            item["endTime"] = now_iso
    return item


def _production_apply_status_model(obj: ProductionLog, status_value: str):
    now = timezone.now()
    s = _production_normalize_status(status_value)
    obj.status = s
    if s == "pending":
        obj.start_time = None
        obj.end_time = None
    elif s == "started":
        if not obj.start_time:
            obj.start_time = now
        obj.end_time = None
    elif s in ("ended", "stopped"):
        if not obj.start_time:
            obj.start_time = now
        if not obj.end_time:
            obj.end_time = now
    return obj


def _production_model_to_dict(obj: ProductionLog):
    return {
        "id": obj.id,
        "date": obj.date.isoformat() if obj.date else "",
        "machineNumber": obj.machine_number or "",
        "moldNumber": obj.mold_number or "",
        "productName": obj.product_name or "",
        "productNameEng": obj.product_name_eng or "",
        "color1": obj.color1 or "",
        "color2": obj.color2 or "",
        "unit": obj.unit or "",
        "quantity": int(obj.quantity or 0),
        "unitQuantity": int(obj.unit_quantity or 0),
        "total": int(obj.total or 0),
        "status": obj.status or "pending",
        "startTime": obj.start_time.isoformat() if obj.start_time else None,
        "endTime": obj.end_time.isoformat() if obj.end_time else None,
        "sortOrder": obj.sort_order or 0,
    }


def _zai_get_config():
    """
    OpenRouter 무료 모델 전용 설정.
    유료 모델(MiniMax 등) 기본값/환경변수는 enforce 로 차단.
    """
    try:
        from . import openrouter_service as ors

        cfg = ors.get_chat_config()
        if cfg:
            return cfg
    except Exception as e:
        logger.warning("openrouter_service get_chat_config failed: %s", e)

    # 최후 폴백 (키만 직접 읽음, 모델은 항상 free)
    base_url = (
        (
            os.getenv("OPENROUTER_BASE_URL")
            or os.getenv("ANTHROPIC_BASE_URL")
            or "https://openrouter.ai"
        )
        .strip()
        .rstrip("/")
    )
    api_key = (
        os.getenv("OPENROUTER_API_KEY") or os.getenv("ANTHROPIC_AUTH_TOKEN") or ""
    ).strip()
    if not base_url or not api_key:
        return None
    return {
        "base_url": base_url,
        "api_key": api_key,
        "model": "openrouter/free",
        "timeout_s": 30,
    }


def _zai_call_messages(
    *, system: str, user: str, max_tokens: int = 2048, temperature: float = 0.3
):
    """OpenRouter 무료 모델만 호출. 유료 모델 경로는 없음."""
    try:
        from . import openrouter_service as ors

        result = ors.chat_completions(
            system=system,
            user=user,
            max_tokens=max_tokens,
            temperature=temperature,
        )
        if result.get("success") and result.get("content"):
            content = str(result["content"]).strip()
            content = _extract_korean_only(content)
            return content or None
        if result.get("error"):
            logger.error(
                "AI free-model call failed (%s): %s",
                result.get("model"),
                result.get("error"),
            )
        return None
    except Exception as e:
        logger.error("AI call failed: %s", e)
        return None


class StandardResultsSetPagination(LimitOffsetPagination):
    default_limit = 10000
    max_limit = 20000


class OutboundRecordListView(generics.ListAPIView):
    queryset = OutboundRecord.objects.all().order_by("-outbound_date")
    serializer_class = OutboundRecordSerializer
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = super().get_queryset()
        start = self.request.query_params.get("start") or self.request.query_params.get(
            "startDate"
        )
        end = self.request.query_params.get("end") or self.request.query_params.get(
            "endDate"
        )

        if start:
            queryset = queryset.filter(outbound_date__gte=start)
        if end:
            queryset = queryset.filter(outbound_date__lte=end)

        return queryset

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        # Return only the list for frontend compatibility
        return Response(response.data["results"])


class InventoryItemViewSet(viewsets.ModelViewSet):
    queryset = InventoryItem.objects.all()
    serializer_class = InventoryItemSerializer


class DataSourceViewSet(viewsets.ModelViewSet):
    queryset = DataSource.objects.all()
    serializer_class = DataSourceSerializer


def _parse_int(val) -> int:
    if val is None:
        return 0
    s = str(val).strip().replace(",", "")
    if not s:
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _parse_date_ymd(val):
    s = ("" if val is None else str(val)).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    try:
        dt = pd.to_datetime(s, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.date()
    except Exception:
        return None


def _parse_datetime(val):
    if val is None:
        return None
    try:
        if isinstance(val, datetime):
            return val
    except Exception:
        pass
    s = str(val).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y.%m.%d %H:%M:%S",
        "%Y.%m.%d %H:%M",
    ):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        dt = pd.to_datetime(s, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.to_pydatetime()
    except Exception:
        return None


def _normalize_cols(cols):
    out = []
    for c in cols:
        s = " ".join(str(c).strip().lower().split())
        out.append(s)
    return out


def _find_col_index(cols, candidates):
    for cand in candidates:
        cand_norm = " ".join(str(cand).strip().lower().split())
        if not cand_norm:
            continue

        # Prefer exact match first (avoids picking '로케이션 유형' when '로케이션' exists)
        for i, c in enumerate(cols):
            if cand_norm == (c or ""):
                return i

        # Fallback to substring match
        for i, c in enumerate(cols):
            if cand_norm in (c or ""):
                return i
    return None


@api_view(["GET"])
def inventory_unified(request):
    # Source of truth: latest baseline upload (single active snapshot)
    latest_upload = InventoryBaselineUpload.objects.order_by("-uploaded_at").first()
    if not latest_upload:
        return Response(
            {
                "success": True,
                "data": [],
                "pagination": {
                    "page": 1,
                    "limit": 1000,
                    "total": 0,
                    "pages": 1,
                    "hasMore": False,
                },
                "summary": {"overall": {}, "filtered": {}, "options": {}},
                "lastUploadDate": None,
                "latestDataInfo": {
                    "latestUploadDate": None,
                    "totalItems": 0,
                    "filteredItems": 0,
                    "dataCompleteness": 100,
                    "hasLatestDataOnly": True,
                },
            }
        )

    as_of = latest_upload.as_of_date
    baseline_items = InventoryBaselineItem.objects.filter(upload=latest_upload)

    master_qs = BarcodeMaster.objects.all()
    master_map = {(m.barcode or "").strip(): m for m in master_qs if (m.barcode or "").strip()}

    # 제품 마스터(MasterSpec) — 단가 + 분류(category_lg)
    # 분류 표시 SoT: category_lg (마스터 페이지와 동일). 수량 계산과 무관.
    price_by_barcode = {}
    price_by_sku = {}
    price_by_name = {}
    category_lg_by_barcode = {}
    category_lg_by_name = {}
    for spec in MasterSpec.objects.all().only(
        "barcode",
        "sku_id",
        "product_name",
        "price",
        "prev_price",
        "price_changed_at",
        "category_lg",
    ):
        bc = (spec.barcode or "").strip()
        sku = (spec.sku_id or "").strip()
        name = (spec.product_name or "").strip()
        lg = (spec.category_lg or "").strip()
        if lg:
            if bc and bc not in category_lg_by_barcode:
                category_lg_by_barcode[bc] = lg
            if name and name not in category_lg_by_name:
                category_lg_by_name[name] = lg
        p = int(spec.price or 0)
        if p <= 0:
            continue
        if bc and bc not in price_by_barcode:
            price_by_barcode[bc] = p
        if sku and sku not in price_by_sku:
            price_by_sku[sku] = p
        if name and name not in price_by_name:
            price_by_name[name] = p

    # Fallback: 출고 분류 (마스터·BM 모두 없을 때만). Max(문자열) 금지 → 최빈값.
    baseline_barcodes_qs = baseline_items.exclude(barcode__isnull=True).exclude(barcode="").values("barcode")
    from collections import Counter, defaultdict

    outbound_category_map = {}
    barcode_cat_counts = defaultdict(Counter)
    for row in (
        OutboundRecord.objects.filter(barcode__in=baseline_barcodes_qs)
        .exclude(category__isnull=True)
        .exclude(category="")
        .values_list("barcode", "category")
    ):
        b = (row[0] or "").strip()
        c = (row[1] or "").strip()
        if b and c:
            barcode_cat_counts[b][c] += 1
    for b, counter in barcode_cat_counts.items():
        if counter:
            outbound_category_map[b] = counter.most_common(1)[0][0]

    # 현재고 가감: sales_api/inventory_stock.py 단일 규칙
    # (기준일 = 출고 후 잔여 스냅샷 → 당일 출고/입고 재차감 금지)
    from .inventory_stock import (
        aggregate_movements_after_baseline,
        compute_current_stock,
        filter_outbound_for_stock,
        stock_value,
    )

    baseline_barcode_list = list(
        baseline_items.exclude(barcode__isnull=True)
        .exclude(barcode="")
        .values_list("barcode", flat=True)
        .distinct()
    )
    # 바코드→상품명 (별칭 출고 통합 차감용)
    name_by_barcode = {}
    base_qty_by_barcode = {}
    for row in baseline_items.exclude(barcode__isnull=True).exclude(barcode="").values(
        "barcode", "product_name", "quantity_box"
    ):
        bc = (row["barcode"] or "").strip()
        if not bc:
            continue
        base_qty_by_barcode[bc] = base_qty_by_barcode.get(bc, 0) + int(
            row.get("quantity_box") or 0
        )
        if row.get("product_name") and bc not in name_by_barcode:
            name_by_barcode[bc] = (row["product_name"] or "").strip()
    # BarcodeMaster / MasterSpec 이름 보강
    extra_names = {}
    try:
        for bm in BarcodeMaster.objects.exclude(barcode="").only("barcode", "product_name"):
            b = (bm.barcode or "").strip()
            if b and (bm.product_name or "").strip():
                extra_names[b] = bm.product_name.strip()
                if b not in name_by_barcode:
                    name_by_barcode[b] = bm.product_name.strip()
    except Exception:
        pass
    try:
        for ms in MasterSpec.objects.exclude(barcode="").only("barcode", "product_name"):
            b = (ms.barcode or "").strip()
            if b and (ms.product_name or "").strip() and b not in name_by_barcode:
                name_by_barcode[b] = ms.product_name.strip()
                extra_names[b] = ms.product_name.strip()
    except Exception:
        pass

    outbound_agg, receipt_agg = aggregate_movements_after_baseline(
        as_of=as_of,
        barcodes=baseline_barcode_list,
        outbound_model=OutboundRecord,
        receipt_model=InventoryReceiptItem,
        name_by_barcode=name_by_barcode,
        extra_name_by_barcode=extra_names,
        base_qty_by_barcode=base_qty_by_barcode,
        apply_aliases=True,
    )

    # 30-day stats for threshold calc (min/max) — 실적 출고만
    end_date = timezone.localdate()
    start_30 = end_date - timedelta(days=29)
    outbound_30_qs = filter_outbound_for_stock(
        OutboundRecord.objects.filter(outbound_date__range=[start_30, end_date])
        .exclude(barcode__isnull=True)
        .exclude(barcode="")
        .filter(barcode__in=baseline_barcodes_qs)
    )
    outbound_30_agg = {
        row["barcode"]: int(row.get("qty") or 0)
        for row in outbound_30_qs.values("barcode").annotate(
            qty=Coalesce(Sum("box_quantity"), 0)
        )
    }

    # 60-day stats for cover days and 10-day target calculation
    start_60 = end_date - timedelta(days=59)
    outbound_60_qs = filter_outbound_for_stock(
        OutboundRecord.objects.filter(outbound_date__range=[start_60, end_date])
        .exclude(barcode__isnull=True)
        .exclude(barcode="")
        .filter(barcode__in=baseline_barcodes_qs)
    )
    outbound_60_agg = {
        row["barcode"]: int(row.get("qty") or 0)
        for row in outbound_60_qs.values("barcode").annotate(
            qty=Coalesce(Sum("box_quantity"), 0)
        )
    }

    # 14-day stats for cover days
    start_14 = end_date - timedelta(days=13)
    outbound_14_qs = filter_outbound_for_stock(
        OutboundRecord.objects.filter(outbound_date__range=[start_14, end_date])
        .exclude(barcode__isnull=True)
        .exclude(barcode="")
        .filter(barcode__in=baseline_barcodes_qs)
    )
    outbound_14_agg = {
        row["barcode"]: int(row.get("qty") or 0)
        for row in outbound_14_qs.values("barcode").annotate(
            qty=Coalesce(Sum("box_quantity"), 0)
        )
    }

    def _status_from_thresholds(
        current_qty: int, min_stock: int, safety_stock: int, max_stock: int
    ) -> str:
        # 우선순위: 위험(긴급발주) > 부족(발주요청) > 과잉 > 안전
        # - 위험: 품절(0 이하) 또는 최소재고 미만
        # - 부족: 최소재고 이상 ~ 안전재고 이하 (안전재고 > 최소재고 일 때만 구간 존재)
        #         안전재고 == 최소재고 이면 최소 도달 시 부족으로 표시
        # - 과잉: 최대재고 초과
        if current_qty <= 0:
            return "critical"
        if min_stock > 0 and current_qty < min_stock:
            return "critical"
        if safety_stock > 0 and current_qty <= safety_stock:
            return "low"
        if max_stock > 0 and current_qty > max_stock:
            return "high"
        return "normal"

    data = []
    for item in baseline_items:
        bc = (item.barcode or "").strip()
        master = master_map.get(bc)
        base_qty = int(item.quantity_box or 0)
        rcv_qty = int(receipt_agg.get(bc) or 0)
        out_qty = int(outbound_agg.get(bc) or 0)
        current_qty = compute_current_stock(base_qty, rcv_qty, out_qty)

        out14 = int(outbound_14_agg.get(bc) or 0)
        avg_daily = (out14 / 14.0) if out14 > 0 else 0.0
        cover_days = (current_qty / avg_daily) if avg_daily > 0 and current_qty > 0 else None

        out30 = int(outbound_30_agg.get(bc) or 0)
        avg_daily_30 = (out30 / 30.0) if out30 > 0 else 0.0
        
        out60 = int(outbound_60_agg.get(bc) or 0)
        avg_daily_60 = (out60 / 60.0) if out60 > 0 else 0.0

        # 정책: 최소재고=3일치, 최대재고=30일치(한달) — 30일 평균 출고 기준
        calc_min_stock = int(round(avg_daily_30 * 3)) if avg_daily_30 > 0 else 0
        calc_max_stock = int(round(avg_daily_30 * 30)) if avg_daily_30 > 0 else 0
        # 안전재고 기본: 7일치 (최소보다 여유 구간을 두어 부족 분류가 동작하도록)
        calc_safety_stock = int(round(avg_daily_30 * 7)) if avg_daily_30 > 0 else 0
        if calc_safety_stock > 0 and calc_min_stock > 0:
            calc_safety_stock = max(calc_safety_stock, calc_min_stock)

        # 임계값의 source of truth:
        # 1) BarcodeMaster에 설정값이 있으면 그것을 우선 사용
        # 2) 없으면(0) 계산값을 fallback
        bm_min_stock = int(getattr(master, "min_stock", 0) or 0) if master else 0
        bm_max_stock = int(getattr(master, "max_stock", 0) or 0) if master else 0
        bm_reorder_point = (
            int(getattr(master, "reorder_point", 0) or 0) if master else 0
        )
        bm_safety_stock = int(getattr(master, "safety_stock", 0) or 0) if master else 0
        bm_lifecycle_status = (
            (getattr(master, "lifecycle_status", None) or "active")
            if master
            else "active"
        )

        min_stock = bm_min_stock if bm_min_stock > 0 else calc_min_stock
        max_stock = bm_max_stock if bm_max_stock > 0 else calc_max_stock

        # safetyStock: 부족(발주요청) 상한
        # - 마스터 값이 있으면 사용 (최소재고 이상으로 보정)
        # - 없으면 계산 안전재고(7일), 그것도 없으면 최소재고
        if bm_safety_stock > 0:
            safety_stock = max(bm_safety_stock, min_stock)
        elif calc_safety_stock > 0:
            safety_stock = max(calc_safety_stock, min_stock)
        else:
            safety_stock = min_stock

        reorder_point = bm_reorder_point if bm_reorder_point > 0 else min_stock

        stock_status = _status_from_thresholds(
            int(current_qty), int(min_stock), int(safety_stock), int(max_stock)
        )

        hidden_reason = None
        if (
            str(bm_lifecycle_status) in ("paused", "discontinued")
            and int(current_qty) == 0
        ):
            hidden_reason = "lifecycle_zero_stock"

        product_name = (
            master.product_name
            if (master and master.product_name)
            else (item.product_name or "-")
        )
        sku_id = (master.sku_id if master else None) or ""
        # 단가: MasterSpec 매입가 (바코드 → SKU → 제품명)
        unit_price = (
            price_by_barcode.get(bc)
            or price_by_sku.get(str(sku_id).strip())
            or price_by_name.get(str(product_name or "").strip())
            or 0
        )

        data.append(
            {
                "id": str(item.id),
                "skuId": (master.sku_id if master else None),
                "productName": product_name,
                "currentStock": int(current_qty),
                "minStock": int(min_stock),
                "maxStock": int(max_stock),
                "reorderPoint": int(reorder_point),
                "safetyStock": int(safety_stock),
                "lifecycleStatus": str(bm_lifecycle_status),
                "hiddenReason": hidden_reason,
                # 분류 표시: MasterSpec.category_lg(마스터) → BM → 출고 최빈값 → 기타
                # 수량 계산(base/receipt/outbound)과는 완전 분리
                "category": (
                    category_lg_by_barcode.get(bc)
                    or category_lg_by_name.get(str(product_name or "").strip())
                    or (master.category if (master and master.category) else "")
                    or outbound_category_map.get(bc)
                    or "기타"
                ),
                "location": (
                    master.location
                    if (master and master.location)
                    else (item.location or "")
                ),
                "barcode": bc,
                "price": int(unit_price),
                "baseStock": int(base_qty),
                "receiptQty": int(rcv_qty),
                "outboundQty": int(out_qty),
                "lastUpdated": latest_upload.uploaded_at.isoformat()
                if latest_upload.uploaded_at
                else None,
                "inventoryDate": as_of.isoformat(),
                "stockStatus": stock_status,
                "coverDays": cover_days,
                "outbound14dTotal": out14,
                "avgDailyOutbound14d": avg_daily,
                "outbound30dTotal": out30,
                "avgDailyOutbound30d": avg_daily_30,
                "avgDailyOutbound60d": avg_daily_60,
            }
        )

    # 화면 집계 요약 (필터 전 전체 기준)
    summary_critical = 0
    summary_low = 0
    summary_normal = 0
    summary_high = 0
    summary_critical_qty = 0
    summary_low_qty = 0
    summary_total_value = 0
    summary_total_qty = 0
    for row in data:
        st = row.get("stockStatus") or "normal"
        qty = int(row.get("currentStock") or 0)
        price = int(row.get("price") or 0)
        val = stock_value(qty, price)
        qty_pos = max(0, qty)
        if st == "critical":
            summary_critical += 1
            summary_critical_qty += qty_pos
        elif st == "low":
            summary_low += 1
            summary_low_qty += qty_pos
        elif st == "high":
            summary_high += 1
        else:
            summary_normal += 1
        summary_total_qty += qty_pos
        summary_total_value += val

    return Response(
        {
            "success": True,
            "data": data,
            "pagination": {
                "page": 1,
                "limit": 1000,
                "total": len(data),
                "pages": 1,
                "hasMore": False,
            },
            "summary": {
                "overall": {
                    "critical": summary_critical,
                    "low": summary_low,
                    "normal": summary_normal,
                    "high": summary_high,
                    "criticalQuantity": summary_critical_qty,
                    "lowQuantity": summary_low_qty,
                    "totalValue": summary_total_value,
                    "totalQuantity": summary_total_qty,
                    "totalItems": len(data),
                },
                "filtered": {},
                "options": {},
            },
            "lastUploadDate": latest_upload.uploaded_at.isoformat()
            if latest_upload.uploaded_at
            else None,
            "latestDataInfo": {
                "latestUploadDate": latest_upload.uploaded_at.isoformat()
                if latest_upload.uploaded_at
                else None,
                "totalItems": len(data),
                "filteredItems": len(data),
                "dataCompleteness": 100,
                "hasLatestDataOnly": True,
            },
        }
    )


@api_view(["GET"])
def inventory_stock_diagnostics(request):
    """
    재고 금액/수량 구성 진단 (읽기 전용).
    오전·오후 숫자 차이 원인 확인용:
      현재고 = baseline + 입고(as_of 이후) − 실출고(as_of 이후)
    """
    from .inventory_stock import (
        aggregate_movements_after_baseline,
        compute_current_stock,
        stock_value,
    )

    latest = InventoryBaselineUpload.objects.order_by("-uploaded_at").first()
    if not latest:
        return Response(
            {"success": False, "message": "재고 스냅샷 없음"},
            status=status.HTTP_404_NOT_FOUND,
        )

    as_of = latest.as_of_date
    barcode = (request.query_params.get("barcode") or "").strip()
    items_qs = InventoryBaselineItem.objects.filter(upload=latest)
    if barcode:
        items_qs = items_qs.filter(barcode=barcode)

    items = list(items_qs.only("barcode", "quantity_box", "product_name"))
    bcs = [(i.barcode or "").strip() for i in items if (i.barcode or "").strip()]
    base_sum = sum(int(i.quantity_box or 0) for i in items)

    out_agg, rcv_agg = aggregate_movements_after_baseline(
        as_of=as_of,
        barcodes=bcs,
        outbound_model=OutboundRecord,
        receipt_model=InventoryReceiptItem,
    )
    out_sum = int(sum(out_agg.values()))
    rcv_sum = int(sum(rcv_agg.values()))

    price_by_bc = {}
    for s in MasterSpec.objects.exclude(barcode="").only("barcode", "price"):
        bc = (s.barcode or "").strip()
        p = int(s.price or 0)
        if bc and p > 0 and bc not in price_by_bc:
            price_by_bc[bc] = p

    total_qty = 0
    total_val = 0
    for it in items:
        bc = (it.barcode or "").strip()
        cur = compute_current_stock(
            it.quantity_box, rcv_agg.get(bc, 0), out_agg.get(bc, 0)
        )
        total_qty += max(0, cur)
        total_val += stock_value(cur, price_by_bc.get(bc, 0))

    out_daily = list(
        OutboundRecord.objects.filter(is_estimated=False, outbound_date__gt=as_of)
        .values("outbound_date")
        .annotate(qty=Coalesce(Sum("box_quantity"), 0), rows=Count("id"))
        .order_by("outbound_date")
    )
    if barcode:
        out_daily = list(
            OutboundRecord.objects.filter(
                is_estimated=False, outbound_date__gt=as_of, barcode=barcode
            )
            .values("outbound_date")
            .annotate(qty=Coalesce(Sum("box_quantity"), 0), rows=Count("id"))
            .order_by("outbound_date")
        )
    rcv_daily = list(
        InventoryReceiptItem.objects.filter(receipt_date__gt=as_of)
        .values("receipt_date")
        .annotate(qty=Coalesce(Sum("quantity_box"), 0), rows=Count("id"))
        .order_by("receipt_date")
    )
    if barcode:
        rcv_daily = list(
            InventoryReceiptItem.objects.filter(receipt_date__gt=as_of, barcode=barcode)
            .values("receipt_date")
            .annotate(qty=Coalesce(Sum("quantity_box"), 0), rows=Count("id"))
            .order_by("receipt_date")
        )

    return Response(
        {
            "success": True,
            "asOfDate": as_of.isoformat() if as_of else None,
            "baselineUploadedAt": latest.uploaded_at.isoformat()
            if latest.uploaded_at
            else None,
            "today": timezone.localdate().isoformat(),
            "formula": "current = baseline + receipts(after as_of) - real_outbound(after as_of)",
            "baseline": {"items": len(items), "qtySum": base_sum},
            "receiptsAfter": {"qtySum": rcv_sum, "daily": [
                {
                    "date": r["receipt_date"].isoformat(),
                    "qty": int(r["qty"] or 0),
                    "rows": int(r["rows"] or 0),
                }
                for r in rcv_daily
            ]},
            "outboundAfterReal": {"qtySum": out_sum, "daily": [
                {
                    "date": r["outbound_date"].isoformat(),
                    "qty": int(r["qty"] or 0),
                    "rows": int(r["rows"] or 0),
                }
                for r in out_daily
            ]},
            "currentStock": {
                "qty": total_qty,
                "value": total_val,
                "rawFormulaQty": base_sum + rcv_sum - out_sum,
            },
            "barcode": barcode or None,
        }
    )


@api_view(["GET", "DELETE"])
def inventory_upload_history(request):
    if request.method == "GET":
        uploads = InventoryBaselineUpload.objects.order_by("-uploaded_at")
        data = []
        for u in uploads:
            data.append(
                {
                    "id": str(u.id),
                    "fileName": ", ".join((u.file_names or [])[:5])
                    if isinstance(u.file_names, list)
                    else "",
                    "uploadDate": u.uploaded_at.isoformat() if u.uploaded_at else None,
                    "inventoryDate": u.as_of_date.isoformat() if u.as_of_date else None,
                    "status": "success",
                    "recordsProcessed": int(u.total_barcodes or 0),
                    "recordsSkipped": 0,
                    "uploadedBy": "system",
                    "fileSize": 0,
                }
            )
        return Response({"success": True, "data": data})

    # DELETE = 전체 리셋 (baseline + receipts). UI "전체 삭제" 경로.
    # DESTRUCTIVE_API_KEY 가 설정된 환경에서만 추가 키 요구 (미설정 시 프론트 호환).
    from .api_guards import destructive_guard_response

    blocked = destructive_guard_response(request)
    if blocked is not None:
        return blocked

    logger.warning(
        "inventory_upload_history FULL RESET requested path=%s",
        request.path,
    )
    with transaction.atomic():
        bi = InventoryBaselineItem.objects.all().delete()
        bu = InventoryBaselineUpload.objects.all().delete()
        ri = InventoryReceiptItem.objects.all().delete()
        ru = InventoryReceiptUpload.objects.all().delete()
    return Response(
        {
            "success": True,
            "message": "reset ok",
            "deleted": {
                "baselineItems": bi[0] if bi else 0,
                "baselineUploads": bu[0] if bu else 0,
                "receiptItems": ri[0] if ri else 0,
                "receiptUploads": ru[0] if ru else 0,
            },
        }
    )


@api_view(["DELETE"])
def inventory_upload_history_by_date(request, date: str):
    """
    특정 기준일(as_of_date)의 baseline 업로드만 삭제.
    입고(receipt)는 건드리지 않음 — 과거에는 전체 리셋이어서 위험했음.
    """
    as_of = _parse_date_ymd(date)
    if not as_of:
        return Response(
            {"success": False, "message": "invalid date (YYYY-MM-DD)"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    uploads = InventoryBaselineUpload.objects.filter(as_of_date=as_of)
    count = uploads.count()
    if count == 0:
        return Response(
            {
                "success": True,
                "message": "no uploads for date",
                "asOfDate": as_of.isoformat(),
                "deletedUploads": 0,
            }
        )

    # FK CASCADE로 items 함께 삭제
    with transaction.atomic():
        deleted, detail = uploads.delete()
    logger.info(
        "inventory_upload_history_by_date as_of=%s deleted_rows=%s detail=%s",
        as_of.isoformat(),
        deleted,
        detail,
    )
    return Response(
        {
            "success": True,
            "message": "deleted baseline uploads for date",
            "asOfDate": as_of.isoformat(),
            "deletedUploads": count,
            "deletedRows": deleted,
        }
    )


@api_view(["POST"])
def inventory_baseline_upload(request):
    """
    재고 기준 스냅샷 업로드.

    안전 규칙 (실서비스):
    1) 엑셀 파싱·검증이 끝난 뒤에만 DB 변경
    2) 트랜잭션으로 baseline 교체 (중간 실패 시 롤백)
    3) 입고(receipt) 이력은 삭제하지 않음 — 현재고 공식은 as_of 이후 입고만 가산
    """
    error_id = str(uuid.uuid4())
    as_of_str = (
        (
            request.data.get("inventoryDate") or request.data.get("asOfDate") or ""
        ).strip()
        if isinstance(request.data, dict)
        else ""
    )
    as_of = _parse_date_ymd(as_of_str)
    if not as_of:
        return Response(
            {"message": "inventoryDate is required (YYYY-MM-DD)"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    files = request.FILES.getlist("files")
    if not files:
        files = request.FILES.getlist("xlsx")
    if not files:
        # backward compatibility with existing UI
        files = request.FILES.getlist("csv")
    if not files:
        return Response(
            {"message": "files are required"}, status=status.HTTP_400_BAD_REQUEST
        )

    total_bytes = 0
    try:
        for f in files:
            total_bytes += int(getattr(f, "size", 0) or 0)
    except (ValueError, TypeError, AttributeError):
        total_bytes = 0

    logger.info(
        "baseline_upload start error_id=%s as_of=%s file_count=%s total_mb=%.2f",
        error_id,
        as_of.isoformat(),
        len(files),
        (total_bytes / 1024 / 1024) if total_bytes else 0.0,
    )
    logger.debug("baseline_upload files=%s", [getattr(f, "name", "") for f in files])

    # --- 1) 파싱만 수행 (DB 변경 없음). 실패 시 기존 스냅샷 유지 ---
    total_rows = 0
    agg = {}
    meta = {}
    file_names = [getattr(f, "name", "") for f in files]

    for f in files:
        try:
            logger.debug(
                "baseline_upload read_excel start error_id=%s file=%s size=%s",
                error_id,
                getattr(f, "name", ""),
                getattr(f, "size", None),
            )
            df = pd.read_excel(f, dtype=str)
            logger.debug(
                "baseline_upload read_excel done error_id=%s file=%s rows=%s cols=%s",
                error_id,
                getattr(f, "name", ""),
                len(df),
                len(df.columns),
            )
        except Exception as e:
            logger.exception(
                "baseline_upload read_excel failed error_id=%s file=%s",
                error_id,
                getattr(f, "name", ""),
            )
            return Response(
                {
                    "message": f"엑셀 파싱 실패: {getattr(f, 'name', '')}: {str(e)}",
                    "errorId": error_id,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        df = df.fillna("")
        cols = _normalize_cols(df.columns)
        bc_idx = _find_col_index(
            cols, ["상품 바코드", "상품바코드", "바코드", "barcode"]
        )
        qty_idx = _find_col_index(cols, ["수량"])
        name_idx = _find_col_index(cols, ["상품명", "품목", "product"])
        # Prefer external SKU id when available
        sku_idx = _find_col_index(
            cols,
            [
                "외부 sku id",
                "external sku id",
                "external skuid",
                "ext sku id",
                "sku id",
                "sku_id",
                "sku-id",
                "skuid",
                "sku",
                "품목코드",
                "상품코드",
                "상품 코드",
                "상품id",
                "상품 id",
            ],
        )
        cat_idx = _find_col_index(
            cols, ["분류", "카테고리", "category", "제품분류", "제품 분류"]
        )
        loc_idx = _find_col_index(
            cols, ["로케이션", "location", "위치", "보관", "창고", "적치"]
        )

        if bc_idx is None or qty_idx is None:
            logger.warning(
                "baseline_upload missing required cols error_id=%s file=%s cols=%s",
                error_id,
                getattr(f, "name", ""),
                cols,
            )
            continue

        total_rows += len(df)
        for _, row in df.iterrows():
            bc = str(row.iloc[bc_idx]).strip()
            if not bc:
                continue
            qty = _parse_int(row.iloc[qty_idx])
            if qty <= 0:
                continue
            agg[bc] = int(agg.get(bc) or 0) + qty
            if bc not in meta:
                meta[bc] = {
                    "product_name": str(row.iloc[name_idx]).strip()
                    if name_idx is not None
                    else "",
                    "sku_id": str(row.iloc[sku_idx]).strip()
                    if sku_idx is not None
                    else "",
                    "category": str(row.iloc[cat_idx]).strip()
                    if cat_idx is not None
                    else "",
                    "location": str(row.iloc[loc_idx]).strip()
                    if loc_idx is not None
                    else "",
                }

    if not agg:
        return Response(
            {
                "message": "유효한 바코드/수량 행이 없습니다. 기존 재고 스냅샷은 유지됩니다.",
                "errorId": error_id,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    # --- 2) 파싱 성공 후에만 baseline 교체 (입고 이력 보존) ---
    try:
        with transaction.atomic():
            InventoryBaselineItem.objects.all().delete()
            InventoryBaselineUpload.objects.all().delete()
            # 주의: InventoryReceipt* 는 삭제하지 않음

            upload = InventoryBaselineUpload.objects.create(
                as_of_date=as_of,
                file_count=len(files),
                file_names=file_names,
                total_rows=int(total_rows),
                total_barcodes=int(len(agg)),
            )

            items = [
                InventoryBaselineItem(
                    upload=upload,
                    barcode=bc,
                    quantity_box=int(qty),
                    product_name=((meta.get(bc) or {}).get("product_name") or "")[:255],
                    location=((meta.get(bc) or {}).get("location") or "")[:255],
                )
                for bc, qty in agg.items()
            ]
            InventoryBaselineItem.objects.bulk_create(items, batch_size=2000)
    except Exception as e:
        logger.exception(
            "baseline_upload db write failed error_id=%s", error_id
        )
        return Response(
            {
                "message": f"재고 스냅샷 저장 실패: {str(e)}",
                "errorId": error_id,
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    # Upsert BarcodeMaster (실패해도 스냅샷은 이미 커밋 — 부가 정보만)
    try:
        barcodes = list(agg.keys())
        existing_map = {
            m.barcode: m for m in BarcodeMaster.objects.filter(barcode__in=barcodes)
        }
        to_create = []
        to_update = []

        for bc in barcodes:
            m = meta.get(bc) or {}
            sku_id_val = (m.get("sku_id") or "").strip()
            cat_val = (m.get("category") or "").strip()
            loc_val = (m.get("location") or "").strip()
            name_val = (m.get("product_name") or "").strip()

            bm = existing_map.get(bc)
            if not bm:
                to_create.append(
                    BarcodeMaster(
                        barcode=bc,
                        sku_id=sku_id_val or "",
                        category=cat_val or "",
                        location=loc_val or "",
                        product_name=name_val[:255] if name_val else "",
                    )
                )
                continue

            changed = False
            if sku_id_val and (bm.sku_id or "").strip() != sku_id_val:
                bm.sku_id = sku_id_val
                changed = True
            if cat_val and (bm.category or "").strip() != cat_val:
                bm.category = cat_val
                changed = True
            if loc_val and (bm.location or "").strip() != loc_val:
                bm.location = loc_val
                changed = True
            if name_val and (bm.product_name or "").strip() != name_val:
                bm.product_name = name_val[:255]
                changed = True

            if changed:
                to_update.append(bm)

        if to_create:
            BarcodeMaster.objects.bulk_create(
                to_create, ignore_conflicts=True, batch_size=2000
            )
        if to_update:
            BarcodeMaster.objects.bulk_update(
                to_update,
                ["sku_id", "category", "location", "product_name"],
                batch_size=2000,
            )
    except Exception:
        logger.exception(
            "baseline_upload barcode_master upsert failed error_id=%s", error_id
        )

    logger.info(
        "baseline_upload done error_id=%s as_of=%s total_rows=%s total_barcodes=%s receipts_preserved=1",
        error_id,
        as_of.isoformat(),
        int(total_rows),
        int(len(agg)),
    )

    return Response(
        {
            "success": True,
            "message": "baseline uploaded",
            "rowsProcessed": len(agg),
            "asOfDate": as_of.isoformat(),
            "receiptsPreserved": True,
        }
    )


@api_view(["POST"])
def inventory_receipts_upload(request):
    error_id = str(uuid.uuid4())
    latest_upload = InventoryBaselineUpload.objects.order_by("-uploaded_at").first()
    if not latest_upload:
        return Response(
            {"message": "기준재고 업로드 후 입고 업로드가 가능합니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    as_of = latest_upload.as_of_date
    file_obj = request.FILES.get("file") or request.FILES.get("xlsx")
    if not file_obj:
        return Response(
            {"message": "file is required"}, status=status.HTTP_400_BAD_REQUEST
        )

    logger.info(
        "receipts_upload start error_id=%s baseline_as_of=%s file=%s size=%s",
        error_id,
        as_of.isoformat() if as_of else None,
        getattr(file_obj, "name", ""),
        getattr(file_obj, "size", None),
    )

    raw = file_obj.read()
    file_hash = hashlib.sha256(raw).hexdigest()
    if InventoryReceiptUpload.objects.filter(file_hash=file_hash).exists():
        return Response({"success": True, "message": "이미 업로드된 파일입니다."})

    try:
        df = pd.read_excel(io.BytesIO(raw), dtype=str).fillna("")
    except Exception as e:
        logger.exception(
            "receipts_upload read_excel failed error_id=%s file=%s",
            error_id,
            getattr(file_obj, "name", ""),
        )
        return Response(
            {
                "message": f"엑셀 파싱 실패: {str(e)}",
                "errorId": error_id,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    cols = _normalize_cols(df.columns)
    bc_idx = _find_col_index(cols, ["상품바코드", "상품 바코드", "바코드", "barcode"])
    qty_idx = _find_col_index(cols, ["입고 수량", "입고수량", "수량"])
    dt_idx = _find_col_index(
        cols, ["입고 일시", "입고일시", "입고일", "datetime", "date"]
    )
    name_idx = _find_col_index(cols, ["상품명", "품목", "product"])

    if bc_idx is None or qty_idx is None or dt_idx is None:
        logger.warning(
            "receipts_upload missing required cols error_id=%s cols=%s", error_id, cols
        )
        return Response(
            {
                "message": "필수 컬럼(상품바코드/입고 수량/입고 일시)을 찾을 수 없습니다.",
                "errorId": error_id,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    upload = InventoryReceiptUpload.objects.create(
        file_name=getattr(file_obj, "name", "") or "receipts.xlsx",
        file_hash=file_hash,
    )

    rows_processed = 0
    rows_skipped = 0
    rows_created = 0
    rows_updated = 0
    rows_unchanged = 0
    rows_invalid = 0

    parsed = []
    for _, row in df.iterrows():
        bc = str(row.iloc[bc_idx]).strip()
        if not bc:
            rows_invalid += 1
            continue
        qty = _parse_int(row.iloc[qty_idx])
        if qty <= 0:
            rows_invalid += 1
            continue
        dt = _parse_datetime(row.iloc[dt_idx])
        if not dt:
            rows_invalid += 1
            continue
        rdate = dt.date()
        # 기준일 = 출고 후 잔여 스냅샷 → 기준일 당일·이전 입고는 스냅샷에 포함되었다고 보고 제외
        from .inventory_stock import is_movement_date_applicable

        if not is_movement_date_applicable(rdate, as_of):
            rows_skipped += 1
            continue
        parsed.append(
            {
                "barcode": bc,
                "receipt_datetime": dt,
                "receipt_date": rdate,
                "quantity_box": int(qty),
                "product_name": (
                    str(row.iloc[name_idx]).strip()[:255]
                    if name_idx is not None
                    else ""
                ),
            }
        )

    # Auto-register new products into catalog (BarcodeMaster)
    try:
        parsed_barcodes = sorted({p.get("barcode") for p in parsed if p.get("barcode")})
        if parsed_barcodes:
            existing = set(
                BarcodeMaster.objects.filter(barcode__in=parsed_barcodes).values_list(
                    "barcode", flat=True
                )
            )
            to_create = []
            # pick first non-empty name per barcode
            name_map = {}
            for p in parsed:
                bc = p.get("barcode")
                if not bc or bc in name_map:
                    continue
                nm = (p.get("product_name") or "").strip()
                if nm:
                    name_map[bc] = nm[:255]
            for bc in parsed_barcodes:
                if bc in existing:
                    continue
                to_create.append(
                    BarcodeMaster(
                        barcode=bc,
                        product_name=(name_map.get(bc) or ""),
                    )
                )
            if to_create:
                BarcodeMaster.objects.bulk_create(
                    to_create, ignore_conflicts=True, batch_size=2000
                )
    except Exception:
        logger.exception(
            "receipts_upload barcode_master auto-register failed error_id=%s", error_id
        )

    # Upsert by (barcode, receipt_datetime)
    # - same qty: unchanged (skip)
    # - different qty: update to new qty (delta semantics)
    # - missing: create
    keys = {(p["barcode"], p["receipt_datetime"]) for p in parsed}
    existing_map = {}
    if keys:
        # SQLite doesn't support composite IN well; fetch by barcode and then filter in memory.
        barcodes = sorted({bc for bc, _ in keys})
        qs = InventoryReceiptItem.objects.filter(barcode__in=barcodes)
        for obj in qs:
            k = ((obj.barcode or "").strip(), obj.receipt_datetime)
            if k in keys:
                existing_map[k] = obj

    to_create = []
    to_update = []

    for p in parsed:
        k = (p["barcode"], p["receipt_datetime"])
        existing = existing_map.get(k)
        if not existing:
            to_create.append(
                InventoryReceiptItem(
                    upload=upload,
                    receipt_datetime=p["receipt_datetime"],
                    receipt_date=p["receipt_date"],
                    barcode=p["barcode"],
                    quantity_box=p["quantity_box"],
                    product_name=p["product_name"],
                )
            )
            rows_created += 1
            rows_processed += 1
            continue

        new_qty = int(p["quantity_box"])
        old_qty = int(existing.quantity_box or 0)
        if new_qty == old_qty:
            rows_unchanged += 1
            rows_processed += 1
            continue

        existing.quantity_box = new_qty
        # Keep most recent name if provided
        if p.get("product_name"):
            existing.product_name = p["product_name"]
        # Tie latest upload reference to most recent file
        existing.upload = upload
        existing.receipt_date = p["receipt_date"]
        to_update.append(existing)
        rows_updated += 1
        rows_processed += 1

    try:
        with transaction.atomic():
            if to_create:
                InventoryReceiptItem.objects.bulk_create(
                    to_create, batch_size=2000, ignore_conflicts=True
                )
            if to_update:
                InventoryReceiptItem.objects.bulk_update(
                    to_update,
                    ["quantity_box", "product_name", "upload", "receipt_date"],
                    batch_size=2000,
                )
    except Exception:
        logger.exception("receipts_upload upsert failed error_id=%s", error_id)
        return Response(
            {
                "message": "입고 업로드 처리 중 오류가 발생했습니다.",
                "errorId": error_id,
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    rows_skipped += int(rows_invalid)

    upload.rows_processed = int(rows_processed)
    upload.rows_skipped = int(rows_skipped)
    upload.save(update_fields=["rows_processed", "rows_skipped"])

    logger.info(
        "receipts_upload done error_id=%s rows_processed=%s rows_skipped=%s",
        error_id,
        int(rows_processed),
        int(rows_skipped),
    )

    return Response(
        {
            "success": True,
            "message": "receipts uploaded",
            "rowsProcessed": rows_processed,
            "rowsSkipped": rows_skipped,
            "rowsCreated": rows_created,
            "rowsUpdated": rows_updated,
            "rowsUnchanged": rows_unchanged,
        }
    )


@api_view(["GET"])
def inventory_unified_download_csv(request):
    # inventory_unified is decorated with @api_view, so it expects a Django HttpRequest.
    # Here we are inside DRF already, so we must pass the underlying HttpRequest.
    raw_request = getattr(request, "_request", request)
    resp = inventory_unified(raw_request)
    try:
        payload = resp.data if hasattr(resp, "data") else None
    except (AttributeError, TypeError):
        payload = None

    rows = []
    if isinstance(payload, dict):
        rows = payload.get("data") or []

    headers = [
        "barcode",
        "skuId",
        "productName",
        "category",
        "location",
        "inventoryDate",
        "currentStock",
        "minStock",
        "safetyStock",
        "reorderPoint",
        "maxStock",
        "stockStatus",
        "coverDays",
        "outbound14dTotal",
        "avgDailyOutbound14d",
        "outbound30dTotal",
        "avgDailyOutbound30d",
        "lifecycleStatus",
    ]

    def _fmt2(v):
        if v is None or v == "":
            return ""
        try:
            return f"{float(v):.2f}"
        except Exception:
            return str(v)

    encoding = (request.query_params.get("encoding") or "").strip().lower()
    if encoding not in ("utf-8", "utf-8-sig", "cp949"):
        encoding = "utf-8-sig"

    out = io.StringIO(newline="")
    writer = csv.writer(out, lineterminator="\n")
    writer.writerow(headers)
    for r in rows:
        if not isinstance(r, dict):
            continue
        writer.writerow(
            [
                (r.get("barcode") or ""),
                (r.get("skuId") or ""),
                (r.get("productName") or ""),
                (r.get("category") or ""),
                (r.get("location") or ""),
                (r.get("inventoryDate") or ""),
                (r.get("currentStock") or 0),
                (r.get("minStock") or 0),
                (r.get("safetyStock") or 0),
                (r.get("reorderPoint") or 0),
                (r.get("maxStock") or 0),
                (r.get("stockStatus") or ""),
                (_fmt2(r.get("coverDays"))),
                (r.get("outbound14dTotal") or 0),
                (_fmt2(r.get("avgDailyOutbound14d"))),
                (r.get("outbound30dTotal") or 0),
                (_fmt2(r.get("avgDailyOutbound30d"))),
                (r.get("lifecycleStatus") or ""),
            ]
        )

    content = out.getvalue()
    out.close()

    filename = f"inventory_unified_{timezone.localdate().isoformat()}.csv"
    data = content.encode(encoding, errors="replace")
    charset = "cp949" if encoding == "cp949" else "utf-8"
    response = HttpResponse(data, content_type=f"text/csv; charset={charset}")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@api_view(["POST"])
def inventory_barcode_location_upsert(request):
    """
    스캐너 수동 로케이션 영구 저장 → BarcodeMaster.
    body: barcode, location, product_name?, sku_id?
    """
    payload = request.data if isinstance(request.data, dict) else {}
    barcode = (payload.get("barcode") or "").strip()
    location = (payload.get("location") or "").strip()
    product_name = (payload.get("product_name") or "").strip()
    sku_id = (payload.get("sku_id") or "").strip()
    if not barcode:
        return Response({"message": "barcode required"}, status=status.HTTP_400_BAD_REQUEST)
    if not location:
        return Response({"message": "location required"}, status=status.HTTP_400_BAD_REQUEST)

    ok, msg = _upsert_barcode_location(
        barcode,
        location=location,
        product_name=product_name,
        sku_id=sku_id,
        clear_if_empty=False,
    )
    if not ok:
        return Response({"message": msg}, status=status.HTTP_400_BAD_REQUEST)

    bm = BarcodeMaster.objects.filter(barcode=barcode).first()
    # MasterSpec 대분류 상태 (질문 모달 여부 판단용)
    spec = MasterSpec.objects.filter(barcode=barcode).first()
    if not spec and product_name:
        spec = MasterSpec.objects.filter(product_name=product_name).first()
    needs_category = True
    category_lg = ""
    if spec:
        category_lg = (spec.category_lg or "").strip()
        needs_category = not category_lg

    return Response(
        {
            "success": True,
            "barcode": barcode,
            "location": (bm.location if bm else location) or location,
            "product_name": (bm.product_name if bm else product_name) or product_name,
            "sku_id": (bm.sku_id if bm else sku_id) or sku_id,
            "master_spec_id": spec.id if spec else None,
            "category_lg": category_lg,
            "needs_category": needs_category,
        }
    )


@api_view(["GET"])
def master_category_lg_options(request):
    """대분류 선택 목록 (신규 품목 질문용)."""
    cats = list(
        MasterSpec.objects.exclude(category_lg="")
        .exclude(category_lg__isnull=True)
        .values_list("category_lg", flat=True)
        .distinct()
        .order_by("category_lg")
    )
    return Response({"success": True, "data": cats})


@api_view(["POST"])
def master_specs_register_from_scan(request):
    """
    스캐너 신규/대분류 공란 품목 등록.
    category_lg 필수 (사용자 질문 결과).
    """
    payload = request.data if isinstance(request.data, dict) else {}
    barcode = (payload.get("barcode") or "").strip()
    product_name = (payload.get("product_name") or "").strip()
    category_lg = (payload.get("category_lg") or "").strip()
    category_md = (payload.get("category_md") or "").strip()
    location = (payload.get("location") or "").strip()
    sku_id = (payload.get("sku_id") or "").strip()
    is_vf = bool(payload.get("is_vf_item") or False)

    if not category_lg:
        return Response(
            {"message": "대분류(category_lg)는 필수입니다. 사용자에게 질문 후 저장하세요."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if not product_name and not barcode:
        return Response(
            {"message": "product_name 또는 barcode 필요"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    spec = None
    if barcode:
        spec = MasterSpec.objects.filter(barcode=barcode).first()
    if not spec and product_name:
        spec = MasterSpec.objects.filter(product_name=product_name).first()

    created = False
    if not spec:
        if not product_name:
            product_name = f"미등록 {barcode}" if barcode else "미등록 품목"
        # unique product_name
        base_name = product_name
        n = 1
        while MasterSpec.objects.filter(product_name=product_name).exists():
            n += 1
            product_name = f"{base_name} ({n})"
        spec = MasterSpec.objects.create(
            product_name=product_name,
            barcode=barcode,
            sku_id=sku_id,
            category_lg=category_lg,
            category_md=category_md,
            is_vf_item=is_vf,
        )
        created = True
    else:
        spec.category_lg = category_lg
        if category_md:
            spec.category_md = category_md
        if barcode and not (spec.barcode or "").strip():
            spec.barcode = barcode
        if sku_id and not (spec.sku_id or "").strip():
            spec.sku_id = sku_id
        if product_name and not (spec.product_name or "").strip():
            spec.product_name = product_name
        if "is_vf_item" in payload:
            spec.is_vf_item = is_vf
        spec.save()

    if barcode and location:
        _upsert_barcode_location(
            barcode,
            location=location,
            product_name=spec.product_name or product_name,
            sku_id=spec.sku_id or sku_id,
        )
    elif barcode:
        _upsert_barcode_location(
            barcode,
            location=None,
            product_name=spec.product_name or product_name,
            sku_id=spec.sku_id or sku_id,
        )

    loc = ""
    if barcode:
        loc = _barcode_location_map([barcode]).get(barcode, "") or location

    return Response(
        {
            "success": True,
            "created": created,
            "spec": _master_spec_dict(spec, location=loc),
        },
        status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
    )


@api_view(["GET"])
def inventory_barcode_master(request):
    q = (
        request.query_params.get("q") or request.query_params.get("search") or ""
    ).strip()
    limit_raw = request.query_params.get("limit")
    try:
        limit = int(limit_raw) if limit_raw else 1000
    except Exception:
        limit = 1000
    limit = max(1, min(limit, 5000))

    # Seed missing barcodes from latest baseline so SKU/threshold can be edited later
    latest_upload = InventoryBaselineUpload.objects.order_by("-uploaded_at").first()
    if latest_upload:
        items = (
            InventoryBaselineItem.objects.filter(upload=latest_upload)
            .exclude(barcode__isnull=True)
            .exclude(barcode="")
        )
        for row in items.values("barcode").annotate(
            product_name=Max("product_name"),
            location=Max("location"),
        )[:2000]:
            bc = (row.get("barcode") or "").strip()
            if not bc:
                continue
            BarcodeMaster.objects.get_or_create(
                barcode=bc,
                defaults={
                    "product_name": (row.get("product_name") or "")[:255],
                    "location": (row.get("location") or "")[:255],
                },
            )

    qs = BarcodeMaster.objects.all().order_by("barcode")
    if q:
        qs = qs.filter(
            models.Q(barcode__icontains=q)
            | models.Q(product_name__icontains=q)
            | models.Q(sku_id__icontains=q)
            | models.Q(category__icontains=q)
            | models.Q(location__icontains=q)
        )

    rows = []
    for bm in qs[:limit]:
        rows.append(
            {
                "id": str(bm.id),
                "barcode": bm.barcode,
                "skuId": bm.sku_id or "",
                "productName": bm.product_name or "",
                "category": bm.category or "",
                "location": bm.location or "",
                "lifecycleStatus": getattr(bm, "lifecycle_status", "active")
                or "active",
                "minStock": int(bm.min_stock or 0),
                "maxStock": int(bm.max_stock or 0),
                "reorderPoint": int(bm.reorder_point or 0),
                "safetyStock": int(bm.safety_stock or 0),
                "notes": bm.notes or "",
                "createdAt": bm.created_at.isoformat() if bm.created_at else None,
                "updatedAt": bm.updated_at.isoformat() if bm.updated_at else None,
            }
        )

    return Response({"success": True, "data": rows})


@api_view(["PATCH"])
def inventory_unified_patch(request, _id: str):
    payload = request.data if isinstance(request.data, dict) else {}

    try:
        item = InventoryBaselineItem.objects.filter(id=_id).first()
    except Exception:
        item = None

    barcode = (payload.get("barcode") or "").strip()
    if not barcode and item:
        barcode = (item.barcode or "").strip()
    if not barcode:
        try:
            bm = BarcodeMaster.objects.filter(id=_id).first()
        except Exception:
            bm = None
        if bm:
            barcode = (bm.barcode or "").strip()
    if not barcode:
        return Response(
            {"success": False, "message": "barcode is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def _to_int(v):
        try:
            if v is None or v == "":
                return None
            return int(float(v))
        except Exception:
            return None

    min_stock = _to_int(payload.get("minStock"))
    max_stock = _to_int(payload.get("maxStock"))
    reorder_point = _to_int(payload.get("reorderPoint"))
    safety_stock = _to_int(payload.get("safetyStock"))

    lifecycle_status = (
        payload.get("lifecycleStatus") or payload.get("lifecycle_status") or ""
    ).strip()

    sku_id = (payload.get("skuId") or payload.get("sku_id") or "").strip()
    category = (payload.get("category") or "").strip()
    location = (payload.get("location") or "").strip()
    product_name = (
        payload.get("productName") or payload.get("product_name") or ""
    ).strip()

    bm, _created = BarcodeMaster.objects.get_or_create(barcode=barcode)
    if sku_id:
        bm.sku_id = sku_id
    if category:
        bm.category = category
    if location:
        bm.location = location
    if product_name:
        bm.product_name = product_name

    if min_stock is not None:
        bm.min_stock = max(0, min_stock)
    if max_stock is not None:
        bm.max_stock = max(0, max_stock)
    if reorder_point is not None:
        bm.reorder_point = max(0, reorder_point)
    if safety_stock is not None:
        bm.safety_stock = max(0, safety_stock)

    if lifecycle_status:
        lifecycle_status = lifecycle_status.lower()
        if lifecycle_status in ("active", "paused", "discontinued"):
            bm.lifecycle_status = lifecycle_status

    bm.save()

    return Response({"success": True, "barcode": barcode})


def _barcode_location_map(barcodes=None) -> dict:
    """barcode → location (BarcodeMaster SoT). barcodes=None 이면 전체."""
    qs = BarcodeMaster.objects.exclude(barcode="").exclude(barcode__isnull=True)
    if barcodes is not None:
        cleaned = {str(b).strip() for b in barcodes if b and str(b).strip()}
        if not cleaned:
            return {}
        qs = qs.filter(barcode__in=cleaned)
    out = {}
    for row in qs.values_list("barcode", "location"):
        bc = (row[0] or "").strip()
        if bc and bc not in out:
            out[bc] = (row[1] or "").strip()
    return out


def _upsert_barcode_location(
    barcode: str,
    *,
    location: str | None = None,
    product_name: str = "",
    sku_id: str = "",
    clear_if_empty: bool = False,
) -> tuple[bool, str]:
    """
    BarcodeMaster 로케이션 등록/수정.
    Returns: (ok, message)
    - location is None → 로케이션 필드 변경 안 함 (메타만 보강 가능)
    - location == "" and clear_if_empty → 로케이션 비움
    """
    bc = (barcode or "").strip()
    if not bc:
        return False, "barcode_empty"

    bm, _created = BarcodeMaster.objects.get_or_create(
        barcode=bc,
        defaults={
            "product_name": (product_name or "")[:255],
            "sku_id": (sku_id or "")[:100],
            "location": (location or "")[:255] if location is not None else "",
        },
    )
    changed = False
    if product_name and not (bm.product_name or "").strip():
        bm.product_name = (product_name or "")[:255]
        changed = True
    if sku_id and not (bm.sku_id or "").strip():
        bm.sku_id = (sku_id or "")[:100]
        changed = True
    if location is not None:
        loc = (location or "").strip()
        if loc:
            if (bm.location or "").strip() != loc:
                bm.location = loc[:255]
                changed = True
        elif clear_if_empty:
            if (bm.location or "").strip():
                bm.location = ""
                changed = True
    if changed or _created:
        bm.save()
    return True, "ok" if not _created else "created"


def _master_spec_dict(s, location: str = "") -> dict:
    return {
        "id": s.id,
        "product_name": s.product_name,
        "product_name_eng": s.product_name_eng,
        "mold_number": s.mold_number,
        "color1": s.color1,
        "color2": s.color2,
        "default_quantity": int(s.default_quantity or 0),
        "sku_id": s.sku_id or "",
        "barcode": s.barcode or "",
        "location": location or "",
        "category_lg": s.category_lg or "",
        "category_md": s.category_md or "",
        "price": int(s.price or 0),
        "lot_number": s.lot_number or "",
        "components": s.components or "",
        "image_url": s.image_url or "",
        "prev_price": int(s.prev_price or 0),
        "price_changed_at": s.price_changed_at.isoformat() if s.price_changed_at else None,
        "is_discontinued": s.is_discontinued,
        "is_no_outbound_3m": s.is_no_outbound_3m,
        "is_vf_item": bool(getattr(s, "is_vf_item", False)),
    }


@api_view(["GET", "POST"])
def master_specs(request):
    if request.method == "GET":
        specs = list(MasterSpec.objects.all().order_by("product_name"))
        loc_map = _barcode_location_map(
            [(s.barcode or "").strip() for s in specs if (s.barcode or "").strip()]
        )
        return Response(
            [
                _master_spec_dict(
                    s, location=loc_map.get((s.barcode or "").strip(), "")
                )
                for s in specs
            ]
        )

    payload = request.data if isinstance(request.data, dict) else {}
    product_name = (payload.get("product_name") or "").strip()
    if not product_name:
        return Response(
            {"message": "product_name is required"}, status=status.HTTP_400_BAD_REQUEST
        )

    try:
        spec = MasterSpec.objects.create(
            product_name=product_name,
            product_name_eng=payload.get("product_name_eng") or "",
            mold_number=payload.get("mold_number") or "",
            color1=payload.get("color1") or "",
            color2=payload.get("color2") or "",
            default_quantity=int(payload.get("default_quantity") or 0),
            sku_id=payload.get("sku_id") or "",
            barcode=payload.get("barcode") or "",
            category_lg=payload.get("category_lg") or "",
            category_md=payload.get("category_md") or "",
            price=int(payload.get("price") or 0),
            lot_number=payload.get("lot_number") or "",
            components=payload.get("components") or "",
            image_url=payload.get("image_url") or "",
            prev_price=int(payload.get("prev_price") or 0),
            price_changed_at=_parse_date_ymd(payload.get("price_changed_at")),
            is_discontinued=bool(payload.get("is_discontinued") or False),
            is_no_outbound_3m=bool(payload.get("is_no_outbound_3m") or False),
            is_vf_item=bool(payload.get("is_vf_item") or False),
        )
    except Exception:
        return Response(
            {"message": "already exists"}, status=status.HTTP_400_BAD_REQUEST
        )

    # 수동 로케이션 입력 → BarcodeMaster 자동 등록/저장
    loc_saved = ""
    if "location" in payload:
        loc_val = str(payload.get("location") or "").strip()
        ok, msg = _upsert_barcode_location(
            spec.barcode,
            location=loc_val,
            product_name=spec.product_name,
            sku_id=spec.sku_id or "",
            clear_if_empty=True,
        )
        if ok:
            loc_saved = loc_val
        elif loc_val and msg == "barcode_empty":
            # 바코드 없이 로케이션만 온 경우: 생성은 됐으나 BM 미등록
            pass
    else:
        loc_saved = _barcode_location_map([spec.barcode]).get(
            (spec.barcode or "").strip(), ""
        )

    return Response(
        _master_spec_dict(spec, location=loc_saved),
        status=status.HTTP_201_CREATED,
    )


def _clear_no_outbound_3m_on_activity(*, barcodes=None, sku_ids=None) -> int:
    """
    출고/입고 발생 시 '3개월 미출고' 플래그만 자동 해제.

    단종(is_discontinued=True) 품목은 절대 자동 변경하지 않음.
    단종 이동/복구는 제품 마스터 UI·API 수동 조작만 허용.
    """
    from django.db.models import Q

    barcodes = {str(b).strip() for b in (barcodes or []) if b and str(b).strip()}
    sku_ids = {str(s).strip() for s in (sku_ids or []) if s and str(s).strip()}
    if not barcodes and not sku_ids:
        return 0

    scope = Q()
    if barcodes:
        scope |= Q(barcode__in=barcodes)
    if sku_ids:
        scope |= Q(sku_id__in=sku_ids)

    return MasterSpec.objects.filter(
        scope,
        is_discontinued=False,
        is_no_outbound_3m=True,
    ).update(is_no_outbound_3m=False)


def _normalize_master_status_flags(is_discontinued, is_no_outbound_3m, *, has_disc, has_no_out):
    """
    상태 플래그 정규화.
    - 단종=True 이면 3개월 미출고는 항상 False (단종 탭 단독 표시)
    - 둘 다 켜지지 않도록 상호 배타
    """
    disc = bool(is_discontinued) if has_disc else None
    no_out = bool(is_no_outbound_3m) if has_no_out else None
    if disc is True:
        no_out = False
    elif no_out is True and disc is None:
        # 미출고로 전환 시 단종은 명시적으로 해제
        disc = False
    elif disc is False and no_out is None:
        pass
    return disc, no_out


@api_view(["PATCH"])
def master_specs_bulk_update(request):
    """제품 마스터 일괄 수정 (분류/색상/상태 + 로케이션).
    로케이션은 BarcodeMaster에 저장(수동 입력 자동 등록). 단종은 수동만.
    """
    payload = request.data if isinstance(request.data, dict) else {}
    ids = payload.get("ids", [])
    if not ids:
        return Response({"message": "수정할 대상 품목 ID 목록(ids)이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST)

    update_data = {}
    for key in [
        "category_lg",
        "category_md",
        "color1",
        "color2",
        "is_discontinued",
        "is_no_outbound_3m",
        "is_vf_item",
    ]:
        if key in payload:
            if key in ["is_discontinued", "is_no_outbound_3m", "is_vf_item"]:
                update_data[key] = bool(payload.get(key))
            else:
                update_data[key] = str(payload.get(key) or "").strip()

    has_location = "location" in payload
    location_val = str(payload.get("location") or "").strip() if has_location else None

    # 단종 ↔ 3개월 미출고 상호 배타 (수동 전환 시 정리)
    has_disc = "is_discontinued" in update_data
    has_no_out = "is_no_outbound_3m" in update_data
    if has_disc or has_no_out:
        disc, no_out = _normalize_master_status_flags(
            update_data.get("is_discontinued"),
            update_data.get("is_no_outbound_3m"),
            has_disc=has_disc,
            has_no_out=has_no_out,
        )
        if disc is not None:
            update_data["is_discontinued"] = disc
        if no_out is not None:
            update_data["is_no_outbound_3m"] = no_out

    if not update_data and not has_location:
        return Response({"message": "수정할 데이터가 없습니다."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        updated_count = 0
        if update_data:
            updated_count = MasterSpec.objects.filter(id__in=ids).update(**update_data)

        loc_updated = 0
        loc_skipped_no_barcode = 0
        if has_location:
            for s in MasterSpec.objects.filter(id__in=ids).only(
                "id", "barcode", "product_name", "sku_id"
            ):
                ok, msg = _upsert_barcode_location(
                    s.barcode,
                    location=location_val or "",
                    product_name=s.product_name or "",
                    sku_id=s.sku_id or "",
                    clear_if_empty=True,
                )
                if ok:
                    loc_updated += 1
                elif msg == "barcode_empty":
                    loc_skipped_no_barcode += 1

        return Response(
            {
                "success": True,
                "updated_count": updated_count,
                "location_updated": loc_updated,
                "location_skipped_no_barcode": loc_skipped_no_barcode,
            }
        )
    except Exception as e:
        return Response({"message": f"일괄 수정 처리 실패: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["PUT", "DELETE"])
def master_specs_detail(request, id: int):
    spec = MasterSpec.objects.filter(id=int(id)).first()
    if not spec:
        return Response({"message": "Not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "DELETE":
        spec.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    payload = request.data if isinstance(request.data, dict) else {}
    has_disc = "is_discontinued" in payload
    has_no_out = "is_no_outbound_3m" in payload
    has_location = "location" in payload
    location_payload = (
        str(payload.get("location") or "").strip() if has_location else None
    )

    for key in [
        "product_name",
        "product_name_eng",
        "mold_number",
        "color1",
        "color2",
        "default_quantity",
        "sku_id",
        "barcode",
        "category_lg",
        "category_md",
        "price",
        "lot_number",
        "components",
        "image_url",
        "prev_price",
        "price_changed_at",
        "is_discontinued",
        "is_no_outbound_3m",
        "is_vf_item",
    ]:
        if key in payload:
            val = payload.get(key)
            if key == "price" or key == "default_quantity" or key == "prev_price":
                try:
                    val = int(val or 0)
                except Exception:
                    val = 0
            elif key == "price_changed_at":
                val = _parse_date_ymd(val)
            elif key in ["is_discontinued", "is_no_outbound_3m", "is_vf_item"]:
                val = bool(val)
            setattr(spec, key, val)

    # 단종은 수동 전환만: 단종 시 미출고 플래그 정리 (자동 배치와 혼선 방지)
    if has_disc or has_no_out:
        disc, no_out = _normalize_master_status_flags(
            spec.is_discontinued,
            spec.is_no_outbound_3m,
            has_disc=True,
            has_no_out=True,
        )
        if disc is not None:
            spec.is_discontinued = disc
        if no_out is not None:
            spec.is_no_outbound_3m = no_out

    spec.save()

    # 수동 로케이션 입력 → BarcodeMaster 자동 등록/저장
    loc_out = ""
    if has_location:
        ok, msg = _upsert_barcode_location(
            spec.barcode,
            location=location_payload or "",
            product_name=spec.product_name or "",
            sku_id=spec.sku_id or "",
            clear_if_empty=True,
        )
        if ok:
            loc_out = location_payload or ""
        elif msg == "barcode_empty" and (location_payload or ""):
            return Response(
                {
                    "message": "로케이션을 저장하려면 바코드가 필요합니다. 바코드를 먼저 입력해 주세요.",
                    "location_error": "barcode_empty",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
    else:
        loc_out = _barcode_location_map([spec.barcode]).get(
            (spec.barcode or "").strip(), ""
        )

    return Response(_master_spec_dict(spec, location=loc_out))


@api_view(["POST"])
def master_specs_sync_outbound_status(request):
    """최근 3개월간 출고 실적이 없는 품목들의 마스터 상태(is_no_outbound_3m)를 수동 갱신하는 API"""
    from django.core.management import call_command
    try:
        call_command("update_outbound_status")
        return Response({"success": True, "message": "출고 상태 정기 배치가 성공적으로 실행되었습니다."})
    except Exception as e:
        return Response({"message": f"배치 실행 실패: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def master_specs_sync_vf_from_stock(request):
    """
    Enhanced 전산 현재고 기준 VF 지정 (1차 단순).
    currentStock > 0 바코드 → is_vf_item=True
    바코드 있고 재고 없음 → is_vf_item=False
    바코드 공란 MasterSpec → 스킵
    """
    from .inventory_stock import (
        aggregate_movements_after_baseline,
        compute_current_stock,
    )

    latest = InventoryBaselineUpload.objects.order_by("-uploaded_at").first()
    if not latest:
        return Response(
            {"message": "재고 스냅샷이 없습니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    as_of = latest.as_of_date
    baseline_items = list(
        InventoryBaselineItem.objects.filter(upload=latest).only("barcode", "quantity_box")
    )
    base_map = {}
    for it in baseline_items:
        bc = (it.barcode or "").strip()
        if not bc:
            continue
        base_map[bc] = base_map.get(bc, 0) + int(it.quantity_box or 0)

    bcs = list(base_map.keys())
    out_agg, rcv_agg = aggregate_movements_after_baseline(
        as_of=as_of,
        barcodes=bcs,
        outbound_model=OutboundRecord,
        receipt_model=InventoryReceiptItem,
    )

    positive = set()
    for bc, base in base_map.items():
        cur = compute_current_stock(
            base, int(rcv_agg.get(bc) or 0), int(out_agg.get(bc) or 0)
        )
        if cur > 0:
            positive.add(bc)

    # baseline 밖 이후 입고 잔여
    from collections import defaultdict

    extra_rcv = defaultdict(int)
    extra_out = defaultdict(int)
    if as_of:
        for row in (
            InventoryReceiptItem.objects.filter(receipt_date__gt=as_of)
            .exclude(barcode="")
            .exclude(barcode__isnull=True)
            .values("barcode")
            .annotate(qty=Coalesce(Sum("quantity_box"), 0))
        ):
            bc = (row["barcode"] or "").strip()
            if bc and bc not in base_map:
                extra_rcv[bc] = int(row["qty"] or 0)
        from .inventory_stock import filter_outbound_for_stock

        for row in (
            filter_outbound_for_stock(
                OutboundRecord.objects.filter(outbound_date__gt=as_of)
                .exclude(barcode="")
                .exclude(barcode__isnull=True)
            )
            .values("barcode")
            .annotate(qty=Coalesce(Sum("box_quantity"), 0))
        ):
            bc = (row["barcode"] or "").strip()
            if bc and bc not in base_map:
                extra_out[bc] = int(row["qty"] or 0)
    for bc in set(extra_rcv) | set(extra_out):
        cur = compute_current_stock(0, extra_rcv.get(bc, 0), extra_out.get(bc, 0))
        if cur > 0:
            positive.add(bc)

    # MasterSpec 반영
    specs = list(MasterSpec.objects.exclude(barcode="").only("id", "barcode", "is_vf_item"))
    to_true_ids = []
    to_false_ids = []
    for s in specs:
        bc = (s.barcode or "").strip()
        if not bc:
            continue
        want = bc in positive
        if want and not s.is_vf_item:
            to_true_ids.append(s.id)
        elif (not want) and s.is_vf_item:
            to_false_ids.append(s.id)

    if to_true_ids:
        MasterSpec.objects.filter(id__in=to_true_ids).update(is_vf_item=True)
    if to_false_ids:
        MasterSpec.objects.filter(id__in=to_false_ids).update(is_vf_item=False)

    # 재고 있는데 마스터 없는 바코드
    master_bcs = {(s.barcode or "").strip() for s in specs}
    skipped_no_master = len(positive - master_bcs)

    return Response(
        {
            "success": True,
            "message": "현재고>0 품목을 VF로 지정하고, 재고 없는 바코드 품목은 VF 해제했습니다.",
            "asOf": as_of.isoformat() if as_of else None,
            "stock_positive_barcodes": len(positive),
            "vf_set_true": len(to_true_ids),
            "vf_set_false": len(to_false_ids),
            "skipped_no_master": skipped_no_master,
            "vf_total_after": MasterSpec.objects.filter(is_vf_item=True).count(),
        }
    )


@api_view(["GET"])
def master_specs_export_xlsx(request):
    """
    마스터 일괄 수정 양식 다운로드.
    쿼리: vf_only=1 | scope=vf|active|all (기본 all)
    컬럼 id 필수 키, location 수정 대상 (BarcodeMaster SoT)
    """
    import io

    scope = (request.query_params.get("scope") or "all").strip().lower()
    vf_only = (request.query_params.get("vf_only") or "").strip() in ("1", "true", "yes")
    if vf_only:
        scope = "vf"

    qs = MasterSpec.objects.all().order_by("product_name")
    if scope == "vf":
        qs = qs.filter(is_vf_item=True)
    elif scope == "active":
        qs = qs.filter(is_discontinued=False, is_no_outbound_3m=False)
    elif scope == "no_outbound_3m":
        qs = qs.filter(is_discontinued=False, is_no_outbound_3m=True)
    elif scope == "discontinued":
        qs = qs.filter(is_discontinued=True)

    specs = list(qs)
    loc_map = _barcode_location_map(
        [(s.barcode or "").strip() for s in specs if (s.barcode or "").strip()]
    )

    rows = []
    for s in specs:
        bc = (s.barcode or "").strip()
        rows.append(
            {
                "id": s.id,
                "product_name": s.product_name,
                "barcode": bc,
                "sku_id": s.sku_id or "",
                "location": loc_map.get(bc, ""),
                "category_lg": s.category_lg or "",
                "category_md": s.category_md or "",
                "color1": s.color1 or "",
                "color2": s.color2 or "",
                "is_vf_item": 1 if s.is_vf_item else 0,
                "is_discontinued": 1 if s.is_discontinued else 0,
                "is_no_outbound_3m": 1 if s.is_no_outbound_3m else 0,
                "price": int(s.price or 0),
            }
        )

    try:
        import pandas as pd
    except Exception:
        return Response(
            {"message": "pandas 필요"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    df = pd.DataFrame(rows)
    readme = pd.DataFrame(
        [
            {"rule": "id", "description": "필수. 수정 대상 MasterSpec PK. 변경 금지"},
            {"rule": "product_name", "description": "참고용. 업로드 시 이름 변경 안 함"},
            {"rule": "barcode / sku_id", "description": "참고. 로케이션은 barcode로 BarcodeMaster 저장"},
            {"rule": "location", "description": "수정 대상. 셀 비움=유지. 삭제 시 __CLEAR__ 입력"},
            {"rule": "category_lg/md, color1/2", "description": "비움=유지"},
            {"rule": "is_discontinued / is_no_outbound_3m", "description": "0/1. 비움=유지. 단종 수동만"},
            {"rule": "is_vf_item", "description": "수정 가능. 1/true=VF설정, 0/false=VF 해제, 비움=유지. CSV 전체 재동기화 시 덮일 수 있음"},
            {"rule": "price", "description": "참고(단가 대량 변경은 FC 단가 업로드 사용)"},
            {"rule": "업로드", "description": "POST /api/master/specs/import-bulk (이 파일 그대로)"},
        ]
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="master_bulk")
        readme.to_excel(writer, index=False, sheet_name="README")
    buf.seek(0)

    filename = f"master_bulk_{scope}_{timezone.localdate().isoformat()}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@api_view(["POST"])
def master_specs_import_bulk(request):
    """
    일괄 수정 양식 재업로드.
    - MasterSpec: category/color/status (셀 비움=유지)
    - location: 비움=유지, __CLEAR__=삭제, 값 있으면 BarcodeMaster 자동 등록
    """
    if "file" not in request.FILES:
        return Response({"message": "파일이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST)

    file_obj = request.FILES["file"]
    name = (file_obj.name or "").lower()
    if not name.endswith((".xlsx", ".xls", ".csv")):
        return Response(
            {"message": "xlsx/xls/csv 만 지원합니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        import io
        import pandas as pd

        raw = file_obj.read()
        if name.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(raw), dtype=str)
        else:
            df = pd.read_excel(io.BytesIO(raw), sheet_name=0, dtype=str)
        df = df.fillna("")
    except Exception as e:
        return Response(
            {"message": f"파일 읽기 실패: {e}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # normalize columns
    colmap = {}
    for c in df.columns:
        key = str(c).strip().lower().replace(" ", "_")
        colmap[c] = key
    df = df.rename(columns=colmap)

    def cell(row, *names):
        for n in names:
            if n in row.index:
                return str(row.get(n) or "").strip()
        return ""

    updated_specs = 0
    updated_locations = 0
    updated_vf = 0
    skipped_no_id = 0
    skipped_no_barcode = 0
    errors = []

    def _as_bool_cell(v: str):
        s = str(v).strip().lower()
        if s in ("1", "true", "yes", "y", "vf", "단종"):
            return True
        if s in ("0", "false", "no", "n"):
            return False
        return None

    for i, row in df.iterrows():
        id_raw = cell(row, "id")
        if not id_raw:
            skipped_no_id += 1
            continue
        try:
            sid = int(float(id_raw))
        except Exception:
            skipped_no_id += 1
            errors.append({"row": int(i) + 2, "error": f"invalid id: {id_raw}"})
            continue

        spec = MasterSpec.objects.filter(id=sid).first()
        if not spec:
            errors.append({"row": int(i) + 2, "error": f"id {sid} not found"})
            continue

        changed = False
        for field, aliases in [
            ("category_lg", ("category_lg", "대분류")),
            ("category_md", ("category_md", "중분류")),
            ("color1", ("color1", "색상1")),
            ("color2", ("color2", "색상2")),
        ]:
            val = cell(row, *aliases)
            if val:
                if getattr(spec, field) != val:
                    setattr(spec, field, val)
                    changed = True

        disc_raw = cell(row, "is_discontinued", "단종")
        no_out_raw = cell(row, "is_no_outbound_3m", "미출고")
        has_disc = disc_raw != ""
        has_no_out = no_out_raw != ""
        if has_disc or has_no_out:
            disc_v = _as_bool_cell(disc_raw) if has_disc else spec.is_discontinued
            no_out_v = _as_bool_cell(no_out_raw) if has_no_out else spec.is_no_outbound_3m
            if disc_v is None:
                disc_v = spec.is_discontinued
            if no_out_v is None:
                no_out_v = spec.is_no_outbound_3m
            disc_v, no_out_v = _normalize_master_status_flags(
                disc_v, no_out_v, has_disc=True, has_no_out=True
            )
            if disc_v is not None and spec.is_discontinued != disc_v:
                spec.is_discontinued = disc_v
                changed = True
            if no_out_v is not None and spec.is_no_outbound_3m != no_out_v:
                spec.is_no_outbound_3m = no_out_v
                changed = True

        # VF 설정/해제 (비움=유지)
        vf_raw = cell(row, "is_vf_item", "vf", "vf품목")
        if vf_raw != "":
            vf_v = _as_bool_cell(vf_raw)
            if vf_v is not None and bool(spec.is_vf_item) != vf_v:
                spec.is_vf_item = vf_v
                changed = True
                updated_vf += 1

        if changed:
            spec.save()
            updated_specs += 1

        loc_raw = cell(row, "location", "로케이션", "위치")
        if loc_raw:
            clear = loc_raw.upper() in ("__CLEAR__", "CLEAR", "삭제")
            loc_val = "" if clear else loc_raw
            ok, msg = _upsert_barcode_location(
                spec.barcode,
                location=loc_val,
                product_name=spec.product_name or "",
                sku_id=spec.sku_id or "",
                clear_if_empty=clear,
            )
            if ok:
                updated_locations += 1
            elif msg == "barcode_empty":
                skipped_no_barcode += 1
                errors.append(
                    {
                        "row": int(i) + 2,
                        "id": sid,
                        "error": "location needs barcode",
                    }
                )

    return Response(
        {
            "success": True,
            "updated_specs": updated_specs,
            "updated_locations": updated_locations,
            "updated_vf": updated_vf,
            "skipped_no_id": skipped_no_id,
            "skipped_no_barcode": skipped_no_barcode,
            "errors": errors[:50],
            "error_count": len(errors),
        }
    )


@api_view(["GET"])
def master_spec_current_stock(request):
    """
    마스터 품목 클릭 리포트용 현재고 조회.
    VF 품목 바코드 기준 — inventory_stock 규칙 (baseline + 이후입고 − 이후출고).
    query: barcode=...  (필수)
    """
    barcode = (request.query_params.get("barcode") or "").strip()
    if not barcode:
        return Response(
            {"message": "barcode required", "found": False, "currentStock": None},
            status=status.HTTP_400_BAD_REQUEST,
        )

    latest_upload = InventoryBaselineUpload.objects.order_by("-uploaded_at").first()
    if not latest_upload:
        return Response(
            {
                "found": False,
                "barcode": barcode,
                "currentStock": None,
                "message": "재고 스냅샷 없음",
            }
        )

    as_of = latest_upload.as_of_date
    item = (
        InventoryBaselineItem.objects.filter(upload=latest_upload, barcode=barcode)
        .order_by("-id")
        .first()
    )
    base_qty = int(item.quantity_box or 0) if item else 0
    in_baseline = item is not None

    from .inventory_stock import (
        aggregate_movements_after_baseline,
        compute_current_stock,
    )

    outbound_agg, receipt_agg = aggregate_movements_after_baseline(
        as_of=as_of,
        barcodes=[barcode],
        outbound_model=OutboundRecord,
        receipt_model=InventoryReceiptItem,
    )
    rcv = int(receipt_agg.get(barcode) or 0)
    out = int(outbound_agg.get(barcode) or 0)
    # baseline 없으면 0 + 이후 이동만 (신규 VF 등) — found=False로 구분
    current = compute_current_stock(base_qty, rcv, out) if in_baseline else compute_current_stock(0, rcv, out)

    # 마지막 입고/출고 (UI 표시용 — 누적 합과 별개)
    last_in = (
        InventoryReceiptItem.objects.filter(barcode=barcode)
        .exclude(receipt_date__isnull=True)
        .order_by("-receipt_date", "-id")
        .first()
    )
    last_inbound_date = None
    last_inbound_qty = 0
    if last_in and last_in.receipt_date:
        last_inbound_date = last_in.receipt_date.isoformat()
        # 같은 날 입고 합
        last_inbound_qty = int(
            InventoryReceiptItem.objects.filter(
                barcode=barcode, receipt_date=last_in.receipt_date
            ).aggregate(s=Coalesce(Sum("quantity_box"), 0))["s"]
            or 0
        )

    from .inventory_stock import filter_outbound_for_stock

    last_out = (
        filter_outbound_for_stock(
            OutboundRecord.objects.filter(barcode=barcode).exclude(
                outbound_date__isnull=True
            )
        )
        .order_by("-outbound_date", "-id")
        .first()
    )
    last_outbound_date = None
    last_outbound_qty = 0
    if last_out and last_out.outbound_date:
        last_outbound_date = last_out.outbound_date.isoformat()
        last_outbound_qty = int(
            filter_outbound_for_stock(
                OutboundRecord.objects.filter(
                    barcode=barcode, outbound_date=last_out.outbound_date
                )
            ).aggregate(s=Coalesce(Sum("box_quantity"), 0))["s"]
            or 0
        )

    bm = BarcodeMaster.objects.filter(barcode=barcode).only("location", "product_name").first()
    return Response(
        {
            "found": in_baseline or rcv > 0 or out > 0,
            "inBaseline": in_baseline,
            "barcode": barcode,
            "currentStock": int(current),
            "baseStock": int(base_qty) if in_baseline else 0,
            "receiptQty": rcv,
            "outboundQty": out,
            "asOf": as_of.isoformat() if as_of else None,
            "lastInboundDate": last_inbound_date,
            "lastInboundQty": last_inbound_qty,
            "lastOutboundDate": last_outbound_date,
            "lastOutboundQty": last_outbound_qty,
            "location": (bm.location or "").strip() if bm else "",
            "productName": (bm.product_name or "") if bm else "",
        }
    )


@api_view(["POST"])
def master_spec_upload_image(request):
    """Product Spec 사진 업로드 API"""
    from django.core.files.storage import FileSystemStorage
    from django.conf import settings
    
    file_obj = request.FILES.get("image") or request.FILES.get("file")
    if not file_obj:
        return Response({"message": "이미지 파일이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST)
        
    try:
        specs_dir = os.path.join(settings.MEDIA_ROOT, "specs")
        os.makedirs(specs_dir, exist_ok=True)
        
        fs = FileSystemStorage(location=specs_dir, base_url="/media/specs/")
        filename = fs.save(file_obj.name, file_obj)
        uploaded_file_url = fs.url(filename)
        return Response({"success": True, "image_url": uploaded_file_url})
    except Exception as e:
        return Response({"message": f"파일 업로드 실패: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _parse_price_cell(val) -> int:
    """엑셀 단가 셀 파싱 (5,280 / 5280 / 5,280원)."""
    try:
        s = str(val if val is not None else "").strip()
        if not s or s.lower() in ("nan", "none", "-", ""):
            return 0
        s = s.replace(",", "").replace("원", "").replace(" ", "")
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _parse_fc_datetime(val):
    """FC 입고 시각 파싱 (2026/07/14 17:46:05, YYYY-MM-DD 등)."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.to_pydatetime() if hasattr(dt, "to_pydatetime") else dt
    except Exception:
        return None


def _sync_master_prices_from_fc_stocked_df(df) -> dict:
    """
    단가 단일 기준: 쿠팡 FC 입고 파일 (Coupang_Stocked_Data_List) 양식.

    필수: SKU번호, SKU명, 단가
    권장: 입고/반출시각 (최신 단가 선정·변동일)
    선택: 바코드/상품바코드

    단가 컬럼(단가)을 MasterSpec.price 에 반영.
    SKU별 최신 입고시각 행의 단가를 사용.
    """
    if df is None or getattr(df, "empty", True):
        return {
            "success": False,
            "message": "데이터가 없습니다.",
            "new_created": 0,
            "updated": 0,
            "unchanged": 0,
            "total_items": 0,
            "price_changed": 0,
        }

    work = df.copy()
    work.columns = [str(c).strip() for c in work.columns]
    cols = list(work.columns)

    sku_col = next((c for c in ["SKU번호", "상품번호"] if c in cols), None)
    name_col = next((c for c in ["SKU명", "상품이름", "상품명"] if c in cols), None)
    # 통일 기준: FC 입고 파일의 '단가' 우선 (구 양식 매입가는 호환용 fallback)
    price_col = next((c for c in ["단가", "매입가"] if c in cols), None)
    date_col = next(
        (c for c in ["입고/반출시각", "입고예정일", "발주등록일시"] if c in cols), None
    )
    barcode_col = next((c for c in ["상품바코드", "바코드"] if c in cols), None)

    if not sku_col or not name_col or not price_col:
        return {
            "success": False,
            "message": (
                "필수 컬럼이 없습니다. FC 입고 파일(Coupang_Stocked_Data_List) 양식이 필요합니다. "
                "필수: SKU번호, SKU명, 단가"
            ),
            "new_created": 0,
            "updated": 0,
            "unchanged": 0,
            "total_items": 0,
            "price_changed": 0,
        }

    # SKU별 최신 단가 취합
    latest_by_sku = {}  # sku -> {dt, price, name, barcode}
    for _, row in work.iterrows():
        sku_id = str(row.get(sku_col, "") or "").strip()
        product_name = str(row.get(name_col, "") or "").strip()
        if not sku_id or not product_name or sku_id.lower() == "nan":
            continue
        if product_name.lower() == "nan":
            continue

        price = _parse_price_cell(row.get(price_col))
        if price <= 0:
            continue

        raw_dt = row.get(date_col) if date_col else None
        dt = _parse_fc_datetime(raw_dt) or datetime.now()
        barcode = ""
        if barcode_col:
            barcode = str(row.get(barcode_col, "") or "").strip()
            if barcode.lower() in ("nan", "none", "-"):
                barcode = ""

        prev = latest_by_sku.get(sku_id)
        if prev is None or dt >= prev["dt"]:
            latest_by_sku[sku_id] = {
                "dt": dt,
                "price": price,
                "name": product_name,
                "barcode": barcode,
            }

    registered_by_sku = {
        s.sku_id.strip(): s for s in MasterSpec.objects.all() if s.sku_id
    }
    registered_by_barcode = {
        s.barcode.strip(): s for s in MasterSpec.objects.all() if s.barcode
    }
    registered_by_name = {
        s.product_name.strip(): s
        for s in MasterSpec.objects.all()
        if s.product_name
    }

    outbound_cats = {}
    for row in (
        OutboundRecord.objects.exclude(category="")
        .exclude(category__isnull=True)
        .values("product_name")
        .annotate(category=Max("category"))
    ):
        outbound_cats[str(row.get("product_name") or "").strip()] = str(
            row.get("category") or ""
        ).strip()

    barcode_cats = {}
    for row in (
        BarcodeMaster.objects.exclude(category="")
        .exclude(category__isnull=True)
        .values("product_name")
        .annotate(category=Max("category"))
    ):
        barcode_cats[str(row.get("product_name") or "").strip()] = str(
            row.get("category") or ""
        ).strip()

    new_created = 0
    updated = 0
    unchanged = 0
    price_changed = 0

    for sku_id, info in latest_by_sku.items():
        latest_price = int(info["price"])
        product_name = info["name"]
        barcode = info["barcode"]
        changed_date = info["dt"].date() if hasattr(info["dt"], "date") else info["dt"]

        spec = None
        if sku_id in registered_by_sku:
            spec = registered_by_sku[sku_id]
        elif barcode and barcode in registered_by_barcode:
            spec = registered_by_barcode[barcode]
        elif product_name in registered_by_name:
            spec = registered_by_name[product_name]

        if spec:
            old_price = int(spec.price or 0)
            dirty = False
            if latest_price != old_price:
                if old_price > 0:
                    spec.prev_price = old_price
                    spec.price_changed_at = changed_date
                spec.price = latest_price
                price_changed += 1
                dirty = True
            if not spec.sku_id:
                spec.sku_id = sku_id
                dirty = True
            if barcode and not spec.barcode:
                spec.barcode = barcode
                dirty = True
            if dirty:
                spec.save()
                updated += 1
            else:
                unchanged += 1
        else:
            inferred_cat = ""
            name_strip = product_name.strip()
            if name_strip in outbound_cats:
                inferred_cat = outbound_cats[name_strip]
            elif name_strip in barcode_cats:
                inferred_cat = barcode_cats[name_strip]
            if not inferred_cat:
                inferred_cat = "미분류"

            MasterSpec.objects.create(
                product_name=product_name,
                sku_id=sku_id,
                barcode=barcode,
                price=latest_price,
                prev_price=0,
                price_changed_at=changed_date,
                category_lg=inferred_cat,
                category_md="미분류",
            )
            new_created += 1
            price_changed += 1

    return {
        "success": True,
        "source": "fc_stocked_data_list",
        "price_column": price_col,
        "new_created": new_created,
        "updated": updated,
        "unchanged": unchanged,
        "price_changed": price_changed,
        "total_items": len(latest_by_sku),
    }


@api_view(["POST"])
def master_spec_upload_excel(request):
    """
    제품 마스터 단가 연동 엑셀 업로드.
    단일 기준: FC 입고 파일(Coupang_Stocked_Data_List) 양식 — 단가 컬럼.
    """
    if "file" not in request.FILES:
        return Response({"message": "파일이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST)

    file_obj = request.FILES["file"]
    if not file_obj.name.endswith((".xlsx", ".xls")):
        return Response({"message": "엑셀 파일만 지원합니다."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        import io

        df = pd.read_excel(io.BytesIO(file_obj.read()), dtype=str)
        df = df.fillna("")
        result = _sync_master_prices_from_fc_stocked_df(df)
        if not result.get("success"):
            return Response(
                {"message": result.get("message") or "단가 동기화 실패"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response(result)
    except Exception as e:
        logger.error(f"Master Excel Upload Error: {e}")
        return Response(
            {"message": f"업로드 처리 실패: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
def master_extract(request):
    existing = set(MasterSpec.objects.values_list("product_name", flat=True))

    def _as_int(v):
        try:
            if v is None:
                return 0
            if isinstance(v, str):
                v = v.replace(",", "").strip()
            return int(float(v))
        except Exception:
            return 0

    # Best row per product_name: latest date, then highest id
    best_by_name = {}
    qs = ProductionLog.objects.exclude(product_name="").order_by("-date", "-id")
    for row in qs.iterator():
        name = (row.product_name or "").strip()
        if name and name not in best_by_name:
            best_by_name[name] = row

    added = 0
    updated = 0
    for name in sorted(best_by_name.keys()):
        row = best_by_name.get(name)
        if not row:
            continue

        eng = (row.product_name_eng or "").strip()
        mold = (row.mold_number or "").strip()
        c1 = (row.color1 or "").strip()
        c2 = (row.color2 or "").strip()
        # Use unit for default_quantity (as requested)
        default_qty = _as_int(row.unit)

        if name in existing:
            spec = MasterSpec.objects.filter(product_name=name).first()
            if not spec:
                continue
            changed = False
            if eng and not (spec.product_name_eng or "").strip():
                spec.product_name_eng = eng
                changed = True
            if mold and not (spec.mold_number or "").strip():
                spec.mold_number = mold
                changed = True
            if c1 and not (spec.color1 or "").strip():
                spec.color1 = c1
                changed = True
            if c2 and not (spec.color2 or "").strip():
                spec.color2 = c2
                changed = True
            if default_qty and int(spec.default_quantity or 0) == 0:
                spec.default_quantity = default_qty
                changed = True
            if changed:
                spec.save()
                updated += 1
            continue

        spec = MasterSpec.objects.create(
            product_name=name,
            product_name_eng=eng,
            mold_number=mold,
            color1=c1,
            color2=c2,
            default_quantity=default_qty,
        )
        existing.add(spec.product_name)
        added += 1

    return Response({"added": added, "updated": updated})


@api_view(["GET"])
def production_list(request):
    paginator = LimitOffsetPagination()
    paginator.default_limit = 100
    paginator.max_limit = 1000

    all_dates = list(
        ProductionLog.objects.values_list("date", flat=True).distinct().order_by("date")
    )
    latest = all_dates[-1].isoformat() if all_dates else None

    # Date filter parameter
    date_param = request.GET.get("date", "").strip()
    get_all = request.GET.get("all", "").lower() == "true"

    # Determine which data to return
    if date_param:
        # Specific date requested
        try:
            date_obj = datetime.fromisoformat(date_param).date()
            data_qs = ProductionLog.objects.filter(date=date_obj)
            data_qs = data_qs.order_by("sort_order", "id")
            latest = date_param
        except ValueError:
            return Response(
                {"error": f"Invalid date format: {date_param}. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST
            )
    elif get_all:
        # All data requested
        data_qs = ProductionLog.objects.all()
        data_qs = data_qs.order_by("date", "sort_order", "id")
    else:
        # Default: latest date only
        data_qs = (
            ProductionLog.objects.filter(date=all_dates[-1])
            if all_dates
            else ProductionLog.objects.none()
        )
        data_qs = data_qs.order_by("sort_order", "id")

    data_page = paginator.paginate_queryset(data_qs, request)
    data_serializer = ProductionLogSerializer(data_page, many=True)
    data = data_serializer.data

    return paginator.get_paginated_response(
        {
            "success": True,
            "latestDate": latest,
            "data": data,
            "allDates": [d.isoformat() for d in all_dates],
            "totalRecords": ProductionLog.objects.count(),
        }
    )


@api_view(["DELETE"])
def production_delete(request, id: int):
    """DELETE /api/production/<id> - Delete a single production log by ID"""
    try:
        item = ProductionLog.objects.filter(id=int(id)).first()
    except (ValueError, TypeError):
        return Response({"message": "Invalid id format"}, status=status.HTTP_400_BAD_REQUEST)

    if not item:
        return Response({"message": "Not found"}, status=status.HTTP_404_NOT_FOUND)

    item.delete()
    return Response({"success": True, "deleted_id": id})


@api_view(["POST"])
def production_bulk_status(request):
    payload = request.data if isinstance(request.data, dict) else {}
    status_value = payload.get("status")
    if not status_value:
        return Response(
            {"message": "status is required"}, status=status.HTTP_400_BAD_REQUEST
        )
    status_value = _production_normalize_status(str(status_value))

    scope = (payload.get("scope") or "").strip().lower()
    ids = payload.get("ids")
    date = (payload.get("date") or "").strip()

    targets = []
    if isinstance(ids, list) and ids:
        idset = [int(x) for x in ids if str(x).isdigit()]
        targets = list(ProductionLog.objects.filter(id__in=idset))
    elif date:
        try:
            targets = list(ProductionLog.objects.filter(date=date))
        except Exception:
            targets = []
    elif scope == "all":
        targets = list(ProductionLog.objects.all())
    else:
        return Response(
            {"message": "ids or date or scope=all is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    updated = 0
    for row in targets:
        _production_apply_status_model(row, status_value)
        row.save()
        updated += 1

    return Response({"success": True, "updated": updated, "status": status_value})


@api_view(["GET"])
def production_template(request):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "date",
            "machineNumber",
            "moldNumber",
            "productName",
            "productNameEng",
            "color1",
            "color2",
            "unit",
            "quantity",
            "unitQuantity",
            "total",
            "status",
        ]
    )
    resp = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="production_template.csv"'
    return resp


@api_view(["POST"])
def production_log(request):
    payload = request.data if isinstance(request.data, dict) else {}
    record = (
        payload.get("record") if isinstance(payload.get("record"), dict) else payload
    )
    date = (record.get("date") or "").strip() if isinstance(record, dict) else ""
    if not date:
        return Response(
            {"message": "date is required"}, status=status.HTTP_400_BAD_REQUEST
        )

    try:
        date_obj = datetime.fromisoformat(date).date()
    except ValueError:
        try:
            date_obj = pd.to_datetime(date).date()
        except Exception:
            return Response(
                {"message": "invalid date"}, status=status.HTTP_400_BAD_REQUEST
            )

    machine = str(record.get("machineNumber") or "").strip()
    mold = str(record.get("moldNumber") or "").strip()
    pname = str(record.get("productName") or "").strip()
    c1 = str(record.get("color1") or "").strip()
    c2 = str(record.get("color2") or "").strip()

    def _to_int(v):
        try:
            if v is None:
                return 0
            if isinstance(v, str):
                v = v.replace(",", "").strip()
            return int(float(v))
        except (ValueError, TypeError):
            return 0

    qty = _to_int(record.get("quantity"))
    unit_qty = _to_int(record.get("unitQuantity"))
    unit_raw = str(record.get("unit") or "").replace(",", "").strip()
    if not unit_qty and unit_raw.isdigit():
        unit_qty = int(unit_raw)
    total = _production_calc_total(qty, unit_qty, record.get("total"))
    status_value = _production_normalize_status(record.get("status") or "pending")

    defaults = {
        "product_name_eng": str(record.get("productNameEng") or "").strip(),
        "unit": str(record.get("unit") or "").strip(),
        "quantity": qty,
        "unit_quantity": unit_qty,
        "total": total,
        "color1": c1,
        "color2": c2,
        "status": status_value,
    }

    obj, created = ProductionLog.objects.update_or_create(
        date=date_obj,
        machine_number=machine,
        mold_number=mold,
        product_name=pname,
        color1=c1,
        color2=c2,
        unit=str(record.get("unit") or "").strip(),
        quantity=qty,
        unit_quantity=unit_qty,
        defaults={
            "product_name_eng": str(record.get("productNameEng") or "").strip(),
            "total": total,
            "status": status_value,
        },
    )
    _production_apply_status_model(obj, status_value)
    obj.total = _production_calc_total(obj.quantity, obj.unit_quantity, obj.total)
    obj.save()

    return Response(
        {"success": True, "record": _production_model_to_dict(obj)},
        status=status.HTTP_201_CREATED,
    )


@api_view(["DELETE"])
def production_log_bulk_delete(request):
    """IDs 리스트로 여러 ProductionLog 삭제"""
    ids = request.data.get("ids", [])
    if not ids:
        return Response(
            {"success": False, "error": "ids required"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        id_list = [int(x) for x in ids if str(x).isdigit()]
    except (ValueError, TypeError):
        return Response(
            {"success": False, "error": "invalid ids format"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    # 빈 배열로 전체 삭제 방지
    if not id_list:
        return Response(
            {"success": False, "error": "no valid ids provided"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    deleted, _ = ProductionLog.objects.filter(id__in=id_list).delete()
    return Response({"success": True, "deleted": deleted})


@api_view(["PUT", "DELETE"])
def production_log_detail(request, id: int):
    try:
        item = ProductionLog.objects.filter(id=int(id)).first()
    except (ValueError, TypeError):
        return Response(
            {"message": "Invalid id format"}, status=status.HTTP_400_BAD_REQUEST
        )
    if not item:
        return Response({"message": "Not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "DELETE":
        item.delete()
        return Response({"success": True})

    payload = request.data if isinstance(request.data, dict) else {}
    if "date" in payload:
        try:
            item.date = datetime.fromisoformat(str(payload.get("date"))).date()
        except ValueError:
            pass
    if "machineNumber" in payload:
        item.machine_number = str(payload.get("machineNumber") or "").strip()
    if "moldNumber" in payload:
        item.mold_number = str(payload.get("moldNumber") or "").strip()
    if "productName" in payload:
        item.product_name = str(payload.get("productName") or "").strip()
    if "productNameEng" in payload:
        item.product_name_eng = str(payload.get("productNameEng") or "").strip()
    if "color1" in payload:
        item.color1 = str(payload.get("color1") or "").strip()
    if "color2" in payload:
        item.color2 = str(payload.get("color2") or "").strip()
    if "unit" in payload:
        item.unit = str(payload.get("unit") or "").strip()
    if "quantity" in payload:
        try:
            item.quantity = int(
                float(str(payload.get("quantity")).replace(",", "").strip())
            )
        except (ValueError, TypeError):
            item.quantity = 0
    if "unitQuantity" in payload:
        try:
            item.unit_quantity = int(
                float(str(payload.get("unitQuantity")).replace(",", "").strip())
            )
        except Exception:
            item.unit_quantity = 0
    if not item.unit_quantity:
        u = str(item.unit or "").replace(",", "").strip()
        if u.isdigit():
            item.unit_quantity = int(u)
    if "status" in payload:
        item.status = _production_normalize_status(
            str(payload.get("status") or "pending")
        )

    _production_apply_status_model(item, item.status)
    item.total = _production_calc_total(item.quantity, item.unit_quantity, item.total)
    item.save()
    return Response({"success": True, "record": _production_model_to_dict(item)})


@api_view(["POST"])
def production_log_bulk_reorder(request):
    """벌크 정렬 순서 업데이트"""
    orders = request.data.get("orders", [])
    if not orders:
        return Response({"success": False, "error": "orders required"}, status=400)
    with transaction.atomic():
        for item in orders:
            pid = item.get("id")
            sort_order = item.get("sort_order", 0)
            ProductionLog.objects.filter(id=pid).update(sort_order=sort_order)
    return Response({"success": True, "updated": len(orders)})


@api_view(["POST"])
def production_log_carry_forward(request):
    """
    미완료 작업(from_date, status != 'ended')을 to_date로 이월.

    POST /api/production-log/carry-forward
    Query params:
      from_date (optional, default: 어제)
      to_date   (optional, default: 오늘)
    Body (optional):
      { "from_date": "YYYY-MM-DD", "to_date": "YYYY-MM-DD" }

    동작:
      - from_date의 status != 'ended' 인 ProductionLog 조회
      - (date, machine_number, mold_number, color1, color2) 조합이
        to_date에 이미 존재하면 -> 건너뜀 (중복 방지)
      - 없으면 -> date를 to_date로 UPDATE
      - ended 상태는 원래 날짜에 유지 (이력 보존)
    """
    today = timezone.localdate()
    yesterday = today - timedelta(days=1)

    payload = request.data if isinstance(request.data, dict) else {}

    # 쿼리파라미터 우선, 없으면 body, 없으면 기본값
    from_date_str = (
        request.query_params.get("from_date")
        or payload.get("from_date")
        or yesterday.isoformat()
    )
    to_date_str = (
        request.query_params.get("to_date")
        or payload.get("to_date")
        or today.isoformat()
    )

    try:
        from_date = datetime.fromisoformat(from_date_str).date()
    except (ValueError, TypeError):
        return Response(
            {"success": False, "error": "invalid from_date format. Use YYYY-MM-DD."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        to_date = datetime.fromisoformat(to_date_str).date()
    except (ValueError, TypeError):
        return Response(
            {"success": False, "error": "invalid to_date format. Use YYYY-MM-DD."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # from_date의 미완료(ended가 아닌) 작업 조회
    sources = list(
        ProductionLog.objects.filter(date=from_date).exclude(status="ended")
    )

    carried = 0
    skipped = 0
    with transaction.atomic():
        for src in sources:
            # to_date에 동일한 (machine, mold, color1, color2) 조합이 있으면 스킵
            exists = ProductionLog.objects.filter(
                date=to_date,
                machine_number=src.machine_number,
                mold_number=src.mold_number,
                color1=src.color1,
                color2=src.color2,
            ).exists()
            if exists:
                skipped += 1
                continue
            # 이월: date만 to_date로 변경 (나머지 필드는 유지)
            ProductionLog.objects.filter(pk=src.pk).update(date=to_date)
            carried += 1

    return Response({
        "success": True,
        "carried": carried,
        "skipped": skipped,
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
    })


@api_view(["DELETE"])
def production_log_by_date(request, date: str):
    try:
        date_obj = datetime.fromisoformat(date).date()
    except Exception:
        return Response({"message": "invalid date"}, status=status.HTTP_400_BAD_REQUEST)
    deleted, _ = ProductionLog.objects.filter(date=date_obj).delete()
    return Response({"success": True, "deleted": deleted})


@api_view(["POST"])
def production_log_move_pending_to_today(request):
    """
    과거 날짜의 pending/start，生产日志移到今天
    POST /api/production-log/move-pending-to-today
    {
      "from_date": "2026-04-21"  // optional, moves all past dates if not specified
    }
    """
    from django.db import transaction
    from django.db.models import Q

    today = timezone.localdate()
    payload = request.data if isinstance(request.data, dict) else {}
    from_date_str = payload.get("from_date")

    try:
        from_date = datetime.fromisoformat(from_date_str).date() if from_date_str else None
    except ValueError:
        return Response({"message": "invalid from_date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

    # Build query - move pending/started from past dates to today
    query = ProductionLog.objects.filter(
        date__lt=today,
        status__in=["pending", "started"]
    )
    if from_date:
        query = query.filter(date=from_date)

    records = list(query.select_related().all())
    count = len(records)

    if count == 0:
        return Response({
            "success": True,
            "message": "이동할生产日志为0",
            "moved_count": 0
        })

    with transaction.atomic():
        # Delete existing records for today that would cause unique constraint conflict
        for record in records:
            # Check if a record with same (machine, mold, product, color1, color2, unit) exists on today
            existing = ProductionLog.objects.filter(
                date=today,
                machine_number=record.machine_number,
                mold_number=record.mold_number,
                product_name=record.product_name,
                color1=record.color1,
                color2=record.color2,
                unit=record.unit
            ).first()

            if existing:
                # Update existing instead of creating new
                existing.quantity = record.quantity
                existing.unit_quantity = record.unit_quantity
                existing.total = record.total
                existing.status = record.status
                existing.sort_order = record.sort_order
                existing.product_name_eng = record.product_name_eng
                existing.save()
                record.delete()
            else:
                # Just update the date
                record.date = today
                record.save()

    return Response({
        "success": True,
        "message": f"{count}개 생산로그를 오늘({today})로 이동했습니다.",
        "moved_count": count
    })


def _ingest_production_dataframe(df, success_message="생산 계획 데이터를 업로드했습니다."):
    """공통: 생산일지 DataFrame → ProductionLog 반영. Response 또는 에러 Response 반환."""
    if df is None or df.empty:
        return Response(
            {"message": "업로드할 데이터가 없습니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # Support common Korean column names by mapping them to canonical keys.
    # 붙여넣기/엑셀 헤더 변형 폭넓게 수용
    col_aliases = {
        "date": ["일자", "날짜", "생산일", "작업일", "date"],
        "machineNumber": [
            "기계번호",
            "기계",
            "호기",
            "호기번호",
            "머신",
            "M/C",
            "MC",
            "machine",
            "machine_number",
            "기계 no",
            "기계No",
        ],
        "moldNumber": ["금형", "금형번호", "금형#", "mold", "mold_number", "moldNumber"],
        "productName": ["제품명", "품명", "상품명", "제품", "product", "product_name", "productName"],
        "productNameEng": ["제품명(영문)", "영문명", "영문", "product_name_eng", "productNameEng"],
        "color1": ["색상", "색상1", "color", "color1"],
        "color2": ["색상2", "color2"],
        "unit": ["단위(문자)", "단위명", "unit", "박스단위"],
        "quantity": ["생산수량", "수량", "박스수", "박스수량", "quantity", "qty"],
        "unitQuantity": ["단위", "단위수량", "개수/박스", "박스당", "unit_quantity", "unitQuantity"],
        "total": ["총계", "합계", "total"],
        "status": ["상태", "status"],
    }

    def _norm_header(h: str) -> str:
        s = str(h or "").strip().replace("\ufeff", "")
        s = s.replace(" ", "").replace("_", "").lower()
        return s

    # 정규화 맵: 변형 헤더 → 원본 컬럼명
    cols_by_norm = {_norm_header(c): c for c in df.columns}

    rename_map = {}
    existing_cols = set(df.columns)
    for canonical, aliases in col_aliases.items():
        if canonical in existing_cols:
            continue
        # 정규화 매칭
        matched = None
        for alias in [canonical] + list(aliases):
            key = _norm_header(alias)
            if key in cols_by_norm:
                matched = cols_by_norm[key]
                break
            if alias in existing_cols:
                matched = alias
                break
        if matched and matched != canonical:
            rename_map[matched] = canonical
    if rename_map:
        df = df.rename(columns=rename_map)

    # machineNumber / moldNumber 없으면 빈 컬럼 추가 (붙여넣기 표에 자주 누락)
    if "machineNumber" not in df.columns:
        df["machineNumber"] = ""
    if "moldNumber" not in df.columns:
        df["moldNumber"] = ""

    required_cols = ["date", "productName"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        return Response(
            {
                "message": f"필수 컬럼이 없습니다: {', '.join(missing)}",
                "hint": "헤더 예: 날짜(또는 date), 제품명, 기계번호(선택), 금형(선택), 수량, 단위",
                "receivedColumns": list(df.columns),
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    # 박스 수량 기준 (1팔레트 기준)
    BOX_QUANTITY_STANDARDS = {
        "슬림형 앞판": 75,
        "에센셜 앞판": 140,
        "해피 바디": 270,
        "와이드 상판": 30,
        "와이드 앞판": 70,
        "와이드 서랍": 180,
        "데크타일": 72,
    }

    # 색상 세트 조합 규칙: 색상1 → 기대되는 색상2
    COLOR_COMBINATION_RULES = {
        "WHITE1": "WHITE 180",
        "WHITE2": "IVORY 1154",
    }

    warnings = []
    rows_processed = 0
    rows_added = 0
    rows_updated = 0

    def _to_int(v):
        try:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return 0
            if isinstance(v, str):
                v = v.replace(",", "").strip()
            return int(float(v))
        except Exception:
            return 0

    with transaction.atomic():
        for _idx, row in df.iterrows():
            product_name = row.get("productName")
            if product_name is None or str(product_name).strip() == "":
                continue

            dt = row.get("date")
            date_str = ""
            try:
                if pd.isna(dt):
                    date_str = ""
                else:
                    date_str = pd.to_datetime(dt).date().isoformat()
            except Exception:
                date_str = str(dt).strip() if dt is not None else ""

            if not date_str:
                continue

            machine = str(row.get("machineNumber") or "").strip()
            mold = str(row.get("moldNumber") or "").strip()
            pname = str(product_name or "").strip()
            c1 = str(row.get("color1") or "").strip()
            c2 = str(row.get("color2") or "").strip()
            try:
                date_obj = datetime.fromisoformat(date_str).date()
            except Exception:
                continue

            qty = _to_int(row.get("quantity"))
            # 엑셀 컬럼 의미:
            #   - unit         : 박스당 개수 (정수)  → DB.unit_quantity
            #   - unitQuantity : 단위명 (예: 'BOX')  → DB.unit (CharField)
            #   - quantity     : 박스 개수           → DB.quantity
            unit_qty = _to_int(row.get("unit")) or _to_int(row.get("unitQuantity"))
            unit_raw_value = row.get("unitQuantity")
            if isinstance(unit_raw_value, (int, float)) and not isinstance(unit_raw_value, bool):
                unit_label = ""
            else:
                unit_label = str(unit_raw_value or "").strip()

            # === 색상 세트 조합 확인 ===
            if c1 and c2:
                if c1 in COLOR_COMBINATION_RULES:
                    expected_color2 = COLOR_COMBINATION_RULES[c1]
                    if c2 != expected_color2:
                        warnings.append(
                            {
                                "type": "color_combination",
                                "row": _idx + 2,
                                "product": pname,
                                "color1": c1,
                                "color2": c2,
                                "expected": expected_color2,
                                "message": f'[{pname}] 색상1 "{c1}"일 때 색상2는 "{expected_color2}"이어야 합니다. (현재: {c2})',
                            }
                        )
                elif c1 not in COLOR_COMBINATION_RULES and c2:
                    warnings.append(
                        {
                            "type": "color_combination_unknown",
                            "row": _idx + 2,
                            "product": pname,
                            "color1": c1,
                            "color2": c2,
                            "message": f"[{pname}] 알 수 없는 색상 조합입니다. (색상1: {c1}, 색상2: {c2})",
                        }
                    )

            # === 박스 수량 확인 ===
            if unit_qty > 0 and pname in BOX_QUANTITY_STANDARDS:
                standard_qty = BOX_QUANTITY_STANDARDS[pname]
                if unit_qty != standard_qty:
                    warnings.append(
                        {
                            "type": "box_quantity",
                            "row": _idx + 2,
                            "product": pname,
                            "actual": unit_qty,
                            "standard": standard_qty,
                            "message": f"[{pname}] 표준 박스 수량은 {standard_qty}이어야 합니다. (현재: {unit_qty})",
                        }
                    )

            total = _production_calc_total(qty, unit_qty, row.get("total"))
            status_value = _production_normalize_status(
                str(row.get("status") or "pending")
            )

            defaults = {
                "product_name_eng": str(row.get("productNameEng") or "").strip(),
                "unit": unit_label,
                "quantity": qty,
                "unit_quantity": unit_qty,
                "total": total,
                "status": status_value,
            }

            obj, created = ProductionLog.objects.update_or_create(
                date=date_obj,
                machine_number=machine,
                mold_number=mold,
                product_name=pname,
                color1=c1,
                color2=c2,
                unit=unit_label,
                quantity=qty,
                unit_quantity=unit_qty,
                defaults=defaults,
            )
            _production_apply_status_model(obj, status_value)
            obj.total = _production_calc_total(
                obj.quantity, obj.unit_quantity, obj.total
            )
            obj.save()

            if created:
                rows_added += 1
            else:
                rows_updated += 1
            rows_processed += 1

    all_dates = list(
        ProductionLog.objects.values_list("date", flat=True)
        .distinct()
        .order_by("date")
    )
    latest_date = all_dates[-1].isoformat() if all_dates else None

    return Response(
        {
            "success": True,
            "message": success_message,
            "rowsProcessed": rows_processed,
            "rowsAdded": rows_added,
            "rowsUpdated": rows_updated,
            "latestDate": latest_date,
            "warnings": warnings,
        }
    )


@api_view(["POST"])
def upload_production_file(request):
    _file = request.FILES.get("productionFile")
    if not _file:
        return Response(
            {"message": "productionFile is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        xls = pd.ExcelFile(_file)

        sheet_name = None
        preferred = "상품목록 구성하기"
        if preferred in (xls.sheet_names or []):
            sheet_name = preferred
        else:
            for name in xls.sheet_names or []:
                try:
                    df_probe = xls.parse(name, nrows=1)
                    cols = [str(c).strip() for c in df_probe.columns]
                    if "date" in cols and "productName" in cols:
                        sheet_name = name
                        break
                except Exception:
                    continue
            if not sheet_name and (xls.sheet_names or []):
                sheet_name = xls.sheet_names[0]

        df = xls.parse(sheet_name) if sheet_name else pd.DataFrame()
        return _ingest_production_dataframe(
            df, success_message="생산 계획 파일을 업로드했습니다."
        )
    except Exception as e:
        return Response(
            {"message": "생산 계획 파일 처리 중 오류가 발생했습니다.", "error": str(e)},
            status=status.HTTP_400_BAD_REQUEST,
        )


def _production_text_looks_headerless(first_cell: str) -> bool:
    """첫 셀이 날짜 형태면 헤더 없이 데이터부터 시작했다고 판단."""
    s = str(first_cell or "").strip().replace("\ufeff", "")
    if not s:
        return False
    # 명시적 헤더 키워드면 헤더 있음
    header_keys = (
        "date",
        "날짜",
        "일자",
        "생산일",
        "productname",
        "제품명",
        "품명",
        "machinenumber",
        "기계",
        "호기",
    )
    if s.lower().replace(" ", "") in header_keys or s in (
        "날짜",
        "일자",
        "date",
        "제품명",
        "품명",
    ):
        return False
    try:
        pd.to_datetime(s)
        return True
    except Exception:
        return False


# 템플릿/엑셀 붙여넣기 기본 열 순서 (production_template.csv 와 맞춤)
# production_template.csv 열 순서와 동일
_PRODUCTION_DEFAULT_COLS = [
    "date",
    "machineNumber",
    "moldNumber",
    "productName",
    "productNameEng",
    "color1",
    "color2",
    "unit",  # 템플릿: 박스당 개수(숫자) 자리 — ingest에서 unit_quantity로 해석
    "quantity",  # 박스 수
    "unitQuantity",  # 템플릿: 단위명 BOX/P/EA
    "total",
    "status",
]


@api_view(["POST"])
def upload_production_text(request):
    """생산일지 표 텍스트(탭/쉼표/공백 구분) 붙여넣기 업로드."""
    payload = request.data if isinstance(request.data, dict) else {}
    text = (payload.get("text") or "").strip()
    if not text:
        return Response(
            {"message": "text is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        # BOM·이상 공백 정리
        text = text.replace("\ufeff", "").replace("\r\n", "\n").replace("\r", "\n")
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        if not lines:
            return Response(
                {"message": "업로드할 데이터가 없습니다."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        first = lines[0]
        # 구분자 추정: 탭 > 세미콜론 > 파이프 > 쉼표 > 2칸 이상 공백
        if "\t" in first:
            sep = "\t"
            parts0 = first.split("\t")
            headerless = _production_text_looks_headerless(parts0[0] if parts0 else "")
            df = pd.read_csv(
                io.StringIO("\n".join(lines)),
                sep=sep,
                dtype=str,
                engine="python",
                header=None if headerless else 0,
            )
        elif first.count(";") >= 2 or first.count(";") > first.count(","):
            sep = ";"
            parts0 = first.split(";")
            headerless = _production_text_looks_headerless(parts0[0] if parts0 else "")
            df = pd.read_csv(
                io.StringIO("\n".join(lines)),
                sep=sep,
                dtype=str,
                engine="python",
                header=None if headerless else 0,
            )
        elif "|" in first:
            sep = "|"
            parts0 = first.split("|")
            headerless = _production_text_looks_headerless(parts0[0] if parts0 else "")
            df = pd.read_csv(
                io.StringIO("\n".join(lines)),
                sep=sep,
                dtype=str,
                engine="python",
                header=None if headerless else 0,
            )
        elif first.count(",") >= 2:
            sep = ","
            parts0 = first.split(",")
            headerless = _production_text_looks_headerless(parts0[0] if parts0 else "")
            df = pd.read_csv(
                io.StringIO("\n".join(lines)),
                sep=sep,
                dtype=str,
                engine="python",
                header=None if headerless else 0,
            )
        else:
            # Excel/웹에서 공백 구분으로 붙는 경우
            parts0 = first.split()
            headerless = _production_text_looks_headerless(parts0[0] if parts0 else "")
            df = pd.read_csv(
                io.StringIO("\n".join(lines)),
                sep=r"\s{2,}|\t",
                dtype=str,
                engine="python",
                header=None if headerless else 0,
            )
            # 한 컬럼으로만 읽히면 단일 공백 분리 재시도
            if df.shape[1] <= 1 and " " in first:
                df = pd.read_csv(
                    io.StringIO("\n".join(lines)),
                    sep=r"\s+",
                    dtype=str,
                    engine="python",
                    header=None if headerless else 0,
                )

        df = df.fillna("")
        # 헤더 없음: 템플릿 열 이름 부여
        if headerless:
            n = df.shape[1]
            cols = list(_PRODUCTION_DEFAULT_COLS[:n])
            while len(cols) < n:
                cols.append(f"col{len(cols)}")
            df.columns = cols
        else:
            df.columns = [str(c).strip().replace("\ufeff", "") for c in df.columns]
            # 헤더가 데이터처럼 읽힌 경우 보정 (첫 컬럼명이 날짜)
            first_col = str(df.columns[0]).strip() if len(df.columns) else ""
            if _production_text_looks_headerless(first_col):
                # 컬럼명이 실제 첫 데이터 → 행으로 복원 후 기본 헤더
                row0 = [str(c).strip() for c in df.columns]
                body = df.astype(str).values.tolist()
                n = len(row0)
                cols = list(_PRODUCTION_DEFAULT_COLS[:n])
                while len(cols) < n:
                    cols.append(f"col{len(cols)}")
                df = pd.DataFrame([row0] + body, columns=cols)

        return _ingest_production_dataframe(
            df, success_message="생산 계획 텍스트를 업로드했습니다."
        )
    except Exception as e:
        return Response(
            {"message": "생산 계획 텍스트 처리 중 오류가 발생했습니다.", "error": str(e)},
            status=status.HTTP_400_BAD_REQUEST,
        )


# =====================================================
# Machine User & Plan APIs (PIN 인증 + AI 추천)
# =====================================================


def _hash_pin(pin: str) -> str:
    """PIN을 SHA-256으로 해시화"""
    return hashlib.sha256(pin.encode()).hexdigest()


@api_view(["POST"])
def machine_login(request):
    """사원번호 + PIN 로그인"""
    employee_number = request.data.get("employee_number", "").strip()
    pin = request.data.get("pin", "").strip()

    if not employee_number or not pin:
        return Response(
            {"success": False, "message": "사원번호와 PIN을 입력하세요."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # 사원번호로 사용자 조회
    user = MachineUser.objects.filter(
        employee_number=employee_number, is_active=True
    ).first()
    if not user:
        return Response(
            {"success": False, "message": "등록된 사용자가 없습니다."},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    # 잠금 체크
    if user.locked_until and user.locked_until > timezone.now():
        remaining = (user.locked_until - timezone.now()).seconds // 60
        return Response(
            {
                "success": False,
                "message": f"잠겼습니다. {remaining}분 후 재시도하세요.",
            },
            status=status.HTTP_401_UNAUTHORIZED,
        )

    # PIN 검증
    if user.user_pin != _hash_pin(pin):
        user.failed_attempts += 1
        if user.failed_attempts >= 5:
            user.locked_until = timezone.now() + timedelta(minutes=5)
            user.save()
            return Response(
                {"success": False, "message": "5회 실패로 5분간 잠겼습니다."},
                status=status.HTTP_401_UNAUTHORIZED,
            )
        user.save()
        return Response(
            {
                "success": False,
                "message": f"PIN이 올바르지 않습니다. (시도 {user.failed_attempts}/5)",
            },
            status=status.HTTP_401_UNAUTHORIZED,
        )

    # 성공: 실패 횟수 초기화
    user.failed_attempts = 0
    user.locked_until = None
    user.save()

    # 사용자 정보 + 등록된 기계 목록 반환
    user_machines = MachineUser.objects.filter(
        employee_number=employee_number, is_active=True
    ).values_list("machine_number", flat=True)

    token = f"{employee_number}:{user.id}:{timezone.now().timestamp()}"
    return Response(
        {
            "success": True,
            "token": token,
            "user_name": user.user_name,
            "employee_number": user.employee_number,
            "machines": list(user_machines),  # 사용자가 등록된 기계 목록
        }
    )


@api_view(["POST"])
def machine_logout(request):
    """로그아웃"""
    return Response({"success": True, "message": "로그아웃되었습니다."})


@api_view(["POST"])
def production_copy_day(request):
    """前一天 生产计划 复制到 指定日期"""
    from_date = request.data.get("from_date")
    to_date = request.data.get("to_date")
    machine_number = request.data.get("machine_number", "")

    if not from_date:
        return Response(
            {"success": False, "message": "from_date가 필요합니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not to_date:
        to_date = (timezone.now() + timedelta(days=1)).date().isoformat()

    # 前一天 计划查询
    plans = MachinePlan.objects.filter(date=from_date)
    if machine_number:
        plans = plans.filter(machine_number=machine_number)

    if not plans.exists():
        return Response(
            {"success": False, "message": f"{from_date}에 생산 계획이 없습니다."},
            status=status.HTTP_404_NOT_FOUND,
        )

    # 复制到新日期
    copied_plans = []
    for plan in plans:
        new_plan = MachinePlan.objects.create(
            date=datetime.fromisoformat(to_date).date(),
            machine_number=plan.machine_number,
            product_name=plan.product_name,
            product_name_eng=plan.product_name_eng,
            mold_number=plan.mold_number,
            color1=plan.color1,
            color2=plan.color2,
            unit=plan.unit,
            quantity=plan.quantity,
            unit_quantity=plan.unit_quantity,
            total=plan.total,
            status="draft",
            ai_reason=f"{from_date}에서 복사됨",
            outbound_data=plan.outbound_data,
        )
        copied_plans.append(
            {
                "product_name": new_plan.product_name,
                "quantity": new_plan.quantity,
                "unit_quantity": new_plan.unit_quantity,
                "total": new_plan.total,
                "color1": new_plan.color1,
            }
        )

    return Response(
        {
            "success": True,
            "message": f"{from_date} → {to_date} {len(copied_plans)}개 계획 복사됨",
            "copied_plans": copied_plans,
        }
    )


@api_view(["GET"])
def machine_plan_list(request):
    """기계별 생산 계획 조회"""
    date = request.query_params.get("date")
    machine_number = request.query_params.get("machine_number")

    if not date:
        # 기본: 내일
        tomorrow = (timezone.now() + timedelta(days=1)).date()
        date = tomorrow.isoformat()

    queryset = MachinePlan.objects.filter(date=date)
    if machine_number:
        queryset = queryset.filter(machine_number=machine_number)

    plans = []
    for plan in queryset:
        plans.append(
            {
                "id": plan.id,
                "date": plan.date.isoformat(),
                "machine_number": plan.machine_number,
                "user_name": plan.user.user_name if plan.user else None,
                "product_name": plan.product_name,
                "product_name_eng": plan.product_name_eng,
                "mold_number": plan.mold_number,
                "color1": plan.color1,
                "color2": plan.color2,
                "unit": plan.unit,
                "quantity": plan.quantity,
                "unit_quantity": plan.unit_quantity,
                "total": plan.total,
                "status": plan.status,
                "ai_reason": plan.ai_reason,
                "outbound_data": plan.outbound_data,
                "created_at": plan.created_at.isoformat() if plan.created_at else None,
            }
        )

    return Response({"success": True, "plans": plans, "date": date})


@api_view(["POST"])
def machine_plan_create(request):
    """생산 계획 생성 (수동 또는 AI 추천 저장)"""
    data = request.data
    date_str = data.get("date")
    machine_number = data.get("machine_number", "").strip()
    product_name = data.get("product_name", "").strip()

    if not date_str or not machine_number or not product_name:
        return Response(
            {
                "success": False,
                "message": "date, machine_number, product_name은 필수입니다.",
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        date_obj = datetime.fromisoformat(date_str).date()
    except Exception:
        return Response(
            {"success": False, "message": "유효하지 않은 날짜입니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # 중복 체크 (같은 날, 기계, 제품)
    existing = MachinePlan.objects.filter(
        date=date_obj,
        machine_number=machine_number,
        product_name=product_name,
        color1=data.get("color1", ""),
    ).first()

    if existing:
        # 기존 것 업데이트
        existing.product_name_eng = data.get("product_name_eng", "")
        existing.mold_number = data.get("mold_number", "")
        existing.color2 = data.get("color2", "")
        existing.unit = data.get("unit", "BOX")
        existing.quantity = data.get("quantity", 0)
        existing.unit_quantity = data.get("unit_quantity", 0)
        existing.total = data.get("total", 0)
        existing.status = data.get("status", "draft")
        existing.ai_reason = data.get("ai_reason", "")
        existing.outbound_data = data.get("outbound_data")
        existing.save()
        return Response(
            {
                "success": True,
                "plan": {"id": existing.id},
                "message": "계획이 업데이트되었습니다.",
            }
        )

    # 생성
    plan = MachinePlan.objects.create(
        date=date_obj,
        machine_number=machine_number,
        product_name=product_name,
        product_name_eng=data.get("product_name_eng", ""),
        mold_number=data.get("mold_number", ""),
        color1=data.get("color1", ""),
        color2=data.get("color2", ""),
        unit=data.get("unit", "BOX"),
        quantity=data.get("quantity", 0),
        unit_quantity=data.get("unit_quantity", 0),
        total=data.get("total", 0),
        status=data.get("status", "draft"),
        ai_reason=data.get("ai_reason", ""),
        outbound_data=data.get("outbound_data"),
    )
    return Response(
        {"success": True, "plan": {"id": plan.id}, "message": "계획이 생성되었습니다."}
    )


@api_view(["PUT", "DELETE"])
def machine_plan_detail(request, plan_id: int):
    """계획 수정/삭제"""
    plan = MachinePlan.objects.filter(id=plan_id).first()
    if not plan:
        return Response(
            {"success": False, "message": "계획을 찾을 수 없습니다."},
            status=status.HTTP_404_NOT_FOUND,
        )

    if request.method == "DELETE":
        plan.delete()
        return Response({"success": True, "message": "삭제되었습니다."})

    # PUT
    data = request.data
    for field in [
        "product_name",
        "product_name_eng",
        "mold_number",
        "color1",
        "color2",
        "unit",
        "quantity",
        "unit_quantity",
        "total",
        "status",
        "ai_reason",
    ]:
        if field in data:
            setattr(plan, field, data[field])

    # total 자동 계산
    plan.total = (plan.unit_quantity or 0) * (plan.quantity or 0)
    plan.save()

    return Response({"success": True, "plan": {"id": plan.id}})


@api_view(["POST"])
def machine_plan_apply(request, plan_id: int):
    """AI 추천 계획 적용 → ProductionLog에 저장"""
    plan = MachinePlan.objects.filter(id=plan_id).first()
    if not plan:
        return Response(
            {"success": False, "message": "계획을 찾을 수 없습니다."},
            status=status.HTTP_404_NOT_FOUND,
        )

    if plan.status == "applied":
        return Response(
            {"success": False, "message": "이미 적용된 계획입니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # ProductionLog에 저장
    prod_log, created = ProductionLog.objects.update_or_create(
        date=plan.date,
        machine_number=plan.machine_number,
        mold_number=plan.mold_number,
        product_name=plan.product_name,
        color1=plan.color1,
        color2=plan.color2,
        defaults={
            "product_name_eng": plan.product_name_eng,
            "unit": plan.unit,
            "quantity": plan.quantity,
            "unit_quantity": plan.unit_quantity,
            "total": plan.total,
            "status": "pending",
        },
    )

    # MachinePlan 상태 업데이트
    plan.status = "applied"
    plan.save()

    return Response(
        {
            "success": True,
            "message": "생산 계획에 적용되었습니다.",
            "production_log_id": prod_log.id,
        }
    )


@api_view(["POST"])
def ai_production_recommend(request):
    """AI 생산 계획 추천 (출고량 분석 포함)"""
    machine_number = request.data.get("machine_number", "").strip()
    product_name = request.data.get("product_name", "").strip()
    target_date = request.data.get("date")

    if not machine_number or not product_name:
        return Response(
            {
                "success": False,
                "message": "machine_number와 product_name은 필수입니다.",
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not target_date:
        target_date = (timezone.now() + timedelta(days=1)).date().isoformat()

    try:
        target_dt = datetime.fromisoformat(target_date)
    except Exception:
        target_dt = timezone.now() + timedelta(days=1)

    # 1. MasterSpec에서 제품 스펙 조회
    spec = MasterSpec.objects.filter(product_name=product_name).first()

    # 2. 해당 제품의 최근 7일 출고량 조회
    start_date = target_dt - timedelta(days=7)
    outbound_qty = (
        OutboundRecord.objects.filter(
            product_name=product_name, outbound_date__gte=start_date.date()
        ).aggregate(total=Coalesce(Sum("box_quantity"), 0))["total"]
        or 0
    )

    daily_outbound = outbound_qty / 7 if outbound_qty > 0 else 0

    # 3. 출고 추세 분석 (최근 7일 vs 이전 7일)
    prev_start = start_date - timedelta(days=7)
    prev_outbound = (
        OutboundRecord.objects.filter(
            product_name=product_name,
            outbound_date__gte=prev_start.date(),
            outbound_date__lt=start_date.date(),
        ).aggregate(total=Coalesce(Sum("box_quantity"), 0))["total"]
        or 0
    )

    trend_percent = 0
    trend_direction = "stable"
    if prev_outbound > 0:
        trend_percent = ((outbound_qty - prev_outbound) / prev_outbound) * 100
        if trend_percent > 5:
            trend_direction = "increasing"
        elif trend_percent < -5:
            trend_direction = "decreasing"

    # 4. 해당 기계의 해당 제품 생산 이력
    prod_history = ProductionLog.objects.filter(
        machine_number=machine_number, product_name=product_name
    ).order_by("-date")[:10]

    recent_qty_list = [p.quantity for p in prod_history if p.quantity]
    avg_production = (
        sum(recent_qty_list) / len(recent_qty_list) if recent_qty_list else 0
    )

    # 5. 권장 생산량 계산
    # 출고 추세 반영: 증가하면 생산량 증가, 감소하면 감소
    if trend_direction == "increasing":
        recommended_qty = int(avg_production * 1.1)  # 10% 증가
    elif trend_direction == "decreasing":
        recommended_qty = int(avg_production * 0.9)  # 10% 감소
    else:
        recommended_qty = int(avg_production)

    # 최소값 보장
    recommended_qty = max(recommended_qty, 1)

    # 단위수량은 MasterSpec 또는 기본값
    unit_quantity = spec.default_quantity if spec and spec.default_quantity > 0 else 10

    total = unit_quantity * recommended_qty

    # AI 추천 이유 생성
    reason = f"최근 7일 평균 출고 {daily_outbound:.0f}개/일"
    if trend_direction == "increasing":
        reason += f", 증가 추세 ({trend_percent:.1f}%↑)"
    elif trend_direction == "decreasing":
        reason += f", 감소 추세 ({trend_percent:.1f}%↓)"
    reason += f". 평균 생산량 {avg_production:.0f}개 기준 권장 {recommended_qty}박스."

    # 6. MachinePlan에 저장 (recommended 상태)
    tomorrow = (timezone.now() + timedelta(days=1)).date()
    plan = MachinePlan.objects.create(
        date=tomorrow,
        machine_number=machine_number,
        product_name=product_name,
        product_name_eng=spec.product_name_eng if spec else "",
        mold_number=spec.mold_number if spec else "",
        color1=spec.color1 if spec else "",
        color2=spec.color2 if spec else "",
        unit="BOX",
        quantity=recommended_qty,
        unit_quantity=unit_quantity,
        total=total,
        status="recommended",
        ai_reason=reason,
        outbound_data={
            "daily_outbound": round(daily_outbound, 1),
            "trend_percent": round(trend_percent, 1),
            "trend_direction": trend_direction,
            "avg_production": round(avg_production, 1),
            "recent_qty_list": recent_qty_list[:5],
        },
    )

    return Response(
        {
            "success": True,
            "recommendation": {
                "plan_id": plan.id,
                "product_name": product_name,
                "color1": plan.color1,
                "unit_quantity": unit_quantity,
                "quantity": recommended_qty,
                "total": total,
                "reason": reason,
                "outbound_data": plan.outbound_data,
            },
        }
    )


@api_view(["POST"])
def ai_production_chat(request):
    """AI 생산 계획 챗 - 자연어로 계획 생성/조회"""
    message = (request.data.get("message") or "").strip()
    machine_number = request.data.get("machine_number", "").strip()
    target_date = request.data.get("date")

    if not message:
        return Response(
            {"success": False, "message": "메시지를 입력하세요."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # machine_number이 없으면 메시지에서 추출 시도
    if not machine_number:
        import re

        machine_pattern = r"(?:(\d+)번|M0*(\d+))"
        machine_match = re.search(machine_pattern, message)
        if machine_match:
            extracted_machine = machine_match.group(1) or machine_match.group(2)
            if extracted_machine:
                machine_number = f"M{extracted_machine.zfill(3)}"

    # 그래도 없으면 빈 문자열로 진행 (기타 작업 - 기계 없는 생산)

    # 기본 날짜: 내일
    if not target_date:
        target_date = (timezone.now() + timedelta(days=1)).date().isoformat()

    # 완료 안 된 계획 자동 이동
    # 이전 날짜 계획 중 미완료 → 오늘로
    # 오늘 계획 중 미완료 → 내일(다음 작업일)로
    today = timezone.localdate()
    tomorrow_date = today + timedelta(days=1)
    # 주말이면 다음 평일
    while tomorrow_date.weekday() >= 5:
        tomorrow_date += timedelta(days=1)

    # 이전 날짜 미완료 계획 → 오늘로 이동
    old_incomplete = MachinePlan.objects.filter(
        date__lt=today, status__in=["pending", "started", "recommended"]
    )
    moved_to_today = old_incomplete.update(date=today)

    # 오늘 미완료 계획 → 내일로 이동
    today_incomplete = MachinePlan.objects.filter(
        date=today, status__in=["pending", "started", "recommended"]
    )
    moved_to_tomorrow = today_incomplete.update(date=tomorrow_date)

    if moved_to_today or moved_to_tomorrow:
        return Response(
            {
                "success": True,
                "action": "moved",
                "message": f"{moved_to_today}개 → 오늘({today})로, {moved_to_tomorrow}개 → 내일({tomorrow_date})로 이동됨",
            }
        )

    # 메시지 파싱 (간단한 패턴 매칭)
    import re

    # 기계번호 추출
    machine_pattern = r"(?:(\d+)번|M0*(\d+))"
    machine_match = re.search(machine_pattern, message)
    if machine_match:
        extracted_machine = machine_match.group(1) or machine_match.group(2)
        if extracted_machine:
            machine_number = f"M{extracted_machine.zfill(3)}"

    # 색상 추출 - 먼저 초기화
    extracted_color = None

    # 색상 추출 - 제품명 앞에 있는 색상을 먼저 찾음
    color_patterns = [
        "아이보리",
        "화이트",
        "브라운",
        "레드",
        "블루",
        "그린",
        "옐로",
        "퍼플",
        "오렌지",
        "핑크",
        "블랙",
        "그레이",
        "ivory",
        "white",
        "brown",
        "red",
        "blue",
        "green",
        "yellow",
    ]
    for cp in color_patterns:
        if cp.lower() in message.lower():
            # 정확한 색상명 매칭
            if cp == "ivory":
                extracted_color = "아이보리"
            elif cp == "white":
                extracted_color = "화이트"
            elif cp == "brown":
                extracted_color = "브라운"
            elif cp == "red":
                extracted_color = "레드"
            elif cp == "blue":
                extracted_color = "블루"
            elif cp == "green":
                extracted_color = "그린"
            elif cp == "yellow":
                extracted_color = "옐로"
            elif cp == "purple":
                extracted_color = "퍼플"
            elif cp == "orange":
                extracted_color = "오렌지"
            elif cp == "pink":
                extracted_color = "핑크"
            elif cp == "black":
                extracted_color = "블랙"
            elif cp == "gray":
                extracted_color = "그레이"
            else:
                extracted_color = cp
            break

    # 수량 단위 변환: 팔레트/박스/EA
    pallet_match = re.search(r"(\d+)\s*팔레트", message)
    box_match = re.search(r"(\d+)\s*박스", message)
    qty_match = re.search(r"(\d+)\s*(?:개|EA)", message)

    if "팔레트" in message or "pallet" in message.lower():
        if pallet_match:
            quantity = int(pallet_match.group(1)) * 125
        else:
            quantity = 125  # 기본 1팔레트
    elif "박스" in message:
        quantity = int(box_match.group(1)) if box_match else 10
    elif qty_match:
        quantity = int(qty_match.group(1))
    else:
        quantity = 125  # 기본값 1팔레트

    # 제품명 추출 - MasterSpec에서 직접 검색
    product_name = None

    # 알려진 제품명 패턴들
    known_products = [
        "토이 아이보리",
        "toy ivory",
        "ivory",
        "로코스",
        "locs",
        "헬로키티",
        "hello kitty",
        "hello",
        "리리카",
        "lilica",
        "마이멜로디",
        "my melody",
        "쿠퍼",
        "kupi",
        "바니",
        "bani",
    ]

    # message에서 알려진 제품명 찾기
    msg_lower = message.lower()
    for p in known_products:
        if p.lower() in msg_lower:
            if p == "toy ivory" or p == "ivory":
                product_name = "토이 아이보리"
            elif p == "locs":
                product_name = "로코스 L"
            elif p == "hello kitty" or p == "hello":
                product_name = "헬로키티"
            else:
                product_name = p
            break

    # 기존 정규식 시도 (알려진 제품이 없을 때)
    if not product_name:
        product_match = re.search(
            r"([가-힣a-zA-Z0-9]+)(?:\s*(?:추가|생산|등록|넣어|주|해주세요|pallet|박스|EA)|$)",
            message,
        )
        product_name = product_match.group(1) if product_match else None

    if product_name:
        # 제품명이 있으면 계획 생성 시도
        # 1순위: ProductionLog (production 데이터優先)
        prod_data = (
            ProductionLog.objects.filter(product_name__icontains=product_name)
            .values("product_name", "color1", "color2", "mold_number", "unit_quantity")
            .first()
        )

        # 2순위: MasterSpec (대체)
        if not prod_data:
            spec = MasterSpec.objects.filter(
                product_name__icontains=product_name
            ).first()
            if not spec:
                specs = MasterSpec.objects.filter(product_name__icontains=product_name)
                if specs.exists():
                    spec = specs.first()
            if spec:
                prod_data = {
                    "product_name": spec.product_name,
                    "color1": spec.color1 or "",
                    "color2": spec.color2 or "",
                    "mold_number": spec.mold_number or "",
                    "unit_quantity": spec.default_quantity
                    if spec.default_quantity > 0
                    else 10,
                }

        if prod_data:
            unit_quantity = prod_data.get("unit_quantity") or 10
            # 메시지에서 색상을 추출했으면 그것을 우선 사용
            if extracted_color:
                color1 = extracted_color
            else:
                color1 = prod_data.get("color1") or ""
            color2 = prod_data.get("color2") or ""
            mold_number = prod_data.get("mold_number") or ""
        else:
            unit_quantity = 10
            color1 = extracted_color or ""
            color2 = ""
            mold_number = ""

        # 색상 clarification 필요 여부 확인 (색상을 추출하지 못했을 때만)
        clarification_needed = []
        if not color1:  # 색상이 없으면 확인
            if product_name and (
                "로코스" in product_name.lower() or "locs" in product_name.lower()
            ):
                # 로코스는 색상 확인 필요
                clarification_needed.append(
                    {
                        "type": "color",
                        "question": "로코스의 색상은 화이트와 아이보리 중 어느 색상인가요?",
                        "options": ["화이트", "아이보리"],
                    }
                )
            elif product_name and (
                "토이" in product_name.lower() or "toy" in product_name.lower()
            ):
                # 토이 系列은 색상 확인 필요
                clarification_needed.append(
                    {
                        "type": "color",
                        "question": "토이 시리즈의 색상은 화이트, 아이보리, 브라운 중 어느 색상인가요?",
                        "options": ["화이트", "아이보리", "브라운"],
                    }
                )

        # 출고량 분석
        outbound_qty = (
            OutboundRecord.objects.filter(
                product_name__icontains=product_name,
                outbound_date__gte=timezone.localdate() - timedelta(days=7),
            ).aggregate(total=Coalesce(Sum("box_quantity"), 0))["total"]
            or 0
        )

        daily_outbound = outbound_qty / 7

        # 권장 수량 (출고량 기반)
        recommended_qty = max(int(daily_outbound * 1.1), quantity)
        total = recommended_qty * unit_quantity

        # 색상 확인이 필요하면 clarificationNeeded 포함 후 생성
        if clarification_needed:
            # clarificationNeeded만是先返回，不 直接 생성
            return Response(
                {
                    "success": True,
                    "action": "clarification",
                    "message": f'"{product_name}"의 색상을 선택해주세요.',
                    "clarification_needed": clarification_needed,
                    "default_values": {
                        "product_name": product_name,
                        "quantity": quantity,
                        "unit_quantity": unit_quantity,
                        "machine_number": machine_number,
                        "date": target_date,
                    },
                }
            )

        # 계획 생성
        plan = MachinePlan.objects.create(
            date=datetime.fromisoformat(target_date).date(),
            machine_number=machine_number,
            product_name=spec.product_name if spec else product_name,
            product_name_eng=spec.product_name_eng if spec else "",
            mold_number=mold_number,
            color1=color1,
            color2=color2,
            unit="BOX",
            quantity=recommended_qty,
            unit_quantity=unit_quantity,
            total=total,
            status="recommended",
            ai_reason=f"AI 챗 생성: 최근 7일 평균 출고 {daily_outbound:.0f}개/일",
            outbound_data={
                "daily_outbound": round(daily_outbound, 1),
                "source": "chat",
            },
        )

        return Response(
            {
                "success": True,
                "action": "created",
                "message": f'"{plan.product_name}" 계획을 생성했습니다.',
                "plan": {
                    "id": plan.id,
                    "product_name": plan.product_name,
                    "quantity": plan.quantity,
                    "unit_quantity": plan.unit_quantity,
                    "total": plan.total,
                    "color1": plan.color1,
                    "color2": plan.color2,
                },
                "outbound_info": {
                    "daily_outbound": round(daily_outbound, 1),
                    "recommended": total,
                },
            }
        )
    else:
        # 제품명이 없으면 현재 계획 조회
        plans = MachinePlan.objects.filter(
            machine_number=machine_number, date=target_date
        )[:5]

        if plans:
            plan_list = "\n".join(
                [
                    f"- {p.product_name}: {p.quantity}박스 × {p.unit_quantity}개 = {p.total}개 ({p.status})"
                    for p in plans
                ]
            )
            return Response(
                {
                    "success": True,
                    "action": "info",
                    "response": f"[{machine_number}] {target_date} 일정:\n{plan_list}",
                }
            )
        else:
            return Response(
                {
                    "success": True,
                    "action": "info",
                    "response": f'[{machine_number}] {target_date} 일정이 비어있습니다.\n"~에 ~추가해줘"라고 말씀하시면 계획을 추가합니다.',
                }
            )


@api_view(["GET"])
def machine_user_list(request):
    """기계별 사용자 목록 조회 (사원번호 필터 지원)"""
    machine_number = request.query_params.get("machine_number")
    employee_number = request.query_params.get("employee_number")

    users = MachineUser.objects.filter(is_active=True)
    if machine_number:
        users = users.filter(machine_number=machine_number)
    if employee_number:
        users = users.filter(employee_number=employee_number)

    return Response(
        {
            "success": True,
            "users": [
                {
                    "id": u.id,
                    "machine_number": u.machine_number,
                    "employee_number": u.employee_number,
                    "user_name": u.user_name,
                }
                for u in users
            ],
        }
    )


@api_view(["POST"])
def machine_user_create(request):
    """기계 사용자 추가 (관리자용) - 사원번호 기반"""
    employee_number = request.data.get("employee_number", "").strip()
    machine_number = request.data.get("machine_number", "").strip()
    user_name = request.data.get("user_name", "").strip()
    pin = request.data.get("pin", "").strip()

    if not employee_number or not machine_number or not user_name or not pin:
        return Response(
            {
                "success": False,
                "message": "employee_number, machine_number, user_name, pin은 필수입니다.",
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    if len(pin) < 4 or len(pin) > 6:
        return Response(
            {"success": False, "message": "PIN은 4~6자리입니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # 사원번호로 기존 사용자 조회 (같은 사람일 경우)
    existing_user = MachineUser.objects.filter(employee_number=employee_number).first()
    if existing_user:
        # 같은 사원번호면 기계만 추가 (복수 기계 가능)
        if MachineUser.objects.filter(
            employee_number=employee_number, machine_number=machine_number
        ).exists():
            return Response(
                {"success": False, "message": "이미 해당 기계에 등록된 사용자입니다."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = MachineUser.objects.create(
            employee_number=employee_number,
            machine_number=machine_number,
            user_name=user_name,
            user_pin=_hash_pin(pin),
            is_active=True,
        )
    else:
        # 새 사용자 생성
        user = MachineUser.objects.create(
            employee_number=employee_number,
            machine_number=machine_number,
            user_name=user_name,
            user_pin=_hash_pin(pin),
            is_active=True,
        )

    return Response(
        {
            "success": True,
            "user": {
                "id": user.id,
                "employee_number": user.employee_number,
                "machine_number": user.machine_number,
                "user_name": user.user_name,
            },
        }
    )


@api_view(["POST"])
def bulk_create_inventory(request):
    items = request.data
    if not isinstance(items, list):
        return Response(
            {"error": "Expected a list of items"}, status=status.HTTP_400_BAD_REQUEST
        )

    created_items = []
    for item_data in items:
        # Simple upsert logic based on name (or other unique key if available)
        # For now, we just create new ones or update if ID is provided
        try:
            if "id" in item_data and item_data["id"]:
                item, created = InventoryItem.objects.update_or_create(
                    id=item_data["id"], defaults=item_data
                )
            else:
                # If no ID, create new
                serializer = InventoryItemSerializer(data=item_data)
                if serializer.is_valid():
                    serializer.save()
                    created_items.append(serializer.data)
        except Exception as e:
            print(f"Error processing item: {e}")
            continue

    return Response(
        {"message": f"Processed {len(items)} items"}, status=status.HTTP_201_CREATED
    )


from django.db import transaction


@api_view(["POST"])
def bulk_create_outbound(request):
    records = request.data
    if not isinstance(records, list):
        return Response(
            {"error": "Expected a list of records"}, status=status.HTTP_400_BAD_REQUEST
        )

    outbound_instances = []
    errors = []

    for record_data in records:
        try:
            # Manual mapping for speed and bulk_create compatibility
            # Assuming input keys match serializer fields exactly or close enough

            # Handle dates
            outbound_date = record_data.get("outbound_date")
            if not outbound_date:
                continue  # Skip invalid dates

            instance = OutboundRecord(
                id=record_data.get("id") or str(uuid.uuid4()),
                product_name=record_data.get("product_name"),
                category=record_data.get("category"),
                quantity=record_data.get("quantity", 0),
                sales_amount=record_data.get("sales_amount", 0),
                outbound_date=outbound_date,
                status=record_data.get("status", "완료"),
                barcode=record_data.get("barcode"),
                box_quantity=record_data.get("box_quantity"),
                unit_count=record_data.get("unit_count"),
                notes=record_data.get("notes"),
                client=record_data.get("client") or "",
            )
            if len(outbound_instances) < 5:
                print(
                    f"DEBUG BACKEND: Date={outbound_date}, SalesAmountInput={record_data.get('sales_amount')}, InstanceAmount={instance.sales_amount}"
                )
            outbound_instances.append(instance)
        except Exception as e:
            errors.append(str(e))
            if len(errors) > 10:  # Don't flood logs
                break

    if outbound_instances:
        # Fallback to simple loop since bulk_create is failing with SQLite driver issues
        # Use transaction to speed up loop
        total_created = 0
        try:
            with transaction.atomic():
                for instance in outbound_instances:
                    try:
                        instance.save()
                        total_created += 1
                    except Exception as e:
                        errors.append(str(e))
                        if len(errors) > 10:
                            break
        except Exception as e:
            errors.append(f"Transaction failed: {str(e)}")

    return Response(
        {
            "message": f"Successfully created {total_created} records",
            "errors_sample": errors[:5] if errors else [],
        },
        status=status.HTTP_201_CREATED,
    )


@api_view(["DELETE"])
def delete_outbound_by_date(request):
    """
    기간 출고 삭제 (파괴적). 실서비스 보호:
    - confirm=true 필수
    - 기간 최대 31일
    - DESTRUCTIVE_API_KEY 설정 시 키 필수
    - 프론트 UI에서 호출하지 않음 (관리/스크립트용)
    """
    from .api_guards import destructive_guard_response

    blocked = destructive_guard_response(request)
    if blocked is not None:
        return blocked

    start = request.query_params.get("start")
    end = request.query_params.get("end")
    confirm = (request.query_params.get("confirm") or "").strip().lower()

    if not start or not end:
        return Response(
            {"error": "Start and end dates are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if confirm not in ("1", "true", "yes"):
        return Response(
            {
                "error": "confirm=true required for destructive delete",
                "hint": "DELETE /api/outbound/delete-range?start=YYYY-MM-DD&end=YYYY-MM-DD&confirm=true",
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    start_d = _parse_date_ymd(start)
    end_d = _parse_date_ymd(end)
    if not start_d or not end_d:
        return Response(
            {"error": "Invalid start/end date (YYYY-MM-DD)"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if end_d < start_d:
        return Response(
            {"error": "end must be >= start"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if (end_d - start_d).days > 31:
        return Response(
            {"error": "date range too large (max 31 days)"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        logger.warning(
            "delete_outbound_by_date start=%s end=%s",
            start_d.isoformat(),
            end_d.isoformat(),
        )
        count, _ = OutboundRecord.objects.filter(
            outbound_date__range=[start_d, end_d]
        ).delete()
        return Response(
            {
                "message": f"Deleted {count} records between {start_d} and {end_d}",
                "deleted": count,
                "start": start_d.isoformat(),
                "end": end_d.isoformat(),
            }
        )
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def outbound_sync(request):
    url = None
    upload_date = None
    start_date = None
    end_date = None

    if isinstance(request.data, dict):
        url = request.data.get("url")
        upload_date = request.data.get("date")
        start_date = request.data.get("start_date")
        end_date = request.data.get("end_date")

    url = url or os.environ.get("OUTBOUND_GOOGLE_SHEET_URL")

    if not url:
        return Response(
            {"error": "OUTBOUND_GOOGLE_SHEET_URL is not set and no url was provided"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def parse_num(val):
        if val is None:
            return 1
        s = str(val).strip()
        if s == "":
            return 1
        s = s.replace(",", "")
        try:
            num = float(s)
            if num <= 0:
                return 1
            return num
        except Exception:
            return 1

    def parse_date(val):
        s = ("" if val is None else str(val)).strip()
        if not s:
            return None
        for fmt in (
            "%Y-%m-%d",
            "%Y.%m.%d",
            "%Y/%m/%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
        ):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        try:
            dt = pd.to_datetime(s, errors="coerce")
            if pd.isna(dt):
                return None
            return dt.date()
        except Exception:
            return None

    try:
        if url.startswith("http"):
            import requests
            import io

            r = requests.get(url, timeout=120)
            r.raise_for_status()
            encodings = ["utf-8-sig", "cp949", "euc-kr", "utf-8"]
            decoded = None
            for encoding in encodings:
                try:
                    decoded = r.content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if decoded is None:
                raise ValueError("Failed to decode CSV with any supported encoding")
            df = pd.read_csv(io.StringIO(decoded), dtype=str).fillna("")
        else:
            df = pd.read_csv(url, dtype=str, encoding="utf-8-sig").fillna("")

        df.columns = [str(c).strip().lstrip("\ufeff") for c in df.columns]
    except Exception as e:
        return Response(
            {"error": f"Failed to fetch/parse CSV: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if df.empty:
        return Response(
            {"success": True, "synced": 0, "message": "No rows found in sheet"},
            status=status.HTTP_200_OK,
        )

    cols = [str(c).strip() for c in df.columns]

    def find_col(candidates):
        for cand in candidates:
            for c in cols:
                if cand in c:
                    return c
        return None

    date_col = find_col(["일자", "출고일", "date"])
    product_col = find_col(["품목", "상품명", "product"])
    category_col = find_col(["분류", "카테고리", "category"])
    barcode_col = find_col(["바코드", "barcode"])
    box_col = find_col(["수량(박스)", "박스"])
    unit_col = find_col(["수량(낱개)", "낱개"])
    amount_col = find_col(["판매금액", "금액", "매출", "amount"])
    notes_col = find_col(["비고", "메모", "note"])
    client_col = find_col(["거래처", "고객", "client"])

    if not date_col or not product_col:
        return Response(
            {"error": "Required columns not found", "columns": cols},
            status=status.HTTP_400_BAD_REQUEST,
        )

    records = []
    dates = []
    now = timezone.now()

    # Date range filter
    filter_start = parse_date(start_date) if start_date else None
    filter_end = parse_date(end_date) if end_date else None
    # Fallback to upload_date if only one date provided
    if upload_date and not filter_start and not filter_end:
        filter_start = parse_date(upload_date)
        filter_end = filter_start

    batch_num = 0
    batch_size = 10000

    for _, row in df.iterrows():
        outbound_date = parse_date(row.get(date_col))
        if not outbound_date:
            continue

        # Apply date range filter
        if filter_start and outbound_date < filter_start:
            continue
        if filter_end and outbound_date > filter_end:
            continue

        product_name = str(row.get(product_col) or "").strip()
        barcode = str(row.get(barcode_col) or "").strip() if barcode_col else ""
        if not product_name and not barcode:
            continue

        box_qty = int(parse_num(row.get(box_col))) if box_col else 1
        if box_qty <= 0:
            box_qty = 1
        unit_qty = int(parse_num(row.get(unit_col))) if unit_col else 0
        sales_amount = parse_num(row.get(amount_col)) if amount_col else 0

        category = str(row.get(category_col) or "").strip() if category_col else ""
        client = str(row.get(client_col) or "").strip() if client_col else ""
        notes = str(row.get(notes_col) or "").strip() if notes_col else ""

        records.append(
            OutboundRecord(
                id=str(uuid.uuid4()),
                outbound_date=outbound_date,
                product_name=product_name or "-",
                category=category or "기타",
                barcode=barcode or None,
                quantity=box_qty,
                box_quantity=box_qty,
                unit_count=unit_qty,
                sales_amount=sales_amount,
                client=client,
                status="완료",
                notes=notes or None,
                is_estimated=False,
                estimate_method="",
                created_at=now,
                updated_at=now,
            )
        )
        dates.append(outbound_date)

    if not records:
        return Response(
            {"success": True, "synced": 0, "message": "No valid records found"},
            status=status.HTTP_200_OK,
        )

    start = min(dates) if dates else None
    end = max(dates) if dates else None

    # No date range? Use filter dates
    if not start:
        start = filter_start if filter_start else (timezone.localdate() - timedelta(days=30))
    if not end:
        end = filter_end if filter_end else timezone.localdate()

    from django.db import transaction

    total_created = len(records)
    total_updated = 0
    total_deleted = 0

    try:
        with transaction.atomic():
            # Delete all existing outbound records within the date range of the current sheet
            deleted_count, _ = OutboundRecord.objects.filter(
                outbound_date__range=[start, end]
            ).delete()
            total_deleted = deleted_count

            # Bulk create all parsed records in batches of 5000
            OutboundRecord.objects.bulk_create(records, batch_size=5000)
    except Exception as e:
        logger.error(f"Sync error: {e}")
        return Response(
            {"error": f"Database sync failed: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    created = total_created
    updated = total_updated
    deleted = total_deleted

    try:
        DataSource.objects.update_or_create(
            type="google_sheets",
            name="Google Sheets - outbound",
            defaults={
                "url": url,
                "is_active": True,
                "last_sync": timezone.now(),
                "sync_data": {
                    "source": "outbound_sync",
                    "start": start.isoformat(),
                    "end": end.isoformat(),
                    "deleted": deleted,
                    "created": created,
                    "updated": updated,
                },
            },
        )
    except Exception:
        pass

    return Response(
        {
            "success": True,
            "url": url,
            "start": start.isoformat(),
            "end": end.isoformat(),
            "deleted": deleted,
            "created": created,
            "updated": updated,
            "synced": created + updated,
            "timestamp": timezone.now().isoformat(),
        }
    )


def _normalize_header(value: str) -> str:
    return (
        (value or "")
        .strip()
        .lower()
        .replace(" ", "")
        .replace("\t", "")
        .replace("\n", "")
    )


def _find_header(headers, candidates):
    normalized = [_normalize_header(h) for h in headers]
    for cand in candidates:
        nc = _normalize_header(cand)
        for idx, h in enumerate(normalized):
            if nc and nc in h:
                return idx
    return None


def _parse_uploaded_csv(file_obj):
    raw = file_obj.read()
    try:
        text = raw.decode("utf-8-sig")
    except Exception:
        text = raw.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return [], []
    headers = [c.strip() for c in rows[0]]
    body = [[c.strip() for c in r] for r in rows[1:]]
    return headers, body


def _process_inventory_csv_rows(headers, rows, now):
    name_idx = _find_header(headers, ["name", "상품명", "품목"])
    cat_idx = _find_header(headers, ["category", "카테고리", "분류"])
    stock_idx = _find_header(headers, ["current_stock", "stock", "재고"])
    min_idx = _find_header(headers, ["minimum_stock", "min", "최소"])
    barcode_idx = _find_header(headers, ["barcode", "바코드"])

    if name_idx is None:
        raise ValueError("CSV 헤더에 상품명이 필요합니다.")

    processed = 0
    for r in rows:
        name = (
            r[name_idx] if name_idx is not None and name_idx < len(r) else ""
        ).strip()
        if not name:
            continue
        category = (
            r[cat_idx] if cat_idx is not None and cat_idx < len(r) else "기타"
        ).strip() or "기타"
        current_stock = int(
            float(
                (r[stock_idx] if stock_idx is not None and stock_idx < len(r) else 0)
                or 0
            )
        )
        minimum_stock = int(
            float((r[min_idx] if min_idx is not None and min_idx < len(r) else 0) or 0)
        )
        barcode = (
            r[barcode_idx] if barcode_idx is not None and barcode_idx < len(r) else ""
        ).strip() or None

        InventoryItem.objects.update_or_create(
            name=name,
            defaults={
                "category": category,
                "current_stock": current_stock,
                "minimum_stock": minimum_stock,
                "barcode": barcode,
                "updated_at": now,
            },
        )
        processed += 1

    return processed


def _process_outbound_csv_rows(headers, rows, now):
    date_idx = _find_header(headers, ["outbound_date", "date", "일자", "출고일"])
    product_idx = _find_header(headers, ["product_name", "product", "품목", "상품명"])
    category_idx = _find_header(headers, ["category", "분류", "카테고리"])
    barcode_idx = _find_header(headers, ["barcode", "바코드"])
    box_idx = _find_header(headers, ["수량(박스)", "box", "박스"])
    unit_idx = _find_header(headers, ["수량(낱개)", "unit", "낱개"])
    amount_idx = _find_header(
        headers, ["판매금액", "sales_amount", "amount", "금액", "매출"]
    )
    notes_idx = _find_header(headers, ["비고", "notes", "메모"])
    client_idx = _find_header(headers, ["거래처", "client", "고객"])

    if date_idx is None or product_idx is None:
        raise ValueError("CSV 헤더에 일자/상품명이 필요합니다.")

    def parse_num(val):
        s = ("" if val is None else str(val)).strip().replace(",", "")
        if not s:
            return 1  # 기본값 1로 변경
        try:
            num = float(s)
            if num <= 0:
                return 1
            elif num > 1000:  # 너무 큰 값 제한
                return min(num, 1000)
            return num
        except Exception:
            return 1  # 변환 실패시 1로

    def parse_date(val):
        s = ("" if val is None else str(val)).strip()
        if not s:
            return None
        for fmt in (
            "%Y-%m-%d",
            "%Y.%m.%d",
            "%Y/%m/%d",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
        ):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        try:
            dt = pd.to_datetime(s, errors="coerce")
            if pd.isna(dt):
                return None
            return dt.date()
        except Exception:
            return None

    instances = []
    dates = []
    for r in rows:
        outbound_date = parse_date(r[date_idx] if date_idx < len(r) else None)
        if not outbound_date:
            continue
        product_name = (r[product_idx] if product_idx < len(r) else "").strip()
        barcode = (
            r[barcode_idx] if barcode_idx is not None and barcode_idx < len(r) else ""
        ).strip() or None
        if not product_name and not barcode:
            continue

        category = (
            r[category_idx]
            if category_idx is not None and category_idx < len(r)
            else ""
        ).strip() or "기타"
        box_qty = int(
            parse_num(r[box_idx] if box_idx is not None and box_idx < len(r) else 0)
        )
        unit_qty = int(
            parse_num(r[unit_idx] if unit_idx is not None and unit_idx < len(r) else 0)
        )
        sales_amount = parse_num(
            r[amount_idx] if amount_idx is not None and amount_idx < len(r) else 0
        )
        notes = (
            r[notes_idx] if notes_idx is not None and notes_idx < len(r) else ""
        ).strip() or None
        client = (
            r[client_idx] if client_idx is not None and client_idx < len(r) else ""
        ).strip()

        instances.append(
            OutboundRecord(
                id=str(uuid.uuid4()),
                outbound_date=outbound_date,
                product_name=product_name or "-",
                category=category,
                barcode=barcode,
                quantity=box_qty,
                box_quantity=box_qty,
                unit_count=unit_qty,
                sales_amount=sales_amount,
                client=client,
                status="완료",
                notes=notes,
                created_at=now,
                updated_at=now,
            )
        )
        dates.append(outbound_date)

    if not instances:
        return 0

    start = min(dates)
    end = max(dates)

    from django.db import transaction

    with transaction.atomic():
        OutboundRecord.objects.filter(outbound_date__range=[start, end]).delete()
        try:
            OutboundRecord.objects.bulk_create(instances, batch_size=5000)
        except Exception:
            for inst in instances:
                inst.save()

        # 신규 출고 품목 자동 복구: 3개월 미출고만 해제 (단종은 수동 유지)
        synced_barcodes = {inst.barcode for inst in instances if inst.barcode}
        if synced_barcodes:
            _clear_no_outbound_3m_on_activity(barcodes=synced_barcodes)

    return len(instances)


@api_view(["POST"])
def upload_csv(request):
    if "csv" not in request.FILES:
        return Response(
            {"message": "CSV 파일이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST
        )

    data_type = (request.data.get("type") or "").strip()
    if data_type not in ("inventory", "outbound"):
        return Response(
            {"message": "올바른 데이터 타입을 선택해주세요."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    file_obj = request.FILES["csv"]
    headers, rows = _parse_uploaded_csv(file_obj)
    if not headers or not rows:
        return Response(
            {"message": "빈 CSV 파일입니다."}, status=status.HTTP_400_BAD_REQUEST
        )

    now = timezone.now()
    try:
        if data_type == "inventory":
            rows_processed = _process_inventory_csv_rows(headers, rows, now)
        else:
            rows_processed = _process_outbound_csv_rows(headers, rows, now)
    except ValueError as e:
        return Response({"message": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    DataSource.objects.create(
        type="csv",
        name=getattr(file_obj, "name", "")
        or getattr(file_obj, "original_name", "")
        or "csv",
        is_active=True,
        last_sync=now,
        sync_data={"headers": headers, "rowsProcessed": rows_processed},
    )

    return Response(
        {
            "message": "CSV 파일이 성공적으로 처리되었습니다.",
            "rowsProcessed": rows_processed,
        }
    )


@api_view(["POST"])
def google_sheets_connect(request):
    url = (request.data.get("url") or "").strip()
    data_type = (request.data.get("type") or "").strip()
    if not url:
        return Response(
            {"message": "구글 시트 URL이 필요합니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if data_type not in ("inventory", "outbound"):
        return Response(
            {"message": "올바른 데이터 타입을 선택해주세요."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    m = None
    try:
        import re

        m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    except Exception:
        m = None
    if not m:
        return Response(
            {"message": "올바른 구글 시트 URL이 아닙니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    sheet_id = m.group(1)
    api_key = (
        os.environ.get("GOOGLE_SHEETS_API_KEY")
        or os.environ.get("GOOGLE_API_KEY")
        or ""
    )
    if not api_key:
        return Response(
            {"message": "Google Sheets API 키가 설정되지 않았습니다."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    range_name = "A:Z"
    api_url = f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{range_name}?key={api_key}"

    try:
        with urllib.request.urlopen(api_url) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        return Response(
            {"message": "구글 시트에서 데이터를 가져올 수 없습니다.", "error": str(e)},
            status=status.HTTP_400_BAD_REQUEST,
        )

    values = payload.get("values") or []
    if not values:
        return Response(
            {"message": "구글 시트에 데이터가 없습니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    headers = [str(c).strip() for c in (values[0] or [])]
    rows = [[str(c).strip() for c in r] for r in values[1:]]

    now = timezone.now()
    try:
        if data_type == "inventory":
            rows_processed = _process_inventory_csv_rows(headers, rows, now)
        else:
            rows_processed = _process_outbound_csv_rows(headers, rows, now)
    except ValueError as e:
        return Response({"message": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    ds, _created = DataSource.objects.update_or_create(
        type="google_sheets",
        name=f"Google Sheets - {data_type}",
        defaults={
            "url": url,
            "is_active": True,
            "last_sync": now,
            "sync_data": payload,
        },
    )

    return Response(
        {
            "message": "구글 시트가 성공적으로 연결되었습니다.",
            "rowsProcessed": rows_processed,
            "dataSource": {
                "id": str(ds.id),
                "type": ds.type,
                "name": ds.name,
                "url": ds.url,
                "isActive": ds.is_active,
                "lastSync": ds.last_sync.isoformat() if ds.last_sync else None,
            },
        }
    )


@api_view(["POST"])
def google_sheets_refresh(request, id: str):
    # existing refresh logic unchanged
    ...
    try:
        ds = DataSource.objects.get(id=id)
    except Exception:
        return Response(
            {"message": "구글 시트 데이터 소스를 찾을 수 없습니다."},
            status=status.HTTP_404_NOT_FOUND,
        )

    if ds.type != "google_sheets" or not ds.url:
        return Response(
            {"message": "구글 시트 데이터 소스를 찾을 수 없습니다."},
            status=status.HTTP_404_NOT_FOUND,
        )

    data_type = "inventory" if "inventory" in (ds.name or "").lower() else "outbound"

    now = timezone.now()
    try:
        import re

        m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", ds.url)
    except Exception:
        m = None
    if not m:
        return Response(
            {"message": "올바른 구글 시트 URL이 아닙니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    sheet_id = m.group(1)
    api_key = (
        os.environ.get("GOOGLE_SHEETS_API_KEY")
        or os.environ.get("GOOGLE_API_KEY")
        or ""
    )
    if not api_key:
        return Response(
            {"message": "Google Sheets API 키가 설정되지 않았습니다."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    range_name = "A:Z"
    api_url = f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/{range_name}?key={api_key}"

    try:
        with urllib.request.urlopen(api_url) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        return Response(
            {"message": "구글 시트에서 데이터를 가져올 수 없습니다.", "error": str(e)},
            status=status.HTTP_400_BAD_REQUEST,
        )

    values = payload.get("values") or []
    if not values:
        return Response(
            {"message": "구글 시트에 데이터가 없습니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    headers = [str(c).strip() for c in (values[0] or [])]
    rows = [[str(c).strip() for c in r] for r in values[1:]]

    try:
        if data_type == "inventory":
            rows_processed = _process_inventory_csv_rows(headers, rows, now)
        else:
            rows_processed = _process_outbound_csv_rows(headers, rows, now)
    except ValueError as e:
        return Response({"message": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    ds.last_sync = now
    ds.is_active = True
    ds.sync_data = payload
    ds.save(update_fields=["last_sync", "is_active", "sync_data"])

    return Response(
        {
            "message": "데이터가 성공적으로 새로고침되었습니다.",
            "lastSync": ds.last_sync.isoformat(),
            "rowsProcessed": rows_processed,
        }
    )


@api_view(["GET"])
def outbound_date_range(request):
    meta = OutboundRecord.objects.aggregate(
        earliestDate=Min("outbound_date"),
        latestDate=Max("outbound_date"),
        totalRecords=Count("id"),
    )
    earliest = meta.get("earliestDate")
    latest = meta.get("latestDate")
    return Response(
        {
            "success": True,
            "data": {
                "earliestDate": earliest.isoformat() if earliest else None,
                "latestDate": latest.isoformat() if latest else None,
                "hasData": bool(earliest and latest),
                "totalRecords": meta.get("totalRecords") or 0,
            },
        }
    )


@api_view(["GET"])
def outbound_barcode_daily(request):
    start = request.query_params.get("startDate")
    end = request.query_params.get("endDate")
    days = request.query_params.get("days")

    qs = OutboundRecord.objects.all()
    if start and end:
        qs = qs.filter(outbound_date__range=[start, end])
    elif days:
        try:
            d = int(days)
        except Exception:
            d = 90
        since = timezone.localdate() - timedelta(days=d)
        qs = qs.filter(outbound_date__gte=since)

    qs = qs.exclude(barcode__isnull=True).exclude(barcode="")

    daily = (
        qs.values("barcode", "product_name", "category", "outbound_date")
        .annotate(
            quantity=Coalesce(Sum("box_quantity"), 0),
            sales_amount=Coalesce(Sum("sales_amount"), Decimal("0")),
        )
        .order_by("barcode", "outbound_date")
    )

    grouped = {}
    total_records = 0
    for row in daily:
        total_records += 1
        bc = row["barcode"]
        g = grouped.get(bc)
        if not g:
            g = {
                "barcode": bc,
                "productName": row.get("product_name") or "-",
                "category": row.get("category") or "-",
                "dailyData": [],
            }
            grouped[bc] = g
        g["dailyData"].append(
            {
                "date": row["outbound_date"].isoformat()
                if row.get("outbound_date")
                else None,
                "quantity": int(row.get("quantity") or 0),
                "salesAmount": float(row.get("sales_amount") or 0),
            }
        )

    data = []
    for bc, g in grouped.items():
        total_qty = sum(int(d.get("quantity") or 0) for d in g["dailyData"])
        days_count = len(g.get("dailyData") or [])
        avg_daily = (total_qty / float(days_count)) if days_count > 0 else 0.0
        g["totalOutbound"] = total_qty
        g["avgDaily"] = float(avg_daily)
        g["calculatedSettings"] = {
            "minStock": int(round(avg_daily * 3)),
            "maxStock": int(round(avg_daily * 30)),
            "reorderPoint": int(round(avg_daily * 3)),
        }
        data.append(g)

    return Response(
        {
            "success": True,
            "data": data,
            "summary": {
                "totalRecords": total_records,
                "totalBarcodes": len(data),
            },
        }
    )


@api_view(["GET"])
def outbound_daily_analysis(request):
    latest = OutboundRecord.objects.aggregate(latest=Max("outbound_date")).get("latest")
    if not latest:
        return Response({"insight": None})

    totals = OutboundRecord.objects.filter(outbound_date=latest).aggregate(
        totalSales=Sum("sales_amount"),
        totalQty=Sum("quantity"),
        totalCount=Count("id"),
    )
    insight = (
        f"### {latest.isoformat()} 출고 요약\n\n"
        f"- 총 건수: {totals.get('totalCount') or 0}건\n"
        f"- 총 박스수량: {int(totals.get('totalQty') or 0)}\n"
        f"- 총 매출: {float(totals.get('totalSales') or 0):,.0f}원\n"
    )
    return Response({"date": latest.isoformat(), "insight": insight})


@api_view(["POST"])
def outbound_ai_analysis(request):
    summary_stats = (
        request.data.get("summaryStats") if isinstance(request.data, dict) else None
    )
    start_date = (
        request.data.get("startDate") if isinstance(request.data, dict) else None
    )
    end_date = request.data.get("endDate") if isinstance(request.data, dict) else None
    category = request.data.get("category") if isinstance(request.data, dict) else None
    search_query = (
        request.data.get("searchQuery") if isinstance(request.data, dict) else None
    )
    product = request.data.get("product") if isinstance(request.data, dict) else None

    total_sales = 0
    try:
        total_sales = float((summary_stats or {}).get("totalSales") or 0)
    except Exception:
        total_sales = 0

    if total_sales <= 0:
        return Response(
            {
                "analysis": "### 데이터 부족\n\n분석할 데이터가 충분하지 않습니다. 데이터를 업로드하거나 동기화해주세요."
            }
        )

    # 1. Django ORM을 이용한 실시간 다차원 정량 통계 집계
    queryset = OutboundRecord.objects.all()
    if start_date:
        queryset = queryset.filter(outbound_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(outbound_date__lte=end_date)
    if category and category != "all":
        queryset = queryset.filter(category=category)
    if search_query:
        queryset = queryset.filter(product_name__icontains=search_query)
    if product:
        queryset = queryset.filter(product_name=product)

    # 품목별 실적 (출고량 기준 TOP 5)
    top_products_qs = (
        queryset.values("product_name")
        .annotate(
            total_qty=Coalesce(Sum("box_quantity"), 0),
            total_sales=Coalesce(Sum("sales_amount"), Decimal("0")),
        )
        .order_by("-total_qty")[:5]
    )
    
    top_products_text = ""
    for idx, p in enumerate(top_products_qs, 1):
        p_name = p["product_name"] or "-"
        p_qty = p["total_qty"] or 0
        p_sales = float(p["total_sales"] or 0)
        p_share = (p_sales / total_sales * 100) if total_sales > 0 else 0
        top_products_text += f"{idx}. {p_name}: 출고량 {p_qty:,} Box, 매출액 {p_sales:,.0f}원 (매출 비중 {p_share:.1f}%)\n"

    # 카테고리별 실적 (매출 기준 TOP 3)
    top_categories_qs = (
        queryset.values("category")
        .annotate(
            total_qty=Coalesce(Sum("box_quantity"), 0),
            total_sales=Coalesce(Sum("sales_amount"), Decimal("0")),
        )
        .order_by("-total_sales")[:3]
    )
    
    top_categories_text = ""
    for idx, c in enumerate(top_categories_qs, 1):
        c_name = c["category"] or "-"
        c_qty = c["total_qty"] or 0
        c_sales = float(c["total_sales"] or 0)
        c_share = (c_sales / total_sales * 100) if total_sales > 0 else 0
        top_categories_text += f"{idx}. {c_name}: 출고량 {c_qty:,} Box, 매출액 {c_sales:,.0f}원 (매출 비중 {c_share:.1f}%)\n"

    # 최고 매출 피크일 (Peak Day)
    peak_day_qs = (
        queryset.values("outbound_date")
        .annotate(
            daily_sales=Coalesce(Sum("sales_amount"), Decimal("0")),
            daily_qty=Coalesce(Sum("box_quantity"), 0),
        )
        .order_by("-daily_sales")
        .first()
    )
    
    peak_day_text = "-"
    if peak_day_qs and peak_day_qs["outbound_date"]:
        pk_date = peak_day_qs["outbound_date"].strftime("%Y-%m-%d")
        pk_sales = float(peak_day_qs["daily_sales"] or 0)
        pk_qty = peak_day_qs["daily_qty"] or 0
        
        # 피크일 당일 최고 매출 기여 품목
        pk_product_qs = (
            queryset.filter(outbound_date=peak_day_qs["outbound_date"])
            .values("product_name")
            .annotate(p_sales=Coalesce(Sum("sales_amount"), Decimal("0")))
            .order_by("-p_sales")
            .first()
        )
        pk_prod_name = pk_product_qs["product_name"] if pk_product_qs else "-"
        peak_day_text = f"{pk_date} (당일 매출: {pk_sales:,.0f}원, 출고량: {pk_qty:,} Box, 최고 공헌 품목: {pk_prod_name})"

    # 일자 수 (전체 일자 집계 로드 없이 count 만)
    try:
        days_count = (
            queryset.values("outbound_date")
            .exclude(outbound_date__isnull=True)
            .distinct()
            .count()
        )
    except Exception:
        days_count = 0
    avg_sales = (total_sales / days_count) if days_count > 0 else 0
    avg_qty = (
        (float((summary_stats or {}).get("totalQty") or 0) / days_count)
        if days_count > 0
        else 0
    )

    system_prompt = (
        "당신은 보노하우스(BONOHOUSE)의 출고 및 매출 데이터를 분석하는 비즈니스 데이터 과학자이자 물류 전략 컨설턴트입니다. "
        "전달된 실데이터 수치(원화 매출액, 제품명, 출고량 등)를 마크다운 형식으로 상세하고 정량적으로 분석하여 보고서를 작성하세요. "
        "추측성 표현은 지양하고, 반드시 전달된 데이터에 기재된 정확한 이름과 수치들을 인용하세요. "
        "한국어로만 정중하고 전문적인 어조로 답변하세요. 각 주제는 2~4문장으로 간결하게."
    )

    user_prompt = (
        "다음 조건의 출고 및 매출 실데이터 통계를 바탕으로 분석 보고서를 작성해 주세요.\n\n"
        "### 1. 분석 기본 조건\n"
        f"- 분석 기간: {start_date or '-'} ~ {end_date or '-'}\n"
        f"- 카테고리 필터: {category or '전체(all)'}\n"
        f"- 상세 검색어: {search_query or '없음(-)'}\n"
        f"- 특정 품목 필터: {product or '없음(-)'}\n\n"
        "### 2. 핵심 종합 지표\n"
        f"- 총 매출액: {total_sales:,.0f}원\n"
        f"- 총 출고량: {float((summary_stats or {}).get('totalQty') or 0):,.0f} Box\n"
        f"- 데이터 수집 일수: {days_count}일\n"
        f"- 일평균 매출액: {avg_sales:,.0f}원\n"
        f"- 일평균 출고량: {avg_qty:,.1f} Box\n\n"
        "### 3. 카테고리별 실적 (매출 기준 TOP 3)\n"
        f"{top_categories_text or '자료 없음'}\n"
        "### 4. 제품별 실적 (출고량 기준 TOP 5)\n"
        f"{top_products_text or '자료 없음'}\n"
        "### 5. 최고 매출 피크일 (Peak Day)\n"
        f"- {peak_day_text}\n\n"
        "### [작성 형식]\n"
        "아래 4개 대주제로 구분하여 각각 2~4문장으로 작성하세요:\n"
        "1. **출고 및 매출 종합 성과 진단**\n"
        "2. **품목 및 카테고리 기여도**\n"
        "3. **판매 안정성 및 리스크**\n"
        "4. **물류 효율 및 액션 제안** (2~3가지)\n"
    )

    try:
        zai_text = _zai_call_messages(
            system=system_prompt, user=user_prompt, max_tokens=600, temperature=0.2
        )
        if isinstance(zai_text, str) and zai_text.strip():
            return Response({"analysis": zai_text})
        # 외부 AI 무응답 → 로컬 정량 요약 (프록시 끊김/대기 대신 즉시 응답)
        local = (
            f"### 출고 AI 분석 (로컬 요약)\n\n"
            f"외부 AI 응답이 없거나 지연되어 **DB 집계 기반 요약**을 표시합니다.\n\n"
            f"**기간:** {start_date or '-'} ~ {end_date or '-'} ({days_count}일)\n\n"
            f"- 총 매출: **{total_sales:,.0f}원** / 일평균 {avg_sales:,.0f}원\n"
            f"- 총 출고: **{float((summary_stats or {}).get('totalQty') or 0):,.0f} Box** / 일평균 {avg_qty:,.1f} Box\n\n"
            f"**카테고리 TOP**\n{top_categories_text or '- 없음'}\n"
            f"**품목 TOP**\n{top_products_text or '- 없음'}\n"
            f"**피크일:** {peak_day_text}\n"
        )
        return Response({"analysis": local, "source": "local-fallback"})
    except urllib.error.HTTPError as e:
        try:
            err_body = e.read().decode("utf-8")
        except Exception:
            err_body = ""
        return Response(
            {
                "analysis": (
                    "### AI 서버 응답 오류\n\n"
                    f"HTTP {getattr(e, 'code', '-')}: {getattr(e, 'reason', 'error')}\n\n"
                    f"(응답 본문 일부)\n\n{err_body[:800]}"
                )
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {
                "analysis": (
                    "### AI 서버 연결 실패\n\n"
                    f"- 사유: {str(e)}\n\n"
                    "(환경변수/네트워크 상태를 확인하세요. 백엔드 `runserver`가 하나만 떠 있는지 확인.)"
                )
            },
            status=status.HTTP_200_OK,
        )


@api_view(["POST"])
def ai_chat(request):
    message = (
        (request.data.get("message") or "").strip()
        if isinstance(request.data, dict)
        else ""
    )
    page_context = (
        request.data.get("pageContext") if isinstance(request.data, dict) else None
    )
    filters = request.data.get("filters") if isinstance(request.data, dict) else {}

    if not message:
        return Response(
            {"answer": "질문을 입력해주세요."}, status=status.HTTP_400_BAD_REQUEST
        )

    # Determine current page type
    page_type = page_context.get("type") if page_context else "vf-outbound"
    page_name = page_context.get("name") if page_context else "VF 출고 대시보드"

    # Get current data context
    from datetime import datetime, timedelta

    today = timezone.localdate()
    yesterday = today - timedelta(days=1)

    # Fetch ALL data sources regardless of page type
    try:
        context_info = {
            "page_name": page_name,
            "page_type": page_type,
            "today": today.isoformat(),
            "yesterday": yesterday.isoformat(),
        }

        # 1. VF Outbound Data - Direct query
        try:
            from .models import OutboundRecord

            vf_total_count = OutboundRecord.objects.count()
            vf_total_quantity = (
                OutboundRecord.objects.aggregate(total=Sum("quantity"))["total"] or 0
            )
            vf_total_sales = (
                OutboundRecord.objects.aggregate(total=Sum("sales_amount"))["total"]
                or 0
            )

            context_info["vf_total_count"] = vf_total_count
            context_info["vf_total_quantity"] = vf_total_quantity
            context_info["vf_total_sales"] = vf_total_sales

            # Today's outbound
            vf_today_quantity = (
                OutboundRecord.objects.filter(outbound_date=today).aggregate(
                    total=Sum("quantity")
                )["total"]
                or 0
            )
            context_info["vf_today_quantity"] = vf_today_quantity

            # Yesterday's outbound
            vf_yesterday_quantity = (
                OutboundRecord.objects.filter(outbound_date=yesterday).aggregate(
                    total=Sum("quantity")
                )["total"]
                or 0
            )
            context_info["vf_yesterday_quantity"] = vf_yesterday_quantity

            # Daily change percentage
            if vf_yesterday_quantity > 0:
                change_pct = (
                    (vf_today_quantity - vf_yesterday_quantity) / vf_yesterday_quantity
                ) * 100
                context_info["vf_daily_change"] = f"{change_pct:+.1f}%"

            # Yesterday's top products
            yesterday_top_products = list(
                OutboundRecord.objects.filter(outbound_date=yesterday)
                .values("product_name")
                .annotate(
                    total_quantity=Sum("quantity"), total_sales=Sum("sales_amount")
                )
                .order_by("-total_quantity")[:5]
            )
            context_info["vf_yesterday_top_products"] = yesterday_top_products

            # Top products (all time)
            top_products = list(
                OutboundRecord.objects.values("product_name")
                .annotate(
                    total_quantity=Sum("quantity"), total_sales=Sum("sales_amount")
                )
                .order_by("-total_quantity")[:5]
            )
            context_info["vf_top_products"] = top_products

            # Category breakdown
            category_breakdown = list(
                OutboundRecord.objects.values("category")
                .annotate(
                    total_quantity=Sum("quantity"), total_sales=Sum("sales_amount")
                )
                .order_by("-total_quantity")[:5]
            )
            context_info["vf_categories"] = category_breakdown

            # Recent daily trend (last 7 days)
            recent_dates = (
                OutboundRecord.objects.filter(outbound_date__gte=today - timedelta(days=7))
                .values("outbound_date")
                .annotate(quantity=Sum("quantity"), sales_amount=Sum("sales_amount"))
                .order_by("outbound_date")
            )
            context_info["vf_daily_trend"] = list(recent_dates)

        except Exception as e:
            logger.warning(f"Failed to fetch VF outbound data: {e}")

        # 2. FC Inbound Data - Direct query
        try:
            from .models import FCInboundRecord

            fc_total_count = FCInboundRecord.objects.count()
            fc_total_quantity = (
                FCInboundRecord.objects.aggregate(total=Sum("quantity"))["total"] or 0
            )

            context_info["fc_total_count"] = fc_total_count
            context_info["fc_total_quantity"] = fc_total_quantity

            # Today's inbound
            fc_today_quantity = (
                FCInboundRecord.objects.filter(receiving_date__date=today).aggregate(
                    total=Sum("quantity")
                )["total"]
                or 0
            )
            context_info["fc_today_quantity"] = fc_today_quantity

            # Yesterday's inbound
            fc_yesterday_quantity = (
                FCInboundRecord.objects.filter(
                    receiving_date__date=yesterday
                ).aggregate(total=Sum("quantity"))["total"]
                or 0
            )
            context_info["fc_yesterday_quantity"] = fc_yesterday_quantity

            # Daily change percentage
            if fc_yesterday_quantity > 0:
                change_pct = (
                    (fc_today_quantity - fc_yesterday_quantity) / fc_yesterday_quantity
                ) * 100
                context_info["fc_daily_change"] = f"{change_pct:+.1f}%"

            # Top products
            top_products = list(
                FCInboundRecord.objects.values("product_name")
                .annotate(total_quantity=Sum("quantity"))
                .order_by("-total_quantity")[:5]
            )
            context_info["fc_top_products"] = top_products

        except Exception as e:
            logger.warning(f"Failed to fetch FC inbound data: {e}")

        # 3. Inventory Data
        try:
            from .models import InventoryItem

            inventory_items = InventoryItem.objects.all()
            total_inventory = inventory_items.count()
            low_stock_items = inventory_items.filter(
                current_stock__lte=models.F("minimum_stock")
            ).count()
            context_info["inventory_total_items"] = total_inventory
            context_info["inventory_low_stock_count"] = low_stock_items

            # Recent inventory movements (receipts)
            from .models import InventoryReceiptItem

            recent_receipts = InventoryReceiptItem.objects.filter(
                receipt__upload_date__gte=today - timedelta(days=7)
            ).count()
            context_info["inventory_recent_receipts"] = recent_receipts
        except Exception as e:
            logger.warning(f"Failed to fetch inventory data: {e}")

        # 4. Delivery Data
        try:
            from .models import DeliveryDailyRecord

            delivery_today = DeliveryDailyRecord.objects.filter(date=today).first()
            if delivery_today:
                context_info["delivery_today_total"] = delivery_today.total
                context_info["delivery_today_by_hour"] = delivery_today.hourly

            delivery_yesterday = DeliveryDailyRecord.objects.filter(
                date=yesterday
            ).first()
            if delivery_yesterday:
                context_info["delivery_yesterday_total"] = delivery_yesterday.total

            # Special notes
            from .models import DeliverySpecialNote

            recent_notes = DeliverySpecialNote.objects.filter(
                note_date__gte=today - timedelta(days=3)
            ).values_list("note_content", flat=True)
            if recent_notes:
                context_info["delivery_special_notes"] = list(recent_notes)
        except Exception as e:
            logger.warning(f"Failed to fetch delivery data: {e}")

        # 5. Production Data - 사용자가 언급한 날짜도 함께 조회
        try:
            from .models import ProductionLog
            from datetime import datetime

            # 사용자가 언급한 날짜 파싱 (04-27, 2026-04-27, 4월27일等形式)
            mentioned_dates = []
            import re
            date_patterns = [
                r'(\d{4})-(\d{2})-(\d{2})',  # 2026-04-27
                r'(\d{2})-(\d{2})',          # 04-27
                r'(\d{1,2})월\s*(\d{1,2})일',  # 4월 27일
            ]
            msg_lower = message.lower()
            for pattern in date_patterns:
                matches = re.findall(pattern, msg_lower)
                for m in matches:
                    try:
                        if len(m) == 3:
                            year, month, day = int(m[0]), int(m[1]), int(m[2])
                            if year < 100:
                                year = 2026
                            mentioned_dates.append(f"{year:04d}-{month:02d}-{day:02d}")
                        elif len(m) == 2:
                            month, day = int(m[0]), int(m[1])
                            mentioned_dates.append(f"2026-{month:02d}-{day:02d}")
                    except:
                        pass

            # 오늘 + 언급된 날짜 모두 조회
            all_target_dates = {today.isoformat()}
            for d in mentioned_dates:
                all_target_dates.add(d)

            all_production_data = {}
            for target_date in all_target_dates:
                logs = ProductionLog.objects.filter(date=target_date)
                total = logs.count()
                if total > 0:
                    active = logs.filter(status="started").count()
                    completed = logs.filter(status="ended").count()
                    pending = logs.filter(status="pending").count()
                    output = sum(log.quantity * log.unit_quantity for log in logs.filter(status="ended"))
                    items = []
                    sorted_logs = logs.order_by('date', 'id')
                    for idx, log in enumerate(sorted_logs[:30], 1):
                        product = log.product_name or '알 수 없음'
                        items.append(f"순번{idx}|ID:{log.id}|날짜:{log.date}|기계:{log.machine_number}|제품:{product}|상태:{log.status}|수량:{log.quantity}")
                    all_production_data[target_date] = {
                        "total": total, "active": active, "completed": completed,
                        "pending": pending, "output": output, "items": items
                    }

            # 대표 값 (오늘 기준)
            today_data = all_production_data.get(today.isoformat(), {})
            context_info["production_active_count"] = today_data.get("active", 0)
            context_info["production_completed_today"] = today_data.get("completed", 0)
            context_info["production_pending_count"] = today_data.get("pending", 0)
            context_info["production_today_total"] = today_data.get("total", 0)
            context_info["production_today_output"] = today_data.get("output", 0)
            context_info["production_items"] = today_data.get("items", [])

            # 모든 날짜 데이터를 AI에게 전달
            if len(all_production_data) > 1:
                all_items = []
                for d, data in sorted(all_production_data.items()):
                    all_items.append(f"=== {d} 생산 데이터 (총 {data['total']}건) ===")
                    all_items.extend(data["items"])
                context_info["production_all_dates"] = all_items
            elif len(all_production_data) == 1:
                # 단일 날짜면 그 날짜를 명시
                for d, data in all_production_data.items():
                    all_items = [f"=== {d} 생산 데이터 (총 {data['total']}건) ==="]
                    all_items.extend(data["items"])
                    context_info["production_all_dates"] = all_items

            logger.info(f"[AI-CHAT] production dates: {list(all_production_data.keys())}")
        except Exception as e:
            logger.warning(f"Failed to fetch production data: {e}")

        # 6. BACO Transfer Data
        try:
            from .models import BarcodeTransferRecord

            today_transfers = BarcodeTransferRecord.objects.filter(
                created_at__date=today
            ).count()
            context_info["baco_today_transfers"] = today_transfers
        except Exception as e:
            logger.warning(f"Failed to fetch BACO transfer data: {e}")

        # Build comprehensive context string
        user_prompt = f"""현재 컨텍스트 (VF/FC 통합 대시보드):
- 현재 페이지: {context_info.get("page_name", "VF/FC 대시보드")}
- 오늘: {context_info.get("today", today)}
- 어제: {context_info.get("yesterday", yesterday)}

=== VF 출고 데이터 ==="""
        if "vf_total_count" in context_info:
            user_prompt += f"\n- VF 총 출고 건수: {context_info['vf_total_count']:,}"
        if "vf_total_sales" in context_info and context_info["vf_total_sales"]:
            user_prompt += f"\n- VF 총 매출: {context_info['vf_total_sales']:,.0f}원"
        if "vf_total_quantity" in context_info and context_info["vf_total_quantity"]:
            user_prompt += f"\n- VF 총 수량: {context_info['vf_total_quantity']:,.0f}"
        if "vf_today_quantity" in context_info:
            user_prompt += (
                f"\n- VF 오늘 출고량: {context_info['vf_today_quantity']:,.0f}"
            )
        if "vf_yesterday_quantity" in context_info:
            user_prompt += (
                f"\n- VF 어제 출고량: {context_info['vf_yesterday_quantity']:,.0f}"
            )
        if "vf_daily_change" in context_info:
            user_prompt += f"\n- VF 전일 대비: {context_info['vf_daily_change']}"
        if "vf_top_products" in context_info and context_info["vf_top_products"]:
            user_prompt += "\n- VF 전체 출고 상위 품목(전체 기간):"
            for i, p in enumerate(context_info["vf_top_products"][:3], 1):
                if isinstance(p, dict):
                    name = p.get("product_name", p.get("name", ""))
                    qty = p.get("total_quantity", p.get("quantity", 0))
                    user_prompt += f"  {i}. {name}: {qty:,.0f}"
        else:
            user_prompt += "\n- VF 출고 데이터: 현재 시스템에 출고 기록이 없습니다"

        if "vf_yesterday_top_products" in context_info and context_info["vf_yesterday_top_products"]:
            user_prompt += f"\n- VF 어제({yesterday}) 출고 상위 품목:"
            for i, p in enumerate(context_info["vf_yesterday_top_products"][:5], 1):
                if isinstance(p, dict):
                    name = p.get("product_name", p.get("name", ""))
                    qty = p.get("total_quantity", p.get("quantity", 0))
                    sales = p.get("total_sales", 0)
                    user_prompt += f"  {i}. {name}: {qty:,.0f}개 (매출 {sales:,.0f}원)"

        # Add daily trend data for specific date queries
        if "vf_daily_trend" in context_info:
            user_prompt += "\n- VF 최근 7일 추이 (날짜별 조회 가능):"
            for trend in context_info["vf_daily_trend"][:10]:
                if isinstance(trend, dict):
                    date = trend.get("outbound_date", trend.get("date", ""))
                    qty = trend.get("quantity", 0)
                    sales = trend.get("sales_amount", 0)
                    user_prompt += f"  * {date}: {qty:,.0f}개 (매출 {sales:,.0f}원)"

        user_prompt += "\n\n=== FC 입고 데이터 ==="
        if "fc_total_count" in context_info:
            user_prompt += f"\n- FC 총 입고 건수: {context_info['fc_total_count']:,}"
        if "fc_total_quantity" in context_info and context_info["fc_total_quantity"]:
            user_prompt += (
                f"\n- FC 총 입고 수량: {context_info['fc_total_quantity']:,.0f}"
            )
        if "fc_today_quantity" in context_info:
            user_prompt += (
                f"\n- FC 오늘 입고량: {context_info['fc_today_quantity']:,.0f}"
            )
        if "fc_yesterday_quantity" in context_info:
            user_prompt += (
                f"\n- FC 어제 입고량: {context_info['fc_yesterday_quantity']:,.0f}"
            )
        if "fc_daily_change" in context_info:
            user_prompt += f"\n- FC 전일 대비: {context_info['fc_daily_change']}"

        user_prompt += "\n\n=== 재고 데이터 ==="
        if "inventory_total_items" in context_info:
            user_prompt += (
                f"\n- 전산 재고 품목 수: {context_info['inventory_total_items']}"
            )
        if "inventory_low_stock_count" in context_info:
            user_prompt += (
                f"\n- 안전재고 미달 품목: {context_info['inventory_low_stock_count']}"
            )
        if "inventory_recent_receipts" in context_info:
            user_prompt += (
                f"\n- 최근 7일 입고 수: {context_info['inventory_recent_receipts']}"
            )

        user_prompt += "\n\n=== 배송 데이터 ==="
        if (
            "delivery_today_total" in context_info
            and context_info["delivery_today_total"] is not None
        ):
            user_prompt += (
                f"\n- 오늘 배송总量: {context_info['delivery_today_total']:,.0f}"
            )
        if (
            "delivery_yesterday_total" in context_info
            and context_info["delivery_yesterday_total"] is not None
        ):
            user_prompt += (
                f"\n- 어제 배송总量: {context_info['delivery_yesterday_total']:,.0f}"
            )
        if "delivery_special_notes" in context_info:
            user_prompt += "\n- 최근 특이사항:"
            for note in context_info["delivery_special_notes"][:3]:
                user_prompt += f"  • {note}"

        user_prompt += "\n\n=== 생산 데이터 ==="
        if "production_today_total" in context_info:
            user_prompt += f"\n- 오늘 전체 생산: {context_info['production_today_total']}건"
        if "production_pending_count" in context_info:
            user_prompt += f"\n- 대기 중 생산: {context_info['production_pending_count']}건"
        if "production_active_count" in context_info:
            user_prompt += (
                f"\n- 진행 중 생산: {context_info['production_active_count']}건"
            )
        if "production_completed_today" in context_info:
            user_prompt += (
                f"\n- 오늘 완료 생산: {context_info['production_completed_today']}건"
            )
        if "production_today_output" in context_info:
            user_prompt += (
                f"\n- 오늘 생산량: {context_info['production_today_output']:,.0f}"
            )
        if "production_items" in context_info and context_info["production_items"]:
            user_prompt += "\n- 오늘 생산 항목 목록:"
            for item in context_info["production_items"]:
                user_prompt += f"\n  {item}"
        if "production_all_dates" in context_info and context_info["production_all_dates"]:
            user_prompt += "\n\n=== 모든 날짜 생산 데이터 ==="
            for line in context_info["production_all_dates"]:
                user_prompt += f"\n  {line}"

        user_prompt += "\n\n=== BACO 데이터 ==="
        if "baco_today_transfers" in context_info:
            user_prompt += (
                f"\n- 오늘 바코드 전송: {context_info['baco_today_transfers']}건"
            )

        user_prompt += f"\n\n사용자 질문: {message}\n\n"

        # Intent-specific guidance
        user_prompt += """답변 가이드:
1. 질문에 대해 구체적으로 답변하세요
2. 가능한 한 실제 데이터를 인용하세요 (VF/FC/재고/배송/생산 모두 활용)
3. **중요: 특정 날짜를 물어보면 위 "VF 최근 7일 추이" 데이터에서 해당 날짜를 찾아서 답변하세요**
4. 판매 추이 분석 시 증감/감소율을 포함하세요
5. 품목별 분석 시 상위 3개를 언급하세요
6. 특이사항이 있으면 명확히 설명하세요
7. 데이터에 없는 날짜를 물어보면 솔직하게 "데이터에 없습니다"라고 말씀하세요
8. VF 출고, FC 입고, 재고, 배송, 생산 데이터를 모두 고려하여 종합적으로 분석하세요
9. **매우 중요: "최근 며칠", "최근 일주일" 등 상대적 표현 대신 실제 날짜 범위를 반드시 명시하세요**
10. **중요: 삭제 요청 시 "생산 항목 목록"의 순번과 ID를 확인하고, 삭제할 ID가 있으면 AI가 직접 삭제 API를 호출하여 삭제합니다. 목록에 없는 순번은 "존재하지 않습니다"라고 답변하세요**
11. **매우 중요: 답변은 간결하게! 불필요한 설명, 참고 사항, 분석,吐槽은 제외하세요. 핵심 데이터만Returned.**
12. **단일 수치 조회 질문 시(예: "어제 출고량?", "오늘 생산량?" 등): 표나 목록 대신 "640개"처럼 단일 값으로Returned.**

**삭제 기능:**
- 사용자가 "순번 3번 삭제", "ID 5번 삭제", "3번 지워줘" 등 삭제 요청을 하면:
  1. 먼저 production_items에서 해당 순번의 ID를 찾음
  2. ID를 찾으면 DELETE /api/production-log/{id} 또는 POST /api/production-log/bulk-delete (ids 배열) 호출하여 삭제
  3. 삭제 결과(성공/실패)를 사용자에게 안내
- **주의: 삭제 전 반드시 목록에서 해당 순번/ID가 실제로 존재하는지 확인하세요**
"""

        # System prompt
        system_prompt = (
            "당신은 VF/FC 통합 데이터 분석 전문가 AI 어시스턴트입니다. "
            "VF 출고, FC 입고, 재고, 배송, 생산 등 전체 데이터에 대해 종합적으로 분석할 수 있습니다. "
            "한국어로만 답변하세요. "
            "답변은 친절하고 전문적인 어조로 작성하세요. "
            "매우 중요: 실제 데이터가 없거나 데이터가 0인 경우, '데이터가 없습니다' 또는 '0건입니다'라고 솔직하게 답변하세요. "
            "절대 존재하지 않는 가상의 수치나 품목을 만들어서 답변하지 마세요. "
            "데이터가 없으면 없다고 명확히 말씀하세요. "
            "가능한 한 구체적인 수치를 제공하세요. "
            "**매우 중요: 답변은 간결하게! 표, 목록, 분석, 참고 사항 없이 핵심 값만Returned.**"
        )

        # =========================================
        # 삭제 요청 처리 (AI 호출 전에 먼저 처리)
        # =========================================
        import re
        delete_patterns = [
            r'순번\s*(\d+)\s*삭제',
            r'순번\s*(\d+)번?\s*삭제',
            r'ID\s*(\d+)\s*삭제',
            r'(\d+)번\s*삭제',           # Must have "삭제" immediately after number+번
            r'삭제\s*(\d+)',
            r'(\d+)번\s*지워[줘]',
            r'지워\s*(\d+)번?',
            r'삭제해\s*(\d+)',
        ]
        delete_match = None
        for pattern in delete_patterns:
            m = re.search(pattern, message)
            if m:
                delete_match = m.group(1)
                break

        if delete_match and 'production_items' in context_info and context_info['production_items']:
            seq_num = int(delete_match)
            items = context_info['production_items']
            # items 형식: "순번{idx}|ID:{log.id}|날짜:...|..."
            target_id = None
            for item in items:
                item_str = str(item)
                # "순번3|ID:5|..." 에서 순번 추출
                seq_match = re.search(r'순번(\d+)', item_str)
                id_match = re.search(r'ID:(\d+)', item_str)
                if seq_match and int(seq_match.group(1)) == seq_num:
                    if id_match:
                        target_id = int(id_match.group(1))
                        break
            
            if target_id:
                try:
                    logger.info(f"[AI-DELETE] Attempting to delete seq_num={seq_num}, target_id={target_id}")
                    logger.info(f"[AI-DELETE] Before delete - ProductionLog id={target_id} exists: {ProductionLog.objects.filter(id=target_id).exists()}")
                    deleted_count, _ = ProductionLog.objects.filter(id=target_id).delete()
                    logger.info(f"[AI-DELETE] After delete - deleted_count={deleted_count}")
                    if deleted_count > 0:
                        return Response({
                            "answer": f"✅ 순번 {seq_num}번 (ID: {target_id}) 생산 로그 삭제가 완료되었습니다. 삭제 후 페이지를 새로고침하여 확인해주세요."
                        })
                    else:
                        return Response({
                            "answer": f"순번 {seq_num}번 (ID: {target_id}) 존재하지 않거나 이미 삭제되었습니다."
                        })
                except Exception as del_err:
                    logger.error(f"Failed to delete production log: {del_err}")
                    return Response({
                        "answer": f"삭제 처리 중 오류가 발생했습니다: {str(del_err)}"
                    })
            else:
                logger.info(f"[AI-DELETE] target_id is None for seq_num={seq_num}. Items available: {len(context_info.get('production_items', []))}")
                return Response({
                    "answer": f"순번 {seq_num}번은 현재 목록에 존재하지 않습니다. 현재 목록의 순번을 확인해주세요."
                })

        # Call AI
        try:
            ai_response = _zai_call_messages(
                system=system_prompt, user=user_prompt, max_tokens=700, temperature=0.3
            )
            if isinstance(ai_response, str) and ai_response.strip():
                return Response({"answer": ai_response})
        except Exception as e:
            logger.error(f"AI call failed: {e}")

        # Fallback response
        return Response(
            {"answer": f"죄송합니다. AI 서비스를 이용할 수 없습니다. 질문: {message}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    except Exception as e:
        logger.error(f"AI chat error: {e}")
        return Response(
            {"answer": f"데이터를 가져오는 중 오류가 발생했습니다: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET", "POST"])
def ai_free_models(request):
    """OpenRouter 무료 모델 목록 조회 / 선택 모델 저장."""
    try:
        from . import openrouter_service as ors

        if request.method == "POST":
            payload = request.data if isinstance(request.data, dict) else {}
            model = (
                payload.get("model")
                or payload.get("selectedModel")
                or payload.get("selected_model")
                or ""
            )
            model = str(model).strip()
            if not model:
                return Response(
                    {"success": False, "message": "model is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            store = ors.set_selected_model(model)
            return Response(
                {
                    "success": True,
                    "selectedModel": store.get("selected_model") or model,
                    "models": store.get("models") or [],
                }
            )

        force_raw = (
            request.GET.get("refresh")
            or request.GET.get("force")
            or ""
        )
        force = str(force_raw).strip().lower() in ("1", "true", "yes", "y")
        if force:
            store = ors.refresh_free_models(force=True)
        else:
            store = ors.list_free_models(auto_refresh=True)

        models = store.get("models") if isinstance(store.get("models"), list) else []
        selected = store.get("selected_model") or ors.get_selected_model()
        payload = {
            "success": True,
            "selectedModel": selected,
            "models": models,
            "source": store.get("source"),
            "count": store.get("count") or len(models),
            "cache_hit": store.get("cache_hit"),
        }
        if store.get("error"):
            payload["error"] = store.get("error")
        return Response(payload)
    except Exception as e:
        logger.error(f"ai_free_models error: {e}")
        return Response(
            {"success": False, "message": str(e), "models": []},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
def ai_analyze(request):
    payload = request.data if isinstance(request.data, dict) else {}
    data = payload.get("data") if isinstance(payload.get("data"), dict) else {}

    date = data.get("date")
    day_of_week = data.get("dayOfWeek")
    total = data.get("total")
    average_total = data.get("averageTotal")
    current_hour = data.get("currentHour")
    data_cutoff_hour = data.get("dataCutoffHour")
    data_lag_hours = data.get("dataLagHours")
    current_cum_at_hour = data.get("currentCumAtHour")
    current_inc_at_hour = data.get("currentIncAtHour")
    recent_inc_trend = (
        data.get("recentIncTrend")
        if isinstance(data.get("recentIncTrend"), list)
        else None
    )
    comparison = (
        data.get("comparison") if isinstance(data.get("comparison"), dict) else None
    )
    weekday_profile = (
        data.get("weekdayProfile")
        if isinstance(data.get("weekdayProfile"), dict)
        else None
    )
    weekday_hourly_inc_profile = (
        data.get("weekdayHourlyIncProfile")
        if isinstance(data.get("weekdayHourlyIncProfile"), dict)
        else None
    )
    ai_predictions = (
        data.get("aiPredictions")
        if isinstance(data.get("aiPredictions"), dict)
        else None
    )
    special_notes = (
        data.get("specialNotes") if isinstance(data.get("specialNotes"), list) else None
    )

    try:
        total_val = int(float(str(total).replace(",", ""))) if total is not None else 0
    except Exception:
        total_val = 0

    try:
        avg_val = (
            int(float(str(average_total).replace(",", "")))
            if average_total is not None
            else 0
        )
    except Exception:
        avg_val = 0

    try:
        current_hour_int = int(current_hour) if current_hour is not None else None
    except Exception:
        current_hour_int = None

    try:
        data_cutoff_hour_int = (
            int(data_cutoff_hour) if data_cutoff_hour is not None else None
        )
    except Exception:
        data_cutoff_hour_int = None

    try:
        data_lag_hours_int = int(data_lag_hours) if data_lag_hours is not None else None
    except Exception:
        data_lag_hours_int = None

    if current_hour_int is None:
        try:
            current_hour_int = int(timezone.localtime(timezone.now()).hour)
        except Exception:
            current_hour_int = None

    analysis_hour_int = (
        data_cutoff_hour_int if data_cutoff_hour_int is not None else current_hour_int
    )

    def _to_int_or_none(v):
        if v is None:
            return None
        try:
            return int(float(str(v).replace(",", "")))
        except Exception:
            return None

    cur_cum_int = _to_int_or_none(current_cum_at_hour)
    cur_inc_int = _to_int_or_none(current_inc_at_hour)

    fallback_expected_cum = None
    fallback_expected_inc = None
    if analysis_hour_int is not None and analysis_hour_int >= 0 and avg_val > 0:
        try:
            progress = max(0.0, min(1.0, float(analysis_hour_int) / 23.0))
            fallback_expected_cum = int(round(avg_val * progress))
            fallback_expected_inc = int(round(avg_val / 23.0))
        except Exception:
            fallback_expected_cum = None
            fallback_expected_inc = None

    diff = total_val - avg_val
    diff_pct = (diff / avg_val * 100) if avg_val else 0

    predicted_23 = None
    if ai_predictions:
        try:
            predicted_23 = ai_predictions.get("hour_23")
            if predicted_23 is not None:
                predicted_23 = int(float(str(predicted_23).replace(",", "")))
        except Exception:
            predicted_23 = None

    def _fmt(v):
        if v is None:
            return "-"
        try:
            return f"{int(v):,}"
        except Exception:
            return str(v)

    def _baseline_block(label, b):
        if not isinstance(b, dict):
            return f"- {label}: 데이터 없음\n"

        sample = b.get("sampleCount")
        avg_cum = _to_int_or_none(b.get("avgCumAtHour"))
        avg_inc = _to_int_or_none(b.get("avgIncAtHour"))
        avg_final = _to_int_or_none(b.get("avgFinal"))

        by_period_cum = (
            b.get("byPeriodCumAtHour")
            if isinstance(b.get("byPeriodCumAtHour"), dict)
            else {}
        )
        by_period_final = (
            b.get("byPeriodFinal") if isinstance(b.get("byPeriodFinal"), dict) else {}
        )

        def _period_line(title, src):
            s = _to_int_or_none(src.get("start"))
            m = _to_int_or_none(src.get("mid"))
            e = _to_int_or_none(src.get("end"))
            return f"  - {title}: 월초 {_fmt(s)} / 월중 {_fmt(m)} / 월말 {_fmt(e)}\n"

        out = (
            f"- {label} (표본 {sample or 0}일):\n"
            f"  - 동시간대 누적 평균: {_fmt(avg_cum)}\n"
            f"  - 동시간대 시간증감 평균: {_fmt(avg_inc)}\n"
            f"  - 최종(23시) 누적 평균: {_fmt(avg_final)}\n"
        )
        out += _period_line("동시간대 누적 평균(월초/중/말)", by_period_cum)
        out += _period_line("최종(23시) 누적 평균(월초/중/말)", by_period_final)
        return out

    def _weekday_profile_block(p):
        if not isinstance(p, dict):
            return "- 요일별 프로필: 데이터 없음\n"

        day_names = {0: "일", 1: "월", 2: "화", 3: "수", 4: "목", 5: "금", 6: "토"}

        lines = ["- 요일별 프로필(최근 데이터):"]
        for dow in range(0, 7):
            item = p.get(str(dow)) if str(dow) in p else p.get(dow)
            if not isinstance(item, dict):
                lines.append(f"  - {day_names.get(dow, str(dow))}: 데이터 없음")
                continue

            sc = _to_int_or_none(item.get("sampleCount"))
            avg_cum = _to_int_or_none(item.get("avgCumAtHour"))
            avg_fin = _to_int_or_none(item.get("avgFinal"))
            lines.append(
                f"  - {day_names.get(dow, str(dow))}: 표본 {sc or 0}일 / 동시간대 누적 {_fmt(avg_cum)} / 최종 {_fmt(avg_fin)}"
            )
        return "\n".join(lines) + "\n"

    def _weekday_hourly_inc_summary(p):
        if not isinstance(p, dict):
            return "- 요일×시간대 증감 프로필: 데이터 없음\n"

        by_dow = p.get("byDow") if isinstance(p.get("byDow"), dict) else None
        if not isinstance(by_dow, dict):
            return "- 요일×시간대 증감 프로필: 데이터 없음\n"

        def _get_bucket(dow):
            b = by_dow.get(str(dow)) if str(dow) in by_dow else by_dow.get(dow)
            return b if isinstance(b, dict) else {}

        def _avg_inc(bucket, hours, min_hour_samples=3):
            vals = []
            used = 0
            for h in hours:
                key = str(int(h))
                item = bucket.get(key)
                if not isinstance(item, dict):
                    continue
                sc = _to_int_or_none(item.get("sampleCount")) or 0
                inc = _to_int_or_none(item.get("avgInc"))
                if sc >= min_hour_samples and inc is not None:
                    vals.append(int(inc))
                    used += 1
            if not vals:
                return None, 0
            return int(round(sum(vals) / len(vals))), used

        # late: 17~23, mid: 12~16
        fri_bucket = _get_bucket(5)
        mid_avg, mid_used = _avg_inc(fri_bucket, range(12, 17))
        late_avg, late_used = _avg_inc(fri_bucket, range(17, 24))

        out = ["- 금요일 시간대별 증감 요약(근거 기반):"]
        out.append(
            f"  - 중반(12~16시) 평균 증감: {_fmt(mid_avg)} (유효시간 {mid_used}/5)"
        )
        out.append(
            f"  - 후반(17~23시) 평균 증감: {_fmt(late_avg)} (유효시간 {late_used}/7)"
        )
        if mid_avg is not None and late_avg is not None:
            out.append(f"  - 후반-중반 차이: {late_avg - mid_avg:+,}")
        else:
            out.append("  - (주의) 표본/유효시간 부족으로 후반 둔화 여부 단정 불가")

        return "\n".join(out) + "\n"

    same8w = comparison.get("sameWeekday8w") if isinstance(comparison, dict) else None
    prev_month = (
        comparison.get("prevMonthSameWeekday") if isinstance(comparison, dict) else None
    )

    trend_lines = []
    if recent_inc_trend:
        for item in recent_inc_trend[-3:]:
            if isinstance(item, dict):
                h = _to_int_or_none(item.get("hour"))
                inc = _to_int_or_none(item.get("inc"))
                if h is not None and inc is not None:
                    trend_lines.append(f"- {h}시 증감: {inc:,}")

    system_prompt = (
        "당신은 출고/매출 데이터 분석 전문가입니다. "
        "반드시 한국어로만 답변하세요. 일본어, 한자, 영문은 절대 사용하지 마세요(やや, 参照 등 금지). "
        "답변은 반드시 5~7줄 이내로 간결하고 명료하게 작성하세요. 각 항목은 한 줄만 작성합니다. "
        "접두어 없이 바로 내용을 시작하세요. "
        "시간 해석 규칙: currentHour는 '현재 시각', dataCutoffHour는 '실제 데이터가 확정(입력/집계)된 마지막 시각'입니다. "
        "동시간대 비교/증감/최근 추이는 반드시 dataCutoffHour 기준으로만 수행하세요(추측 금지). "
        "dataLagHours>0 또는 dataCutoffHour < currentHour 인 경우, 아직 집계 전 시간대가 존재할 수 있으므로 0/None 증감을 '주문 증가 멈춤'으로 단정하지 마세요. "
        "입력 지연/시스템 오류 등의 원인을 추정하지 말고, 필요한 경우 '집계/입력 현황 확인 필요'로만 표현하세요. "
        "dataLagHours가 2 이상이거나 데이터 공백이 명확할 때만 '확인 필요'로 1줄 언급하세요. "
        "비교는 '동시간대 누적'과 '동시간대 증감'을 최우선으로 사용하고, '최종(23시) 평균'은 별도로 분리해서 언급하세요. "
        "기준선/표본이 부족해도 가능한 범위에서 최선의 분석을 제시하세요. "
        "비교 기준선이 비어있으면 [폴백 기준]을 사용하세요. "
        "요일별 패턴은 avgCumAtHour 기준으로 1줄만 간략히 언급하세요. "
        "금요일 후반 둔화는 weekdayHourlyIncProfile 근거가 있을 때만 1줄 언급하세요. "
        "데이터에 없는 내용은 추측하지 마세요."
    )

    user_prompt = (
        "현재 주문(누적) 데이터를 '동시간대 기준'으로 비교 분석하여 리포트를 작성해주세요.\n\n"
        f"- 날짜: {date or '-'}\n"
        f"- 요일: {day_of_week or '-'}\n"
        f"- currentHour(현재 시각): {current_hour_int if current_hour_int is not None else '-'}\n"
        f"- dataCutoffHour(확정 데이터 기준 시각): {analysis_hour_int if analysis_hour_int is not None else '-'}\n"
        f"- dataLagHours: {data_lag_hours_int if data_lag_hours_int is not None else '-'}\n"
        f"- 현재 누적(마지막 입력 기준): {total_val:,}건\n"
        f"- 동시간대 누적(dataCutoffHour 기준): {_fmt(cur_cum_int)}건\n"
        f"- 동시간대 시간증감(dataCutoffHour 기준): {_fmt(cur_inc_int)}건\n"
        f"- AI 예측 23시 최종 누적: {_fmt(predicted_23)}건\n\n"
        "[폴백 기준(기준선 부족 시 사용)]\n"
        "- 참고 averageTotal(최종 누적 평균일 가능성): " + f"{avg_val:,}" + "\n"
        "- (근사) 동시간대 기대 누적: " + f"{_fmt(fallback_expected_cum)}" + "\n"
        "- (근사) 시간당 평균 증가량: " + f"{_fmt(fallback_expected_inc)}" + "\n\n"
        "[비교 기준 요약]\n"
        "(주의) '동일 요일 평균(averageTotal)'은 과거 최종 누적 기반일 수 있으므로 참고용으로만 사용하세요.\n"
        f"- 참고 averageTotal: {avg_val:,}\n\n"
        "[동시간대/최종 기준선]\n"
        + _baseline_block("같은 요일 최근 8주", same8w)
        + _baseline_block("같은 요일 이전 월", prev_month)
        + "\n[요일별 프로필(패턴 검증)]\n"
        + _weekday_profile_block(weekday_profile)
        + "\n[금요일 후반 둔화 검증(요일×시간대 증감)]\n"
        + _weekday_hourly_inc_summary(weekday_hourly_inc_profile)
        + "\n"
        + (
            "[최근 3시간 증감 추이]\n" + "\n".join(trend_lines) + "\n\n"
            if trend_lines
            else ""
        )
        + "아래 5개 항목을 각 1줄로 간결하게 작성하세요:\n"
        + "1) 동시간대 현황: 최근 8주 평균과 비교, 현재 누적 및 시간증감\n"
        + "2) 이전 월 비교: 이전 월 같은 요일 동시간대 대비 차이\n"
        + "3) 최근 증감 추이: 최근 3시간 증감 패턴, 가속/감속 여부\n"
        + "4) 최종 예측: 23시 AI 예측값과 과거 평균 비교\n"
        + "5) 요일 패턴: 금요일 특성 및 후반 증감 경향\n"
    )

    if special_notes:
        lines = []
        for n in special_notes[:20]:
            if not isinstance(n, dict):
                continue
            dt = (
                (n.get("event_datetime") or n.get("eventDateTime") or "").strip()
                if isinstance(n.get("event_datetime") or n.get("eventDateTime"), str)
                else ""
            )
            pname = (
                (n.get("product_name") or n.get("productName") or "").strip()
                if isinstance(n.get("product_name") or n.get("productName"), str)
                else ""
            )
            barcode = (
                (n.get("barcode") or "").strip()
                if isinstance(n.get("barcode"), str)
                else ""
            )
            sku = (
                (n.get("sku_id") or n.get("skuId") or "").strip()
                if isinstance(n.get("sku_id") or n.get("skuId"), str)
                else ""
            )
            memo = (
                (n.get("memo") or n.get("text") or "").strip()
                if isinstance(n.get("memo") or n.get("text"), str)
                else ""
            )
            qty = n.get("quantity")
            try:
                qty_str = (
                    f"{int(float(str(qty).replace(',', ''))):,}"
                    if qty is not None and str(qty).strip() != ""
                    else "-"
                )
            except Exception:
                qty_str = str(qty) if qty is not None else "-"
            parts = []
            if dt:
                parts.append(dt)
            if pname:
                parts.append(pname)
            if barcode:
                parts.append(f"barcode:{barcode}")
            if sku:
                parts.append(f"sku:{sku}")
            parts.append(f"qty:{qty_str}")
            if memo:
                parts.append(memo)
            if parts:
                lines.append("- " + " / ".join(parts))
        if lines:
            user_prompt += (
                "\n[특이사항 메모(업무 컨텍스트)]\n" + "\n".join(lines) + "\n"
            )

    try:
        zai_text = _zai_call_messages(
            system=system_prompt, user=user_prompt, max_tokens=800, temperature=0.3
        )
        if isinstance(zai_text, str) and zai_text.strip():
            return Response(
                {"success": True, "insight": zai_text}, status=status.HTTP_200_OK
            )
    except Exception:
        pass

    insight = (
        f"## AI 분석 (간이)\n\n"
        f"- 기준일: {date or '-'} ({day_of_week or '-'})\n"
        f"- 현재 누적: {total_val:,}\n"
        f"- 동요일 평균: {avg_val:,}\n"
        f"- 평균 대비: {diff:+,} ({diff_pct:+.1f}%)\n"
    )

    if predicted_23 is not None:
        insight += f"- 23시 예측: {predicted_23:,}\n"

    insight += (
        "\n### 해석\n\n"
        "- 0(미입력) 구간은 예측값으로 보완됩니다.\n"
        "- 실제 값이 업데이트되면 예측/분석이 다시 계산됩니다.\n"
    )

    return Response({"success": True, "insight": insight}, status=status.HTTP_200_OK)


@api_view(["GET", "POST"])
def delivery_notes(request):
    if request.method == "GET":
        date_str = (request.query_params.get("date") or "").strip()
        if not date_str:
            return Response(
                {"success": False, "message": "date query param required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        qs = DeliverySpecialNote.objects.filter(date=date_str).order_by(
            "-event_datetime", "-created_at"
        )
        return Response(
            {"success": True, "data": DeliverySpecialNoteSerializer(qs, many=True).data}
        )

    payload = request.data if isinstance(request.data, dict) else {}
    date_str = (payload.get("date") or "").strip()
    if not date_str:
        return Response(
            {"success": False, "message": "date required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    event_dt_raw = (
        payload.get("event_datetime")
        if "event_datetime" in payload
        else payload.get("eventDateTime")
    )
    event_dt = None
    if isinstance(event_dt_raw, str) and event_dt_raw.strip():
        try:
            event_dt = datetime.fromisoformat(event_dt_raw.strip())
        except Exception:
            event_dt = None

    quantity = payload.get("quantity")
    qty_val = None
    if quantity is not None and str(quantity).strip() != "":
        try:
            qty_val = int(float(str(quantity).replace(",", "")))
        except Exception:
            qty_val = None

    note = DeliverySpecialNote.objects.create(
        date=date_str,
        event_datetime=event_dt,
        product_name=str(
            payload.get("product_name") or payload.get("productName") or ""
        ).strip(),
        barcode=(
            str(payload.get("barcode")).strip()
            if payload.get("barcode") is not None
            else None
        )
        or None,
        sku_id=str(payload.get("sku_id") or payload.get("skuId") or "").strip(),
        quantity=qty_val,
        memo=str(payload.get("memo") or payload.get("text") or "").strip(),
    )
    return Response(
        {"success": True, "data": DeliverySpecialNoteSerializer(note).data},
        status=status.HTTP_201_CREATED,
    )


def _delivery_day_final(record) -> int:
    """
    일 마감 출고량 SoT = hour_23 (또는 마지막 누적 hour).
    주의: 일부 날짜 total 이 hour_23 의 ~9배로 오염됨 → total 신뢰 금지.
    """
    hourly = getattr(record, "hourly", None) or {}
    h23 = int(hourly.get("hour_23") or 0)
    if h23 > 0:
        return h23
    mx = 0
    for h in range(24):
        mx = max(mx, int(hourly.get(f"hour_{h:02d}") or 0))
    if mx > 0:
        return mx
    # total 은 hour 와 거의 같을 때만 사용 (오염 방지)
    t = int(getattr(record, "total", 0) or 0)
    return t if 0 < t < 2500 else 0


def _delivery_hour_value(record, hour: int) -> int:
    hourly = getattr(record, "hourly", None) or {}
    return int(hourly.get(f"hour_{hour:02d}") or 0)


@api_view(["POST"])
def ai_predict_hourly(request):
    """
    시간별 마감(23시) 예측.
    - 일 최종 = hour_23 (total 필드 오염 무시)
    - 완료비율 + 잔여증가 중앙값 블렌드
    - 동일 요일 hour_23 중앙값으로 하한 보정 (과소 예측 방지)
    """
    import statistics
    from datetime import timedelta

    payload = request.data if isinstance(request.data, dict) else {}

    current_hour = payload.get("currentHour")
    current_data = (
        payload.get("currentData")
        if isinstance(payload.get("currentData"), dict)
        else {}
    )
    try:
        current_hour_int = max(0, min(23, int(current_hour)))
    except Exception:
        current_hour_int = 0

    # 현재 누적: hour_XX 우선, total 은 오염 가능
    total_int = 0
    hour_key = f"hour_{current_hour_int:02d}"
    try:
        for h in range(current_hour_int, -1, -1):
            k = f"hour_{h:02d}"
            v = int(float(str(current_data.get(k) or 0).replace(",", "")))
            if v > 0:
                total_int = v
                current_hour_int = h
                break
    except Exception:
        total_int = 0
    if total_int <= 0:
        try:
            t = int(float(str(current_data.get("total") or 0).replace(",", "")))
            if 0 < t < 2500:
                total_int = t
        except Exception:
            pass

    today = timezone.localdate()
    lookback = today - timedelta(days=84)
    past_records = list(
        DeliveryDailyRecord.objects.filter(date__gte=lookback, date__lt=today).order_by(
            "-date"
        )
    )

    current_weekday = today.weekday()

    def get_period(day):
        if day <= 7:
            return "month_start"
        if day >= 22:
            return "month_end"
        return "month_mid"

    current_period = get_period(today.day)

    same_day_all = [r for r in past_records if r.date.weekday() == current_weekday]
    same_finals = sorted(
        f for f in (_delivery_day_final(r) for r in same_day_all) if f > 0
    )
    same_med = float(statistics.median(same_finals)) if same_finals else 0.0

    def _is_sane(r):
        f = _delivery_day_final(r)
        return f > 0 and (same_med <= 0 or f <= same_med * 2.5)

    same_day_records = [r for r in same_day_all if _is_sane(r)]
    period_records = [
        r for r in same_day_records if get_period(r.date.day) == current_period
    ]
    if len(period_records) < 3:
        period_records = same_day_records
    if len(period_records) < 3:
        period_records = [r for r in past_records if _delivery_day_final(r) > 0]

    # ── 1) 완료비율법 (hour_h / hour_23) ──
    ratios = []
    for r in period_records:
        hv = _delivery_hour_value(r, current_hour_int)
        fv = _delivery_day_final(r)
        if hv > 0 and fv >= hv and fv > 0:
            ratios.append(hv / fv)

    predicted_from_ratio = None
    ratio_med = None
    if ratios and total_int > 0:
        ratio_med = statistics.median(ratios)
        if 0.05 <= ratio_med <= 0.98:
            predicted_from_ratio = int(round(total_int / ratio_med))

    # ── 2) 잔여 증가분 중앙값 ──
    increments = []
    for r in period_records:
        hv = _delivery_hour_value(r, current_hour_int)
        fv = _delivery_day_final(r)
        if hv > 0 and fv > hv:
            inc = fv - hv
            if same_med > 0 and inc > same_med * 0.9:
                continue
            if inc > 500:
                continue
            weight = 2 if (today - r.date).days <= 21 else 1
            increments.extend([inc] * weight)

    if increments:
        predicted_increment = int(statistics.median(increments))
    else:
        # 조기 시간대 기본 잔여 (동일 요일 중앙 기준)
        if same_med > 0 and total_int > 0:
            predicted_increment = max(30, int(same_med - total_int))
        else:
            predicted_increment = 80

    predicted_from_inc = total_int + max(0, predicted_increment)

    # 블렌드
    if predicted_from_ratio is not None:
        # 오전(비율 불안정)은 잔여/중앙값 비중↑
        w_ratio = 0.55 if current_hour_int < 12 else 0.75
        predicted_total = int(
            w_ratio * predicted_from_ratio + (1 - w_ratio) * predicted_from_inc
        )
        model_name = "completion_ratio_h23_v3"
    else:
        predicted_total = predicted_from_inc
        model_name = "residual_h23_v3"

    # 동일 요일 hour_23 중앙값 앵커 (과소 예측 핵심 방지)
    if same_med > 0:
        # 하한: 중앙값의 92% 또는 현재+여유
        floor = int(max(total_int + 20, same_med * 0.92))
        # 상한: 중앙값 130%
        ceil = int(same_med * 1.30)
        predicted_total = max(floor, min(int(predicted_total), ceil))
        # 현재 진행이 중앙 대비 빠르면 상향
        if ratio_med and ratio_med > 0 and total_int > 0:
            pace = (total_int / max(same_med, 1)) / max(ratio_med, 0.05)
            if pace > 1.05:
                predicted_total = max(
                    predicted_total, int(min(ceil, total_int / ratio_med))
                )

    predicted_total = max(int(predicted_total), total_int)

    if current_hour_int >= 23 and total_int > 0:
        predicted_total = total_int

    return Response(
        {
            "success": True,
            "predictions": {"hour_23": int(predicted_total)},
            "metadata": {
                "model": model_name,
                "period": current_period,
                "current_hour": current_hour_int,
                "current_total": total_int,
                "same_dow_median_h23": int(same_med) if same_med else 0,
                "completion_ratio_median": round(float(ratio_med), 4)
                if ratio_med
                else None,
                "predicted_from_ratio": predicted_from_ratio,
                "predicted_from_increment": predicted_from_inc,
                "increment_samples": len(increments),
                "same_day_records": len(same_day_records),
                "server_today": today.isoformat(),
                "note": "day_final=hour_23 (total field ignored when polluted)",
            },
        },
        status=status.HTTP_200_OK,
    )


@api_view(["GET"])
def get_outbound_meta(request):
    meta = OutboundRecord.objects.aggregate(
        earliestDate=Min("outbound_date"),
        latestDate=Max("outbound_date"),
    )

    earliest = meta.get("earliestDate")
    latest = meta.get("latestDate")

    return Response(
        {
            "earliestDate": earliest.isoformat() if earliest else None,
            "latestDate": latest.isoformat() if latest else None,
        }
    )


@api_view(["GET"])
def get_outbound_stats(request):
    start = request.query_params.get("start") or request.query_params.get("startDate")
    end = request.query_params.get("end") or request.query_params.get("endDate")
    group_by = request.query_params.get("groupBy", "day")
    category = request.query_params.get("category")
    search = request.query_params.get("search")
    product = request.query_params.get("product")
    barcode = (request.query_params.get("barcode") or "").strip()

    queryset = OutboundRecord.objects.all()

    if start:
        queryset = queryset.filter(outbound_date__gte=start)
    if end:
        queryset = queryset.filter(outbound_date__lte=end)

    if category and category != "all":
        if category == "__others__":
            top_cats = list(
                queryset.values("category")
                .annotate(salesAmount=Sum("sales_amount"))
                .order_by("-salesAmount")
                .values_list("category", flat=True)[:10]
            )
            if top_cats:
                queryset = queryset.exclude(category__in=top_cats)
        else:
            queryset = queryset.filter(category=category)
    if search:
        queryset = queryset.filter(product_name__icontains=search)
    if product:
        queryset = queryset.filter(product_name=product)
    if barcode:
        queryset = queryset.filter(barcode=barcode)

    # 1. Summary
    summary = queryset.aggregate(
        totalCount=Count("id"),
        totalQuantity=Coalesce(Sum("box_quantity"), 0),
        totalSalesAmount=Coalesce(Sum("sales_amount"), Decimal("0")),
        estimatedCount=Count("id", filter=Q(is_estimated=True)),
        estimatedDays=Count("outbound_date", filter=Q(is_estimated=True), distinct=True),
    )

    # 2. Daily Trend (Group By)
    # SQLite supports TruncDate, TruncMonth etc.
    trunc_func = TruncDay
    if group_by == "week":
        trunc_func = TruncWeek
    elif group_by == "month":
        trunc_func = TruncMonth
    elif group_by == "year":
        trunc_func = TruncYear

    daily_trend = (
        queryset.annotate(date=trunc_func("outbound_date"))
        .values("date")
        .annotate(
            quantity=Coalesce(Sum("box_quantity"), 0),
            salesAmount=Coalesce(Sum("sales_amount"), Decimal("0")),
            estimated_count=Count("id", filter=Q(is_estimated=True)),
            real_count=Count("id", filter=Q(is_estimated=False)),
        )
        .order_by("date")
    )

    # Format date for frontend
    # isEstimated: 해당 버킷이 예측 보정만 있음 (전부 노랑)
    # hasEstimated: 실적+보정이 섞임 (노랑 표시 + "포함" 안내)
    trend_data = []
    for item in daily_trend:
        if item["date"]:
            est_n = int(item.get("estimated_count") or 0)
            real_n = int(item.get("real_count") or 0)
            is_est_only = est_n > 0 and real_n == 0
            has_est = est_n > 0 and real_n > 0
            trend_data.append(
                {
                    "date": item["date"].strftime("%Y-%m-%d"),
                    "quantity": item["quantity"] or 0,
                    "salesAmount": item["salesAmount"] or 0,
                    "isEstimated": is_est_only,
                    "hasEstimated": has_est,
                    "is_estimated": is_est_only,
                    "has_estimated": has_est,
                }
            )

    # 2-2. Previous Year Trend (선택 기간의 1년 전 동일 구간, 동일 필터)
    prev_year_trend = []
    if start and end:
        try:
            start_dt = datetime.strptime(start, "%Y-%m-%d")
            end_dt = datetime.strptime(end, "%Y-%m-%d")
            # 윤년(2/29) 대비: replace 실패 시 하루 줄여서 재시도
            try:
                prev_start_dt = start_dt.replace(year=start_dt.year - 1)
            except ValueError:
                prev_start_dt = start_dt.replace(year=start_dt.year - 1, day=28)
            try:
                prev_end_dt = end_dt.replace(year=end_dt.year - 1)
            except ValueError:
                prev_end_dt = end_dt.replace(year=end_dt.year - 1, day=28)

            prev_start = prev_start_dt.strftime("%Y-%m-%d")
            prev_end = prev_end_dt.strftime("%Y-%m-%d")

            # 동일 필터(category, search, product)를 전년 구간에 적용
            prev_queryset = OutboundRecord.objects.all()
            prev_queryset = prev_queryset.filter(outbound_date__gte=prev_start)
            prev_queryset = prev_queryset.filter(outbound_date__lte=prev_end)
            if category and category != "all":
                if category == "__others__":
                    prev_top_cats = list(
                        prev_queryset.values("category")
                        .annotate(salesAmount=Sum("sales_amount"))
                        .order_by("-salesAmount")
                        .values_list("category", flat=True)[:10]
                    )
                    if prev_top_cats:
                        prev_queryset = prev_queryset.exclude(category__in=prev_top_cats)
                else:
                    prev_queryset = prev_queryset.filter(category=category)
            if search:
                prev_queryset = prev_queryset.filter(product_name__icontains=search)
            if product:
                prev_queryset = prev_queryset.filter(product_name=product)

            prev_daily = (
                prev_queryset.annotate(date=trunc_func("outbound_date"))
                .values("date")
                .annotate(
                    quantity=Coalesce(Sum("box_quantity"), 0),
                    salesAmount=Coalesce(Sum("sales_amount"), Decimal("0")),
                )
                .order_by("date")
            )
            for item in prev_daily:
                if item["date"]:
                    prev_year_trend.append(
                        {
                            "date": item["date"].strftime("%Y-%m-%d"),
                            "quantity": item["quantity"] or 0,
                            "salesAmount": item["salesAmount"] or 0,
                        }
                    )
        except (ValueError, TypeError):
            prev_year_trend = []

    # 3. Category Breakdown
    category_breakdown = (
        queryset.values("category")
        .annotate(
            quantity=Coalesce(Sum("box_quantity"), 0),
            salesAmount=Coalesce(Sum("sales_amount"), Decimal("0")),
        )
        .order_by("-salesAmount")
    )

    return Response(
        {
            "summary": {
                "totalCount": summary["totalCount"] or 0,
                "totalQuantity": summary["totalQuantity"] or 0,
                "totalSalesAmount": summary["totalSalesAmount"] or 0,
                "estimatedCount": summary.get("estimatedCount") or 0,
                "estimatedDays": summary.get("estimatedDays") or 0,
            },
            "estimateMeta": {
                "rows": summary.get("estimatedCount") or 0,
                "days": summary.get("estimatedDays") or 0,
            },
            "dailyTrend": trend_data,
            "prevYearTrend": prev_year_trend,
            "categoryBreakdown": category_breakdown,
        }
    )


@api_view(["GET"])
def get_outbound_top_products(request):
    start = request.query_params.get("start") or request.query_params.get("startDate")
    end = request.query_params.get("end") or request.query_params.get("endDate")
    category = request.query_params.get("category")
    search = request.query_params.get("search")
    product = request.query_params.get("product")
    try:
        limit = int(request.query_params.get("limit") or 100)
    except Exception:
        limit = 100
    limit = max(1, min(limit, 500))

    queryset = OutboundRecord.objects.all()
    if start:
        queryset = queryset.filter(outbound_date__gte=start)
    if end:
        queryset = queryset.filter(outbound_date__lte=end)
    if category and category != "all":
        if category == "__others__":
            top_cats = list(
                queryset.values("category")
                .annotate(salesAmount=Sum("sales_amount"))
                .order_by("-salesAmount")
                .values_list("category", flat=True)[:10]
            )
            if top_cats:
                queryset = queryset.exclude(category__in=top_cats)
        else:
            queryset = queryset.filter(category=category)
    if search:
        queryset = queryset.filter(product_name__icontains=search)
    if product:
        queryset = queryset.filter(product_name=product)

    rows = (
        queryset.values("product_name")
        .annotate(
            quantity=Coalesce(Sum("box_quantity"), 0),
            salesAmount=Coalesce(Sum("sales_amount"), Decimal("0")),
        )
        .order_by("-quantity")[:limit]
    )

    return Response(
        [
            {
                "name": r.get("product_name") or "-",
                "quantity": r.get("quantity") or 0,
                "salesAmount": r.get("salesAmount") or 0,
            }
            for r in rows
        ]
    )


@api_view(["GET"])
def get_outbound_pivot(request):
    start = request.query_params.get("start") or request.query_params.get("startDate")
    end = request.query_params.get("end") or request.query_params.get("endDate")
    row = request.query_params.get("row", "category")
    group_by = request.query_params.get("groupBy", "day")
    category = request.query_params.get("category")
    search = request.query_params.get("search")
    product = request.query_params.get("product")
    try:
        limit = int(request.query_params.get("limit") or 100)
    except Exception:
        limit = 100
    limit = max(1, min(limit, 500))

    if row not in ["category", "product"]:
        return Response(
            {"message": "row must be category or product"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if group_by not in ["day", "week", "month", "year"]:
        return Response(
            {"message": "groupBy must be day, week, month, or year"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    queryset = OutboundRecord.objects.all()
    if start:
        queryset = queryset.filter(outbound_date__gte=start)
    if end:
        queryset = queryset.filter(outbound_date__lte=end)
    if category and category != "all":
        if category == "__others__":
            top_cats = list(
                queryset.values("category")
                .annotate(salesAmount=Sum("sales_amount"))
                .order_by("-salesAmount")
                .values_list("category", flat=True)[:10]
            )
            if top_cats:
                queryset = queryset.exclude(category__in=top_cats)
        else:
            queryset = queryset.filter(category=category)
    if search:
        queryset = queryset.filter(product_name__icontains=search)
    if product:
        queryset = queryset.filter(product_name=product)

    trunc_func = TruncDay
    if group_by == "week":
        trunc_func = TruncWeek
    elif group_by == "month":
        trunc_func = TruncMonth
    elif group_by == "year":
        trunc_func = TruncYear

    row_field = "category" if row == "category" else "product_name"

    if row_field == "product_name" and not product:
        top_products = list(
            queryset.values("product_name")
            .annotate(salesAmount=Sum("sales_amount"))
            .order_by("-salesAmount")
            .values_list("product_name", flat=True)[:limit]
        )
        if top_products:
            queryset = queryset.filter(product_name__in=top_products)
    grouped = (
        queryset.annotate(date=trunc_func("outbound_date"))
        .values(row_field, "date")
        .annotate(
            quantity=Coalesce(Sum("box_quantity"), 0),
            salesAmount=Coalesce(Sum("sales_amount"), Decimal("0")),
        )
        .order_by(row_field, "date")
    )

    pivot = {}
    for item in grouped:
        key = item.get(row_field) or "-"
        date_val = item.get("date")
        if not date_val:
            continue

        # Format date_key based on group_by
        if group_by == "month":
            date_key = date_val.strftime("%Y-%m")
        elif group_by == "week":
            date_key = date_val.strftime("%Y-%m-%d")  # Week start date
        else:  # day
            date_key = date_val.strftime("%Y-%m-%d")

        if key not in pivot:
            pivot[key] = {"values": {}, "total": {"quantity": 0, "salesAmount": 0}}

        q = item.get("quantity") or 0
        s = item.get("salesAmount") or 0

        pivot[key]["values"][date_key] = {
            "quantity": q,
            "salesAmount": s,
        }
        pivot[key]["total"]["quantity"] += q
        pivot[key]["total"]["salesAmount"] += s

    rows = [
        {
            "key": key,
            "values": data["values"],
            "total": data["total"],
        }
        for key, data in pivot.items()
    ]
    rows.sort(key=lambda r: r.get("total", {}).get("salesAmount", 0) or 0, reverse=True)
    return Response(rows)


import openpyxl
import io
import csv
import zipfile


@api_view(["POST"])
def parse_excel_delivery(request):
    if "file" not in request.FILES:
        return Response(
            {"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST
        )

    file_obj = request.FILES["file"]

    try:
        # Read file into BytesIO
        file_content = file_obj.read()

        if len(file_content) == 0:
            return Response({"error": "Empty file"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            sheet = wb.active
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                rows.append(cleaned_row)
        except (zipfile.BadZipFile, OSError) as e:
            # Fallback to CSV parsing
            try:
                # Decode bytes to string (assume utf-8, fallback to cp949/euc-kr if needed)
                try:
                    text_content = file_content.decode("utf-8")
                except UnicodeDecodeError:
                    text_content = file_content.decode("cp949")

                csv_reader = csv.reader(io.StringIO(text_content))
                rows = list(csv_reader)
            except Exception as csv_e:
                raise e  # Re-raise original Excel error if CSV also fails

        return Response({"rows": rows})
    except Exception as e:
        return Response(
            {"error": f"Failed to parse file: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


def _normalize_stock_status(current_stock: int, minimum_stock: int) -> str:
    if current_stock <= 0:
        return "critical"
    if current_stock <= minimum_stock:
        return "low"
    max_stock = (minimum_stock or 0) * 3
    if max_stock > 0 and current_stock > max_stock:
        return "high"
    return "normal"


def _order_recommendation(stock_status: str) -> str:
    if stock_status == "critical":
        return "즉시 발주"
    if stock_status == "low":
        return "발주 권장"
    if stock_status == "high":
        return "과재고"
    return "적정"


@api_view(["GET"])
def inventory_integrated(request):
    search = (request.query_params.get("search") or "").strip()
    stock_status = (request.query_params.get("stockStatus") or "").strip()

    all_qs = InventoryItem.objects.all()
    qs = all_qs.order_by("name")
    if search:
        qs = qs.filter(
            models.Q(name__icontains=search)
            | models.Q(category__icontains=search)
            | models.Q(barcode__icontains=search)
        )

    items = []
    for item in qs:
        current_stock = int(item.current_stock or 0)
        min_stock = int(item.minimum_stock or 0)
        max_stock = min_stock * 3
        computed_status = _normalize_stock_status(current_stock, min_stock)
        is_order_required = computed_status in ("critical", "low")

        payload = {
            "id": str(item.id),
            "productName": item.name,
            "barcode": item.barcode or None,
            "currentStock": current_stock,
            "minStock": min_stock,
            "maxStock": max_stock,
            "stockStatus": computed_status,
            "reliability": 100,
            "location": "창고 A",
            "category": item.category,
            "lastUpdated": item.updated_at.isoformat() if item.updated_at else None,
            "orderRecommendation": _order_recommendation(computed_status),
            "isOrderRequired": is_order_required,
            "hasInventoryData": True,
            "inventoryId": str(item.id),
            "hasBarcodeMaster": bool(item.barcode),
            "createdAt": item.created_at.isoformat()
            if item.created_at
            else timezone.now().isoformat(),
            "updatedAt": item.updated_at.isoformat()
            if item.updated_at
            else timezone.now().isoformat(),
        }

        if stock_status:
            if stock_status == "order_required" and not is_order_required:
                continue
            if (
                stock_status in ("critical", "low", "normal", "high")
                and payload["stockStatus"] != stock_status
            ):
                continue
            if stock_status == "no_barcode" and payload["barcode"]:
                continue

        items.append(payload)

    summary = {
        "totalItems": all_qs.count(),
        "filteredItems": len(items),
        "orderRequired": sum(1 for i in items if i["isOrderRequired"]),
        "criticalStock": sum(1 for i in items if i["stockStatus"] == "critical"),
        "lowStock": sum(1 for i in items if i["stockStatus"] == "low"),
        "highStock": sum(1 for i in items if i["stockStatus"] == "high"),
        "withoutBarcode": sum(1 for i in items if not i["barcode"]),
        "withInventoryData": len(items),
    }

    return Response(
        {
            "items": items,
            "summary": summary,
            "message": "재고 데이터를 성공적으로 불러왔습니다.",
        }
    )


def _decode_bytes(data: bytes) -> str:
    try:
        return data.decode("utf-8").replace("\ufeff", "")
    except UnicodeDecodeError:
        return data.decode("cp949").replace("\ufeff", "")


@api_view(["POST"])
def inventory_import_csv(request):
    files = request.FILES.getlist("csv")
    if not files:
        return Response(
            {"message": "파일이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST
        )

    total_processed = 0
    for file_obj in files:
        content = _decode_bytes(file_obj.read())
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if len(rows) < 2:
            continue
        headers = [h.strip() for h in rows[0]]
        normalized = [h.lower().replace(" ", "") for h in headers]

        name_idx = next(
            (
                i
                for i, h in enumerate(normalized)
                if "name" in h or "상품명" in h or "품목" in h
            ),
            -1,
        )
        cat_idx = next(
            (
                i
                for i, h in enumerate(normalized)
                if "category" in h or "분류" in h or "카테고리" in h
            ),
            -1,
        )
        stock_idx = next(
            (i for i, h in enumerate(normalized) if "stock" in h or "재고" in h), -1
        )
        barcode_idx = next(
            (i for i, h in enumerate(normalized) if "barcode" in h or "바코드" in h), -1
        )

        if name_idx == -1:
            continue

        for row in rows[1:]:
            if not row or all(not str(c).strip() for c in row):
                continue
            name = row[name_idx].strip() if name_idx < len(row) else ""
            if not name:
                continue
            category = (
                row[cat_idx].strip() if cat_idx > -1 and cat_idx < len(row) else "기타"
            )
            barcode = (
                row[barcode_idx].strip()
                if barcode_idx > -1 and barcode_idx < len(row)
                else None
            )
            try:
                current_stock = (
                    int(str(row[stock_idx]).replace(",", "").strip())
                    if stock_idx > -1 and stock_idx < len(row)
                    else 0
                )
            except Exception:
                current_stock = 0

            defaults = {
                "category": category or "기타",
                "current_stock": current_stock,
            }
            if barcode:
                defaults["barcode"] = barcode

            InventoryItem.objects.update_or_create(
                name=name,
                defaults=defaults,
            )
            total_processed += 1

    return Response(
        {
            "message": "파일이 성공적으로 업로드되었습니다.",
            "rowsProcessed": total_processed,
        }
    )


@api_view(["GET"])
def inventory_template(request):
    content = "name,category,current_stock,minimum_stock,barcode\n"
    resp = HttpResponse(content, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = 'attachment; filename="inventory_template.csv"'
    return resp


@api_view(["POST"])
def inventory_apply_calculated_thresholds(request):
    payload = request.data if isinstance(request.data, dict) else {}
    products = payload.get("products")
    if not isinstance(products, list):
        return Response(
            {"error": "products must be a list"}, status=status.HTTP_400_BAD_REQUEST
        )

    products = [p.strip() for p in products if isinstance(p, str) and p.strip()]
    if not products:
        return Response({"success": True, "applied": 0}, status=status.HTTP_200_OK)

    period = (payload.get("period") or "3month").strip()
    days = 90
    if period == "1month":
        days = 30
    elif period == "6month":
        days = 180

    since = timezone.localdate() - timedelta(days=days)
    qs = (
        OutboundRecord.objects.filter(outbound_date__gte=since)
        .exclude(barcode__isnull=True)
        .exclude(barcode="")
        .filter(barcode__in=products)
    )

    daily = qs.values("barcode", "outbound_date").annotate(
        qty=Coalesce(Sum("box_quantity"), 0)
    )

    by_barcode = {}
    for row in daily:
        bc = row.get("barcode")
        if not bc:
            continue
        agg = by_barcode.get(bc)
        if not agg:
            agg = {"total": 0, "days": set()}
            by_barcode[bc] = agg
        agg["total"] += int(row.get("qty") or 0)
        if row.get("outbound_date"):
            agg["days"].add(row["outbound_date"])

    applied = 0
    for bc in products:
        agg = by_barcode.get(bc) or {"total": 0, "days": set()}
        days_count = len(agg["days"])
        avg_daily = (agg["total"] / float(days_count)) if days_count > 0 else 0.0
        min_stock = int(round(avg_daily * 3))
        max_stock = int(round(avg_daily * 30))
        reorder_point = int(round(avg_daily * 3))

        bm, _created = BarcodeMaster.objects.get_or_create(barcode=bc)
        bm.min_stock = min_stock
        bm.max_stock = max_stock
        bm.reorder_point = reorder_point
        bm.save(update_fields=["min_stock", "max_stock", "reorder_point", "updated_at"])
        applied += 1

    return Response({"success": True, "applied": applied}, status=status.HTTP_200_OK)


def _delivery_row_to_payload(record: DeliveryDailyRecord) -> dict:
    hourly = record.hourly or {}
    payload = {
        "date": record.date.isoformat(),
        "dayOfWeek": record.day_of_week or "",
        "total": int(record.total or 0),
    }
    for h in range(24):
        key = f"hour_{h:02d}"
        payload[key] = int(hourly.get(key) or 0)
    return payload


def _recompute_delivery_total(hourly: dict) -> int:
    for h in range(23, -1, -1):
        key = f"hour_{h:02d}"
        try:
            val = int(hourly.get(key) or 0)
        except Exception:
            val = 0
        if val > 0:
            return val
    return 0


@api_view(["GET", "POST"])
def delivery_hourly(request):
    if request.method == "GET":
        days = request.query_params.get("days")
        try:
            days_int = int(days) if days else 365
        except Exception:
            days_int = 365

        cutoff = timezone.localdate() - timedelta(days=days_int)
        qs = DeliveryDailyRecord.objects.filter(date__gte=cutoff).order_by("date")
        return Response(
            {
                "success": True,
                "data": [_delivery_row_to_payload(r) for r in qs],
            }
        )

    entries = request.data
    if not isinstance(entries, list) or len(entries) == 0:
        return Response(
            {"success": False, "message": "entries array required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    today = timezone.localdate()
    record, _created = DeliveryDailyRecord.objects.get_or_create(date=today)

    hourly = record.hourly or {}
    for entry in entries:
        hour = entry.get("hour")
        quantity = entry.get("quantity")
        try:
            h = int(hour)
        except Exception:
            continue
        if h < 0 or h > 23:
            continue
        try:
            q = int(quantity)
        except Exception:
            q = 0
        hourly[f"hour_{h:02d}"] = q

    record.hourly = hourly
    record.total = _recompute_delivery_total(hourly)
    record.save()

    return Response(
        {
            "success": True,
            "date": record.date.isoformat(),
            "row": _delivery_row_to_payload(record),
        }
    )


@api_view(["GET"])
def delivery_range(request):
    start = request.query_params.get("start")
    end = request.query_params.get("end")
    if not start or not end:
        return Response(
            {"success": False, "message": "start and end required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    qs = DeliveryDailyRecord.objects.filter(date__gte=start, date__lte=end).order_by(
        "date"
    )
    data = [_delivery_row_to_payload(r) for r in qs]
    return Response(
        {
            "success": True,
            "start": start,
            "end": end,
            "count": len(data),
            "data": data,
        }
    )


@api_view(["GET"])
def delivery_weekday_hourly_ratio(request):
    """
    요일별 시간대별 비율 API
    - 최근 4주 데이터에서 각 요일의 시간대별 비율 계산
    - 예측 시 시간대별 패턴 파악용
    Response: {
      weekday: {  # 0=일, 1=월, ..., 6=토
        hour_00: ratio,
        hour_01: ratio,
        ...
      }
    }
    """
    from datetime import timedelta

    today = timezone.localdate()
    four_weeks_ago = today - timedelta(days=28)

    records = DeliveryDailyRecord.objects.filter(
        date__gte=four_weeks_ago, date__lt=today, total__gt=0
    )

    # 요일별 합계 (각 시간대별)
    weekday_hourly_sums = {i: [0] * 24 for i in range(7)}
    weekday_totals = {i: 0 for i in range(7)}

    for record in records:
        if not record.hourly:
            continue
        day = record.date.weekday()
        day_total = int(record.total or 0)
        if day_total <= 0:
            continue

        weekday_totals[day] += day_total
        hourly = record.hourly or {}
        for h in range(24):
            key = f"hour_{h:02d}"
            val = int(hourly.get(key, 0))
            weekday_hourly_sums[day][h] += val

    # 비율 계산
    result = {}
    day_names = ["일", "월", "화", "수", "목", "금", "토"]

    for day in range(7):
        if weekday_totals[day] <= 0:
            # 기본 비율 (보통 낮 12-14시, 저녁 17-20시巅峰)
            result[day_names[day]] = {
                "hour_00": 0.01,
                "hour_01": 0.01,
                "hour_02": 0.01,
                "hour_03": 0.01,
                "hour_04": 0.01,
                "hour_05": 0.01,
                "hour_06": 0.02,
                "hour_07": 0.03,
                "hour_08": 0.04,
                "hour_09": 0.05,
                "hour_10": 0.06,
                "hour_11": 0.07,
                "hour_12": 0.08,
                "hour_13": 0.07,
                "hour_14": 0.06,
                "hour_15": 0.05,
                "hour_16": 0.06,
                "hour_17": 0.08,
                "hour_18": 0.07,
                "hour_19": 0.05,
                "hour_20": 0.04,
                "hour_21": 0.03,
                "hour_22": 0.02,
                "hour_23": 0.02,
            }
            continue

        result[day_names[day]] = {}
        for h in range(24):
            ratio = weekday_hourly_sums[day][h] / weekday_totals[day]
            result[day_names[day]][f"hour_{h:02d}"] = round(ratio, 4)

    return Response(
        {
            "success": True,
            "data": result,
            "meta": {
                "days": 28,
                "min_records": min(weekday_totals.values()) if weekday_totals else 0,
            },
        }
    )


@api_view(["GET"])
def delivery_daily_prediction(request):
    """
    2-stage 일별 예측 API (고도화 버전)
    - Stage 1: 이동 평균 + 절사 평균으로 일별 총량 예측
    - Stage 2: 시간대별 비율로 시간별 분포 예측

    개선 사항:
    - 데이터 개수에 따른 절사 평균 분기 처리
    - 상하위 이상치 동시 처리 (상한 200%, 하한 30%)
    - 동적 월초말계수 (과거 데이터에서 자동 계산)

    Query params:
    - days: 학습 데이터 일수 (default: 90)
    - start_date: 예측 시작 날짜 (default: tomorrow)
    - num_days: 예측할 일수 (default: 1, 최대 7)

    Response: {
      predictions: [
        { date, predicted_total, confidence, day_of_week, product_predictions: [...] },
        ...
      ],
      hourly_predictions: {
        'YYYY-MM-DD': { hour_00, hour_01, ... },
        ...
      }
    }
    """
    from datetime import timedelta
    import numpy as np

    days = request.query_params.get("days", "90")
    start_date_str = request.query_params.get("start_date")
    num_days_str = request.query_params.get("num_days", "1")

    try:
        days_int = int(days)
    except:
        days_int = 90

    try:
        num_days = min(int(num_days_str), 7)  # 최대 7일
    except:
        num_days = 1

    # 시작 날짜 결정 (기본: 내일)
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        except:
            start_date = timezone.localdate() + timedelta(days=1)
    else:
        start_date = timezone.localdate() + timedelta(days=1)

    # 학습 데이터 조회 (total 오염 가능 → hour_23 기준 필터는 아래에서)
    cutoff = timezone.localdate() - timedelta(days=days_int)
    records = list(
        DeliveryDailyRecord.objects.filter(
            date__gte=cutoff, date__lt=timezone.localdate()
        ).order_by("date")
    )
    # hour_23 기준 유효 일만
    records = [r for r in records if _delivery_day_final(r) > 0]

    if len(records) < 7:
        return Response(
            {"success": False, "message": "학습 데이터가 부족합니다 (최소 7일 필요)"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def _robust_median(values):
        """이상치 완화 중앙값 (IQR 밖 제거 후 median, 부족하면 원본 median)."""
        xs = [float(v) for v in values if v is not None and v > 0]
        if not xs:
            return 0.0
        if len(xs) < 5:
            return float(np.median(xs))
        q1, q3 = np.percentile(xs, [25, 75])
        iqr = max(q3 - q1, 1.0)
        lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
        filtered = [v for v in xs if lo <= v <= hi]
        if len(filtered) < 3:
            filtered = xs
        return float(np.median(filtered))

    def calculate_dynamic_factors(records):
        """중앙값 기반 요일·월 구간 계수 — day_final=hour_23."""
        all_totals = [_delivery_day_final(r) for r in records]
        all_totals = [t for t in all_totals if t > 0]
        if not all_totals:
            return {"month": {}, "dow": {}, "overall_median": 0.0, "overall_avg": 0.0}

        overall_med = _robust_median(all_totals)
        overall_avg = float(np.mean(all_totals))
        base = overall_med if overall_med > 0 else overall_avg

        month_start = [
            _delivery_day_final(r) for r in records if r.date.day <= 7
        ]
        month_mid = [
            _delivery_day_final(r) for r in records if 7 < r.date.day < 22
        ]
        month_end = [
            _delivery_day_final(r) for r in records if r.date.day >= 22
        ]
        month_start = [x for x in month_start if x > 0]
        month_mid = [x for x in month_mid if x > 0]
        month_end = [x for x in month_end if x > 0]

        def _fac(vals):
            m = _robust_median(vals)
            if base <= 0 or m <= 0:
                return 1.0
            return float(np.clip(m / base, 0.75, 1.25))

        month_factors = {
            "month_start": round(_fac(month_start), 3),
            "month_mid": round(_fac(month_mid), 3),
            "month_end": round(_fac(month_end), 3),
        }

        dow_totals = {i: [] for i in range(7)}
        for r in records:
            f = _delivery_day_final(r)
            if f > 0:
                dow_totals[r.date.weekday()].append(f)

        defaults = {0: 1.05, 1: 1.02, 2: 1.05, 3: 1.02, 4: 1.0, 5: 0.95, 6: 0.95}
        dow_factors = {}
        for i in range(7):
            if dow_totals[i]:
                m = _robust_median(dow_totals[i])
                dow_factors[i] = (
                    round(float(np.clip(m / base, 0.70, 1.35)), 3)
                    if base > 0
                    else defaults[i]
                )
            else:
                dow_factors[i] = defaults[i]

        dow_medians = {
            i: int(_robust_median(dow_totals[i])) if dow_totals[i] else 0
            for i in range(7)
        }

        return {
            "month": month_factors,
            "dow": dow_factors,
            "dow_medians": dow_medians,
            "overall_median": overall_med,
            "overall_avg": overall_avg,
        }

    factors_data = calculate_dynamic_factors(records)
    month_period_factors = factors_data["month"]
    dynamic_dow_factors = factors_data["dow"]
    dow_medians = factors_data.get("dow_medians") or {}
    base_median = factors_data.get("overall_median") or 0
    base_average = factors_data.get("overall_avg") or base_median

    HOLIDAYS = ["01-01", "03-01", "05-05", "06-06", "08-15", "10-03", "10-09", "12-25"]

    def get_holiday_factor(date):
        date_str = f"{date.month:02d}-{date.day:02d}"
        if date_str in HOLIDAYS:
            return 0.50
        prev_date = date - timedelta(days=1)
        prev_str = f"{prev_date.month:02d}-{prev_date.day:02d}"
        if prev_str in HOLIDAYS:
            return 0.75
        next_date = date + timedelta(days=1)
        next_str = f"{next_date.month:02d}-{next_date.day:02d}"
        if next_str in HOLIDAYS:
            return 0.85
        return 1.00

    # ====== Stage 1: 총량 예측 (동일 요일 hour_23 중앙값 블렌드) ======
    all_totals = [_delivery_day_final(r) for r in records if _delivery_day_final(r) > 0]
    recent_28 = all_totals[-28:] if len(all_totals) >= 7 else all_totals
    recent_4week_med = _robust_median(recent_28) if recent_28 else base_median
    recent_4week_avg = float(np.mean(recent_28)) if recent_28 else float(base_average)
    base_trim = _robust_median(all_totals) if all_totals else base_median

    day_names = ["월", "화", "수", "목", "금", "토", "일"]
    server_today = timezone.localdate()

    predictions = []
    hourly_predictions = {}

    for day_offset in range(num_days):
        target_date = start_date + timedelta(days=day_offset)
        target_dow = target_date.weekday()
        target_day = target_date.day

        if target_day <= 7:
            period_factor = month_period_factors.get("month_start", 1.0)
            target_period = "month_start"
        elif target_day >= 22:
            period_factor = month_period_factors.get("month_end", 1.0)
            target_period = "month_end"
        else:
            period_factor = month_period_factors.get("month_mid", 1.0)
            target_period = "month_mid"

        # 동일 요일 절대 중앙값 (최근 학습 구간)
        same_dow_med = float(dow_medians.get(target_dow) or 0)
        if same_dow_med <= 0:
            same_dow_med = base_trim if base_trim > 0 else recent_4week_med

        # 블렌드: 동일요일 55% + 최근28일 중앙 25% + 전체 강건 중앙 20%
        blended = (
            0.55 * same_dow_med
            + 0.25 * (recent_4week_med or same_dow_med)
            + 0.20 * (base_trim or same_dow_med)
        )

        # 월 구간은 약보정만 (이미 중앙값 기반 factor, 1.0 근처)
        holiday_factor = get_holiday_factor(target_date)
        predicted_total = blended * float(period_factor) * float(holiday_factor)

        # 클립: 동일 요일 중앙값 기준 0.55~1.45
        if same_dow_med > 0:
            predicted_total = float(
                np.clip(predicted_total, same_dow_med * 0.55, same_dow_med * 1.45)
            )
        predicted_total = int(max(1, round(predicted_total)))

        # 표시용 요일 계수 = 동일요일중앙 / 전체중앙
        dow_factor = dynamic_dow_factors.get(target_dow, 1.0)

        weekday_ratios = _get_weekday_hourly_ratios(target_dow, records)

        # 누적 곡선: 비율을 누적 점유율로 해석할 수 있게 정규화
        # (기존 코드는 hour 합 = total 이 되도록 스케일)
        hourly_prediction = {}
        for h in range(24):
            ratio = weekday_ratios.get(f"hour_{h:02d}", 0.01)
            hourly_prediction[f"hour_{h:02d}"] = max(0, int(predicted_total * ratio))

        predicted_sum = sum(hourly_prediction.values())
        if predicted_sum > 0 and predicted_total > 0:
            scale = predicted_total / predicted_sum
            for h in range(24):
                key = f"hour_{h:02d}"
                hourly_prediction[key] = int(hourly_prediction[key] * scale)
            # 누적 단조 증가 보정
            running = 0
            for h in range(24):
                key = f"hour_{h:02d}"
                # ratios 가 증분인지 누적인지 불명 → 증분 가정 시 누적으로 재구성
                running = max(running, hourly_prediction[key])
            # 비율 합이 1에 가깝다면 증분; 아니면 누적값으로 재스케일
            # 안전하게: 시간 순 누적 최대가 predicted_total 이 되도록 선형 보간 곡선
            # 동일 요일 평균 누적 비율 재계산
            pass

        # Stage 2 재구성: 누적 비율 기반
        cum_ratios = _get_weekday_hourly_cumulative_ratios(target_dow, records)
        if cum_ratios:
            hourly_prediction = {}
            for h in range(24):
                cr = cum_ratios.get(f"hour_{h:02d}", (h + 1) / 24.0)
                hourly_prediction[f"hour_{h:02d}"] = int(round(predicted_total * cr))
            # 단조 증가
            prev = 0
            for h in range(24):
                key = f"hour_{h:02d}"
                hourly_prediction[key] = max(prev, hourly_prediction[key])
                prev = hourly_prediction[key]
            hourly_prediction["hour_23"] = predicted_total

        product_ratios = _get_product_ratios()
        product_prediction_list = []
        for p in product_ratios:
            qty = int(predicted_total * p["ratio"])
            if qty > 0:
                product_prediction_list.append(
                    {
                        "barcode": p["barcode"],
                        "product_name": p["product_name"],
                        "category": p["category"],
                        "predicted_quantity": qty,
                    }
                )
        product_prediction_list = sorted(
            product_prediction_list, key=lambda x: x["predicted_quantity"], reverse=True
        )[:10]

        predictions.append(
            {
                "date": target_date.isoformat(),
                "predicted_total": int(predicted_total),
                "day_of_week": day_names[target_dow],
                "confidence": "high"
                if len(all_totals) >= 30
                else "medium"
                if len(all_totals) >= 14
                else "low",
                "period": target_period,
                "factors": {
                    "same_dow_median": int(same_dow_med),
                    "recent_4week_median": int(recent_4week_med or 0),
                    "base_median": int(base_trim or 0),
                    "base_average": int(base_average or 0),
                    "dow_factor": round(float(dow_factor), 2),
                    "period_factor": round(float(period_factor), 2),
                    "holiday_factor": round(float(holiday_factor), 2),
                    "blended_before_period": int(blended),
                },
                "product_predictions": product_prediction_list,
            }
        )

        hourly_predictions[target_date.isoformat()] = hourly_prediction

    return Response(
        {
            "success": True,
            "predictions": predictions,
            "hourly_predictions": hourly_predictions,
            "meta": {
                "start_date": start_date.isoformat(),
                "server_today": server_today.isoformat(),
                "num_days": num_days,
                "training_samples": len(all_totals),
                "base_average": int(base_average or 0),
                "base_median": int(base_trim or 0),
                "month_period_factors": month_period_factors,
                "dow_factors": {str(k): v for k, v in dynamic_dow_factors.items()},
                "dow_medians": {str(k): v for k, v in dow_medians.items()},
                "recent_4week_avg": int(recent_4week_avg or 0),
                "recent_4week_median": int(recent_4week_med or 0),
                "model": "same_dow_median_blend_v3",
            },
        }
    )


def _get_weekday_hourly_ratios(target_dow: int, records) -> dict:
    """요일별 시간대별 비율 계산 (값/일총량 평균 — 레거시 호환)."""
    cum = _get_weekday_hourly_cumulative_ratios(target_dow, records)
    if not cum:
        return {f"hour_{h:02d}": round(1.0 / 24, 4) for h in range(24)}
    # 누적 → 대략 증분
    ratios = {}
    prev = 0.0
    for h in range(24):
        c = float(cum.get(f"hour_{h:02d}", 0))
        ratios[f"hour_{h:02d}"] = round(max(0.0, c - prev), 4)
        prev = c
    s = sum(ratios.values()) or 1.0
    return {k: round(v / s, 4) for k, v in ratios.items()}


def _get_weekday_hourly_cumulative_ratios(target_dow: int, records) -> dict:
    """
    동일 요일 시간대 누적 비율 중앙값.
    hour_h / day_total 의 중앙값, 단조 증가·hour_23=1.0 정규화.
    """
    from datetime import timedelta
    import statistics as _stats

    today = timezone.localdate()
    four_weeks_ago = today - timedelta(days=84)

    # hour -> list of ratios
    by_hour = {h: [] for h in range(24)}
    for record in records:
        if record.date < four_weeks_ago:
            continue
        if record.date.weekday() != target_dow:
            continue
        # SoT: hour_23 (total 오염 무시)
        total = _delivery_day_final(record)
        if total <= 0 or total > 2500:
            continue
        hourly = record.hourly or {}
        prev_r = 0.0
        for h in range(24):
            v = int(hourly.get(f"hour_{h:02d}", 0) or 0)
            if v <= 0 and h < 12:
                continue
            r = min(1.0, max(0.0, v / total)) if total else 0.0
            r = max(prev_r, r)
            by_hour[h].append(r)
            prev_r = r

    if not any(by_hour[h] for h in range(24)):
        # 기본 S-curve 누적
        return {f"hour_{h:02d}": round((h + 1) / 24.0, 4) for h in range(24)}

    ratios = {}
    prev = 0.0
    for h in range(24):
        xs = by_hour[h]
        if xs:
            med = float(_stats.median(xs))
        else:
            med = prev if h > 0 else 0.05
        med = max(prev, min(1.0, med))
        ratios[f"hour_{h:02d}"] = round(med, 4)
        prev = med
    # 23시 = 1.0
    ratios["hour_23"] = 1.0
    # 재단조
    prev = 0.0
    for h in range(24):
        key = f"hour_{h:02d}"
        ratios[key] = round(max(prev, float(ratios[key])), 4)
        prev = ratios[key]
    ratios["hour_23"] = 1.0
    return ratios


def _get_product_ratios() -> list:
    """최근 30일 출고 데이터를 바탕으로 품목별 비중 계산"""
    from datetime import timedelta
    from django.db.models import Sum
    from .models import OutboundRecord

    today = timezone.localdate()
    thirty_days_ago = today - timedelta(days=30)

    # 최근 30일 총 출고량
    total_qty = (
        OutboundRecord.objects.filter(
            outbound_date__gte=thirty_days_ago, outbound_date__lt=today
        ).aggregate(total=Sum("quantity"))["total"]
        or 0
    )

    if total_qty <= 0:
        return []

    # 품목별 집계
    product_stats = (
        OutboundRecord.objects.filter(
            outbound_date__gte=thirty_days_ago, outbound_date__lt=today
        )
        .values("barcode", "product_name", "category")
        .annotate(sum_qty=Sum("quantity"))
        .order_by("-sum_qty")
    )

    ratios = []
    for p in product_stats:
        ratios.append(
            {
                "barcode": p["barcode"],
                "product_name": (p["product_name"] or "Unknown"),
                "category": (p["category"] or "기타"),
                "ratio": p["sum_qty"] / total_qty,
            }
        )

    return ratios


def _upsert_delivery_from_payload(payload: dict):
    date_str = (payload.get("date") or "").strip()
    if not date_str:
        return False
    try:
        date_obj = datetime.fromisoformat(date_str).date()
    except Exception:
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            return False

    record, _created = DeliveryDailyRecord.objects.get_or_create(date=date_obj)
    hourly = record.hourly or {}
    for h in range(24):
        key = f"hour_{h:02d}"
        if key not in payload:
            continue
        try:
            hourly[key] = int(payload.get(key) or 0)
        except Exception:
            hourly[key] = 0

    total = payload.get("total")
    try:
        total_val = (
            int(total) if total is not None else _recompute_delivery_total(hourly)
        )
    except Exception:
        total_val = _recompute_delivery_total(hourly)
    if total_val <= 0:
        total_val = _recompute_delivery_total(hourly)

    day_of_week = payload.get("dayOfWeek") or payload.get("day_of_week") or ""

    record.day_of_week = str(day_of_week)
    record.total = total_val
    record.hourly = hourly
    record.save()
    return True


@api_view(["POST"])
def delivery_import(request):
    file_obj = request.FILES.get("file")
    if not file_obj:
        return Response(
            {"success": False, "message": "file field required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    name = (file_obj.name or "").lower()
    raw = file_obj.read()

    imported = 0

    if name.endswith(".json"):
        try:
            text = _decode_bytes(raw)
            obj = json.loads(text)
        except Exception as e:
            return Response(
                {"success": False, "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        arr = obj if isinstance(obj, list) else obj.get("delivery_data")
        if not isinstance(arr, list):
            return Response(
                {"success": False, "message": "Invalid JSON format"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for row in arr:
            if isinstance(row, dict) and _upsert_delivery_from_payload(row):
                imported += 1

        return Response({"success": True, "result": {"count": imported}})

    text = _decode_bytes(raw)
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) < 2:
        return Response(
            {"success": False, "message": "Empty file"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    header = [str(h).strip() for h in rows[0]]
    normalized = [h.replace(" ", "") for h in header]
    date_idx = 0

    matrix = any(h in ("00", "0", "01") for h in normalized) and any(
        h in ("23",) for h in normalized
    )

    if matrix:
        hour_map = {}
        for idx, h in enumerate(normalized):
            if h.isdigit():
                hh = int(h)
                if 0 <= hh <= 23:
                    hour_map[idx] = f"hour_{hh:02d}"

        for row in rows[1:]:
            if not row:
                continue
            date_str = str(row[date_idx]).strip() if date_idx < len(row) else ""
            if not date_str:
                continue
            payload = {"date": date_str}
            for idx, key in hour_map.items():
                if idx < len(row):
                    try:
                        payload[key] = int(str(row[idx]).replace(",", "").strip() or 0)
                    except Exception:
                        payload[key] = 0
            if _upsert_delivery_from_payload(payload):
                imported += 1
    else:
        for row in rows[1:]:
            if len(row) < 3:
                continue
            date_str = str(row[0]).strip()
            hour_str = str(row[1]).strip()
            qty_str = str(row[2]).strip()
            try:
                hour_val = int(hour_str)
            except Exception:
                continue
            if hour_val < 0 or hour_val > 23:
                continue
            try:
                qty_val = int(qty_str)
            except Exception:
                qty_val = 0
            payload = {
                "date": date_str,
                f"hour_{hour_val:02d}": qty_val,
            }
            if _upsert_delivery_from_payload(payload):
                imported += 1

    return Response({"success": True, "result": {"count": imported}})


@api_view(["POST"])
def delivery_import_excel(request):
    if "file" not in request.FILES:
        return Response(
            {"success": False, "message": "file field required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    file_obj = request.FILES["file"]
    file_content = file_obj.read()
    if len(file_content) == 0:
        return Response(
            {"success": False, "message": "Empty file"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
    except Exception as e:
        return Response(
            {"success": False, "message": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )

    if len(rows) < 2:
        return Response(
            {"success": False, "message": "Empty or invalid Excel file"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    headers = [str(h).strip() for h in rows[0]]
    imported = 0
    for values in rows[1:]:
        if not values:
            continue
        payload = {}
        for idx, header in enumerate(headers):
            if idx >= len(values):
                continue
            clean = str(header).strip()
            val = values[idx]
            if clean in ("날짜", "일자", "date"):
                payload["date"] = val
            elif clean in ("요일", "dayOfWeek"):
                payload["dayOfWeek"] = val
            elif clean in ("합계", "총계", "누적", "total"):
                payload["total"] = val
            else:
                if clean.isdigit():
                    h = int(clean)
                    if 0 <= h <= 23:
                        payload[f"hour_{h:02d}"] = val
        if _upsert_delivery_from_payload(payload):
            imported += 1
    return Response(
        {"success": True, "result": {"count": imported}, "created": imported}
    )


@api_view(["GET"])
def delivery_export_xlsx(request):
    start = request.query_params.get("start")
    end = request.query_params.get("end")
    qs = DeliveryDailyRecord.objects.all().order_by("date")
    if start:
        qs = qs.filter(date__gte=start)
    if end:
        qs = qs.filter(date__lte=end)

    wb = openpyxl.Workbook()
    ws = wb.active
    header = ["date", "dayOfWeek", "total"] + [f"{h:02d}" for h in range(24)]
    ws.append(header)

    for rec in qs:
        payload = _delivery_row_to_payload(rec)
        row = [payload["date"], payload["dayOfWeek"], payload["total"]]
        for h in range(24):
            row.append(payload[f"hour_{h:02d}"])
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="delivery_export.xlsx"'
    return resp


@api_view(["GET", "POST", "DELETE"])
def baco_transfer_stats(request):
    if request.method == "DELETE":
        BarcodeTransferRecord.objects.all().delete()
        return Response({"success": True, "message": "Transferred data cleared."})

    if request.method == "POST":
        new_data = request.data
        if not isinstance(new_data, list):
            return Response(
                {"success": False, "error": "Array expected"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        new_count = 0
        for item in new_data:
            tracking = (item.get("trackingNo") or "").strip()
            barcode = (item.get("barcode") or "").strip()
            if not tracking or not barcode:
                continue
            obj, created = BarcodeTransferRecord.objects.get_or_create(
                tracking_no=tracking,
                defaults={
                    "barcode": barcode,
                    "product_name": (item.get("productName") or "").strip(),
                    "category": (item.get("category") or "").strip(),
                },
            )
            if created:
                new_count += 1

        return Response(
            {
                "success": True,
                "message": f"Data transferred successfully. Added {new_count} new items.",
            }
        )

    raw = list(
        BarcodeTransferRecord.objects.all()
        .order_by("created_at")
        .values("tracking_no", "barcode", "product_name", "category")
    )
    aggregated = (
        BarcodeTransferRecord.objects.values("barcode", "product_name", "category")
        .annotate(count=Count("tracking_no"))
        .order_by("-count")
    )
    data = []
    for row in aggregated:
        data.append(
            {
                "barcode": row.get("barcode"),
                "productName": row.get("product_name") or "-",
                "category": row.get("category") or "-",
                "count": row.get("count") or 0,
            }
        )

    return Response(
        {
            "success": True,
            "timestamp": timezone.now().isoformat(),
            "rawData": raw,
            "data": data,
            "totalItems": BarcodeTransferRecord.objects.count(),
        }
    )


@api_view(["GET"])
def outbound_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "outbound_template"
    headers = [
        "outbound_date",
        "product_name",
        "category",
        "barcode",
        "quantity",
        "box_quantity",
        "unit_count",
        "sales_amount",
        "client",
        "status",
        "notes",
    ]
    ws.append(headers)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="outbound_upload_template.xlsx"'
    return resp


@api_view(["POST"])
def outbound_upload_excel(request):
    """
    바코드 통계 엑셀 업로드 (바코드통계_YYYYMMDD.xlsx)
    컬럼: ['바코드', '제품명', '대분류', '수량']
    파일명에서 날짜 추출 (예: 바코드통계_20260401.xlsx -> 2026-04-01)
    """
    if "file" not in request.FILES:
        return Response(
            {"success": False, "message": "file field required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    file_obj = request.FILES["file"]
    filename = file_obj.name

    # 파일명에서 날짜 추출 (바코드통계_20260401.xlsx -> 2026-04-01)
    import re

    date_match = re.search(r"(\d{4})(\d{2})(\d{2})", filename)
    if date_match:
        outbound_date = (
            f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
        )
    else:
        # 프론트엔드에서 date 파라미터로 받을 경우
        outbound_date = request.data.get("date")
        if not outbound_date:
            return Response(
                {
                    "success": False,
                    "message": "Cannot extract date from filename. Please provide date parameter.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

    file_content = file_obj.read()
    if len(file_content) == 0:
        return Response(
            {"success": False, "message": "Empty file"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
    except Exception as e:
        return Response(
            {"success": False, "message": str(e)}, status=status.HTTP_400_BAD_REQUEST
        )

    if len(rows) < 2:
        return Response(
            {"success": False, "message": "Empty or invalid Excel file"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # 헤더 파싱 (바코드, 제품명, 대분류, 수량)
    headers = [str(h).strip() for h in rows[0]]
    barcode_idx = _find_col_index(headers, ["바코드", "barcode", "BARCODE"])
    product_idx = _find_col_index(
        headers, ["제품명", "productName", "PRODUCT_NAME", "PRODUCT", "제품"]
    )
    category_idx = _find_col_index(headers, ["대분류", "category", "CATEGORY", "분류"])
    quantity_idx = _find_col_index(
        headers, ["수량", "quantity", "QUANTITY", "개수", "출고수량"]
    )

    if barcode_idx is None and product_idx is None:
        return Response(
            {
                "success": False,
                "message": "Cannot find barcode or product name columns",
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    # 해당 날짜 기존 데이터 삭제 (중복 합산 방지)
    deleted_count, _ = OutboundRecord.objects.filter(
        outbound_date=outbound_date
    ).delete()

    # 데이터 일괄 저장
    outbound_instances = []
    imported = 0
    errors = []

    for values in rows[1:]:
        if not values:
            continue
        try:
            barcode = (
                values[barcode_idx].strip()
                if barcode_idx is not None and barcode_idx < len(values)
                else ""
            )
            product_name = (
                values[product_idx].strip()
                if product_idx is not None and product_idx < len(values)
                else ""
            )
            category = (
                values[category_idx].strip()
                if category_idx is not None and category_idx < len(values)
                else "기타"
            )
            quantity_str = (
                values[quantity_idx].strip()
                if quantity_idx is not None and quantity_idx < len(values)
                else "0"
            )

            if not product_name:  # 제품명이 없으면 스킵
                continue

            # 수량 파싱
            try:
                quantity = (
                    int(quantity_str.replace(",", "").strip()) if quantity_str else 0
                )
            except ValueError:
                quantity = 0

            instance = OutboundRecord(
                outbound_date=outbound_date,
                product_name=product_name,
                barcode=barcode if barcode else None,
                category=category if category else "기타",
                quantity=quantity,
                sales_amount=0,  # 엑셀에 금액 정보 없음
                client="",
                status="완료",
            )
            outbound_instances.append(instance)
            imported += 1

        except Exception as e:
            errors.append(str(e))
            if len(errors) > 10:
                break

    # 벌크 인서트
    if outbound_instances:
        try:
            with transaction.atomic():
                OutboundRecord.objects.bulk_create(outbound_instances, batch_size=500)

                # 신규 출고 품목 자동 복구 (최근 3개월 미출고 상태 해제)
                synced_barcodes = {inst.barcode for inst in outbound_instances if inst.barcode}
                if synced_barcodes:
                    _clear_no_outbound_3m_on_activity(barcodes=synced_barcodes)
        except Exception as e:
            logger.error(f"Bulk insert failed, falling back to individual saves: {e}")
            # 폴백: 개별 저장
            for instance in outbound_instances:
                try:
                    instance.save()
                except Exception as e2:
                    errors.append(str(e2))
                    if len(errors) > 10:
                        break

    return Response(
        {
            "success": True,
            "message": f"{outbound_date} 데이터 {imported}건 저장 완료 (기존 데이터 {deleted_count}건 삭제)",
            "result": {
                "date": outbound_date,
                "imported": imported,
                "deleted": deleted_count,
                "errors": errors[:10] if errors else [],
            },
        }
    )


@api_view(["GET"])
def outbound_download_excel(request):
    start = request.query_params.get("start")
    end = request.query_params.get("end")

    qs = OutboundRecord.objects.all().order_by("outbound_date")
    if start:
        qs = qs.filter(outbound_date__gte=start)
    if end:
        qs = qs.filter(outbound_date__lte=end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "outbound"

    headers = [
        "outbound_date",
        "product_name",
        "category",
        "barcode",
        "quantity",
        "box_quantity",
        "unit_count",
        "sales_amount",
        "client",
        "status",
        "notes",
    ]
    ws.append(headers)

    for r in qs:
        ws.append(
            [
                r.outbound_date.isoformat() if r.outbound_date else "",
                r.product_name,
                r.category,
                r.barcode or "",
                r.quantity or 0,
                r.box_quantity or 0,
                r.unit_count or 0,
                float(r.sales_amount or 0),
                r.client or "",
                r.status or "",
                r.notes or "",
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="outbound_data.xlsx"'
    return resp


# ============================================================================
# Inbound Order Management (입고 가능 수량)
# ============================================================================


@api_view(["POST"])
def inbound_order_upload(request):
    """입고 발주서 파일 업로드 (VF xlsx / 미입고 csv)"""
    error_id = str(uuid.uuid4())
    file_obj = request.FILES.get("file")
    if not file_obj:
        return Response(
            {"message": "file is required"}, status=status.HTTP_400_BAD_REQUEST
        )

    file_name = getattr(file_obj, "name", "") or "unknown"
    file_name_lower = file_name.lower()

    # 파일 타입 판별
    if "vf" in file_name_lower and file_name_lower.endswith(".xlsx"):
        file_type = "vf_xlsx"
    elif "미입고" in file_name_lower or "unreceived" in file_name_lower:
        file_type = "unreceived_csv"
    elif file_name_lower.endswith(".csv"):
        file_type = "unreceived_csv"  # 기본 CSV는 미입고로 간주
    elif file_name_lower.endswith(".xlsx"):
        file_type = "vf_xlsx"  # 기본 xlsx는 VF로 간주
    else:
        return Response(
            {"message": "지원하지 않는 파일 형식입니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    logger.info(
        "inbound_upload start error_id=%s file_type=%s file=%s",
        error_id,
        file_type,
        file_name,
    )

    try:
        raw = file_obj.read()
        if file_type == "vf_xlsx":
            df = pd.read_excel(io.BytesIO(raw), dtype=str, sheet_name="상품목록")
        else:
            # CSV 처리
            try:
                text = raw.decode("utf-8-sig")
            except Exception:
                text = raw.decode("cp949", errors="ignore")
            df = pd.read_csv(io.StringIO(text), dtype=str)
    except Exception as e:
        logger.exception("inbound_upload parse failed error_id=%s", error_id)
        return Response(
            {"message": f"파일 파싱 실패: {str(e)}", "errorId": error_id},
            status=status.HTTP_400_BAD_REQUEST,
        )

    df = df.fillna("")
    cols = _normalize_cols(df.columns)

    # 컬럼 인덱스 찾기
    if file_type == "vf_xlsx":
        # VF 발주서 업로드.xlsx 컬럼
        order_no_idx = _find_col_index(cols, ["발주번호", "order no", "orderno"])
        order_status_idx = _find_col_index(cols, ["발주상태", "order status"])
        barcode_idx = _find_col_index(cols, ["상품바코드", "바코드", "barcode"])
        product_name_idx = _find_col_index(cols, ["상품이름", "상품명", "product name"])
        ordered_qty_idx = _find_col_index(cols, ["발주수량", "ordered qty"])
        confirmed_qty_idx = _find_col_index(cols, ["확정수량", "confirmed qty"])
        expected_date_idx = _find_col_index(cols, ["입고예정일", "expected date"])
        product_no_idx = _find_col_index(
            cols, ["상품번호", "sku id", "sku_id", "product no"]
        )
    else:
        # 발주서 미입고 물량.csv 컬럼
        order_no_idx = _find_col_index(cols, ["발주번호", "order no", "orderno"])
        order_status_idx = _find_col_index(
            cols, ["발주현황", "발주상태", "order status"]
        )
        barcode_idx = _find_col_index(
            cols, ["sku barcode", "상품바코드", "바코드", "barcode"]
        )
        product_name_idx = _find_col_index(
            cols, ["sku 이름", "상품이름", "상품명", "product name"]
        )
        ordered_qty_idx = _find_col_index(cols, ["발주수량", "ordered qty"])
        confirmed_qty_idx = _find_col_index(cols, ["확정수량", "confirmed qty"])
        received_qty_idx = _find_col_index(cols, ["입고수량", "received qty"])
        expected_date_idx = _find_col_index(cols, ["입고예정일", "expected date"])
        product_no_idx = _find_col_index(
            cols, ["상품번호", "sku id", "sku_id", "product no"]
        )

    # 필수 컬럼 확인
    if order_no_idx is None or barcode_idx is None or confirmed_qty_idx is None:
        logger.warning(
            "inbound_upload missing required cols error_id=%s cols=%s",
            error_id,
            cols,
        )
        return Response(
            {
                "message": "필수 컬럼(발주번호/바코드/확정수량)을 찾을 수 없습니다.",
                "errorId": error_id,
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    # 업로드 레코드 생성
    upload = InboundOrderUpload.objects.create(
        file_name=file_name,
        file_type=file_type,
        rows_total=len(df),
        status="pending",
    )

    rows_parsed = 0
    rows_skipped = 0
    lines_to_create = []

    for _, row in df.iterrows():
        order_no = str(row.iloc[order_no_idx]).strip()
        barcode = str(row.iloc[barcode_idx]).strip()

        if not order_no or not barcode:
            rows_skipped += 1
            continue

        confirmed_qty = _parse_int(row.iloc[confirmed_qty_idx])
        if confirmed_qty <= 0:
            rows_skipped += 1
            continue

        order_status = (
            str(row.iloc[order_status_idx]).strip()
            if order_status_idx is not None
            else ""
        )
        product_name = (
            str(row.iloc[product_name_idx]).strip()
            if product_name_idx is not None
            else ""
        )
        ordered_qty = (
            _parse_int(row.iloc[ordered_qty_idx]) if ordered_qty_idx is not None else 0
        )
        received_qty = (
            _parse_int(row.iloc[received_qty_idx])
            if file_type == "unreceived_csv" and received_qty_idx is not None
            else 0
        )
        expected_date = (
            _parse_date_ymd(row.iloc[expected_date_idx])
            if expected_date_idx is not None
            else None
        )
        product_no = (
            str(row.iloc[product_no_idx]).strip() if product_no_idx is not None else ""
        )

        lines_to_create.append(
            InboundOrderLine(
                upload=upload,
                barcode=barcode,
                order_no=order_no,
                order_status=order_status,
                product_name=product_name,
                product_no=product_no,
                ordered_qty=ordered_qty,
                confirmed_qty=confirmed_qty,
                received_qty=received_qty,
                expected_date=expected_date,
            )
        )
        rows_parsed += 1

    # 일괄 생성
    if lines_to_create:
        InboundOrderLine.objects.bulk_create(lines_to_create, batch_size=2000)

    # 업로드 상태 업데이트
    upload.rows_parsed = rows_parsed
    upload.rows_skipped = rows_skipped
    upload.status = "success"
    upload.save(update_fields=["rows_parsed", "rows_skipped", "status"])

    logger.info(
        "inbound_upload done error_id=%s rows_parsed=%s rows_skipped=%s",
        error_id,
        rows_parsed,
        rows_skipped,
    )

    return Response(
        {
            "success": True,
            "message": "입고 발주서 파일이 업로드되었습니다.",
            "uploadId": str(upload.id),
            "fileType": file_type,
            "rowsParsed": rows_parsed,
            "rowsSkipped": rows_skipped,
        }
    )


@api_view(["GET", "DELETE"])
def inbound_order_latest(request):
    """최신 입고 발주서 데이터 조회 / 최신 업로드 초기화(삭제)"""
    latest_upload = (
        InboundOrderUpload.objects.filter(status="success")
        .order_by("-uploaded_at")
        .first()
    )
    if request.method == "DELETE":
        with transaction.atomic():
            deleted_lines_count, _ = InboundOrderLine.objects.all().delete()
            deleted_uploads_count, _ = InboundOrderUpload.objects.all().delete()

        return Response(
            {
                "success": True,
                "deleted": True,
                "deletedUploadsCount": int(deleted_uploads_count or 0),
                "deletedLinesCount": int(deleted_lines_count or 0),
            }
        )

    if not latest_upload:
        return Response(
            {
                "success": True,
                "data": [],
                "uploadInfo": None,
            }
        )

    # 정책 적용
    policy = InboundPolicy.objects.first()
    status_mode = (
        (getattr(policy, "status_mode", "") or "").strip().lower() if policy else ""
    )
    statuses = (getattr(policy, "statuses", None) or []) if policy else []
    statuses_norm = [str(s).strip() for s in statuses if str(s).strip()]

    lines_qs = InboundOrderLine.objects.filter(upload=latest_upload)

    if statuses_norm:
        q = models.Q()
        for s in statuses_norm:
            q |= models.Q(order_status__iexact=s)

        if status_mode == "exclude":
            lines_qs = lines_qs.exclude(q)
        elif status_mode == "include":
            lines_qs = lines_qs.filter(q)

    lines = []
    for line in lines_qs:
        lines.append(
            {
                "id": str(line.id),
                "barcode": line.barcode,
                "orderNo": line.order_no,
                "orderStatus": line.order_status,
                "productName": line.product_name,
                "productNo": line.product_no,
                "orderedQty": line.ordered_qty,
                "confirmedQty": line.confirmed_qty,
                "receivedQty": line.received_qty,
                "expectedDate": line.expected_date.isoformat()
                if line.expected_date
                else None,
            }
        )

    return Response(
        {
            "success": True,
            "data": lines,
            "uploadInfo": {
                "id": str(latest_upload.id),
                "fileName": latest_upload.file_name,
                "fileType": latest_upload.file_type,
                "uploadedAt": latest_upload.uploaded_at.isoformat(),
                "rowsTotal": latest_upload.rows_total,
                "rowsParsed": latest_upload.rows_parsed,
                "rowsSkipped": latest_upload.rows_skipped,
            },
        }
    )


@api_view(["GET", "POST"])
def inbound_policy(request):
    """입고 발주서 필터링 정책 조회/설정"""
    if request.method == "GET":
        policy = InboundPolicy.objects.first()
        if not policy:
            return Response(
                {
                    "statusMode": "exclude",
                    "statuses": [],
                }
            )
        return Response(
            {
                "statusMode": policy.status_mode,
                "statuses": policy.statuses or [],
            }
        )

    payload = request.data if isinstance(request.data, dict) else {}
    status_mode = (
        payload.get("statusMode") or payload.get("status_mode") or "exclude"
    ).strip()
    statuses = payload.get("statuses") or []

    if status_mode not in ("exclude", "include"):
        return Response(
            {"message": 'statusMode must be "exclude" or "include"'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not isinstance(statuses, list):
        return Response(
            {"message": "statuses must be an array"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    policy = InboundPolicy.objects.first()
    if policy:
        policy.status_mode = status_mode
        policy.statuses = statuses
        policy.save(update_fields=["status_mode", "statuses", "updated_at"])
    else:
        policy = InboundPolicy.objects.create(
            status_mode=status_mode,
            statuses=statuses,
        )

    return Response(
        {
            "success": True,
            "statusMode": policy.status_mode,
            "statuses": policy.statuses or [],
        }
    )


@api_view(["GET"])
def get_fc_inbound_records(request):
    start = request.query_params.get("start")
    end = request.query_params.get("end")
    try:
        limit = int(request.query_params.get("limit") or 10000)
    except Exception:
        limit = 10000

    queryset = FCInboundRecord.objects.all()
    filters = {}
    if start:
        filters["inbound_date__gte"] = start
    if end:
        filters["inbound_date__lte"] = end

    if filters:
        queryset = queryset.filter(**filters)

    queryset = queryset.order_by("-inbound_date")[:limit]

    serializer = FCInboundRecordSerializer(queryset, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def get_fc_inbound_stats(request):
    start = request.query_params.get("start") or request.query_params.get("startDate")
    end = request.query_params.get("end") or request.query_params.get("endDate")
    group_by = request.query_params.get("groupBy", "day")
    category = request.query_params.get("category")
    search = request.query_params.get("search")
    product = request.query_params.get("product")
    logistics_center = request.query_params.get(
        "logisticsCenter"
    ) or request.query_params.get("logistics_center")

    queryset = FCInboundRecord.objects.all()

    if start:
        queryset = queryset.filter(inbound_date__gte=start)
    if end:
        queryset = queryset.filter(inbound_date__lte=end)

    if category and category != "all":
        if category == "__others__":
            top_cats = list(
                queryset.values("category")
                .annotate(totalQty=Sum("quantity"))
                .order_by("-totalQty")
                .values_list("category", flat=True)[:10]
            )
            if top_cats:
                queryset = queryset.exclude(category__in=top_cats)
        else:
            queryset = queryset.filter(category=category)

    if logistics_center:
        # 쉼표로 구분된 여러 물류 센터 처리
        centers = [c.strip() for c in logistics_center.split(",") if c.strip()]
        if centers:
            queryset = queryset.filter(logistics_center__in=centers)
    else:
        # 파라미터 없을 때 VF67 제외 (기본 동작)
        queryset = queryset.exclude(logistics_center="VF67")

    if search:
        queryset = queryset.filter(product_name__icontains=search)

    if product:
        queryset = queryset.filter(product_name=product)

    # Separate aggregations to avoid mixed type issues
    count_result = queryset.aggregate(totalCount=Count("id"))
    quantity_result = queryset.aggregate(totalQuantity=Coalesce(Sum("quantity"), 0))
    supply_result = queryset.aggregate(
        totalSupplyAmount=Coalesce(
            Sum("supply_amount"), Value(Decimal("0"), output_field=DecimalField())
        )
    )

    summary = {
        "totalCount": count_result["totalCount"],
        "totalQuantity": quantity_result["totalQuantity"],
        "totalSupplyAmount": supply_result["totalSupplyAmount"],
    }

    trunc_func = TruncDay
    if group_by == "week":
        trunc_func = TruncWeek
    elif group_by == "month":
        trunc_func = TruncMonth

    # Get dates first
    dates = (
        queryset.annotate(date=trunc_func("inbound_date"))
        .values("date")
        .annotate(count=Count("id"))
        .order_by("date")
    )

    trend_data = []
    for item in dates:
        if item["date"]:
            # Format date based on group_by
            if group_by == "month":
                date_str = item["date"].strftime("%Y-%m")
            elif group_by == "week":
                date_str = item["date"].strftime("%Y-%m-%d")
            else:  # day
                date_str = item["date"].strftime("%Y-%m-%d")

            # Get quantity and supply amount for this date period
            if group_by == "month":
                # 월별: 해당 월의 1일부터 마지막일까지 범위 필터링
                from datetime import timedelta

                month_start = item["date"]
                # 다음 달 1일에서 하루를 빼면 현재 달의 마지막일
                from datetime import date

                if month_start.month == 12:
                    month_end = date(month_start.year + 1, 1, 1) - timedelta(days=1)
                else:
                    month_end = date(
                        month_start.year, month_start.month + 1, 1
                    ) - timedelta(days=1)
                day_qs = queryset.filter(inbound_date__range=[month_start, month_end])
            elif group_by == "week":
                # 주별: 해당 주의 월요일부터 일요일까지 범위 필터링
                from datetime import timedelta

                week_start = item["date"]
                week_end = week_start + timedelta(days=6)
                day_qs = queryset.filter(inbound_date__range=[week_start, week_end])
            else:
                # 일별: 해당 날짜의 데이터
                day_qs = queryset.filter(inbound_date=date_str)

            qty = day_qs.aggregate(total=Coalesce(Sum("quantity"), 0))["total"]
            supply = day_qs.aggregate(
                total=Coalesce(
                    Sum("supply_amount"),
                    Value(Decimal("0"), output_field=DecimalField()),
                )
            )["total"]
            trend_data.append(
                {
                    "date": date_str,
                    "quantity": qty or 0,
                    "supplyAmount": float(supply or 0),
                }
            )

    category_breakdown = (
        queryset.values("category")
        .annotate(quantity=Coalesce(Sum("quantity"), 0))
        .order_by("-quantity")
    )

    # Add supply_amount separately to avoid mixed type error
    for item in category_breakdown:
        cat = item["category"]
        cat_supply = queryset.filter(category=cat).aggregate(
            supplyAmount=Coalesce(
                Sum("supply_amount"), Value(Decimal("0"), output_field=DecimalField())
            )
        )
        item["supplyAmount"] = float(cat_supply["supplyAmount"] or 0)

    return Response(
        {
            "summary": {
                "totalCount": summary["totalCount"] or 0,
                "totalQuantity": summary["totalQuantity"] or 0,
                "totalSupplyAmount": float(summary["totalSupplyAmount"] or 0),
            },
            "dailyTrend": trend_data,
            "categoryBreakdown": category_breakdown,
        }
    )


@api_view(["GET"])
def get_fc_inbound_top_products(request):
    start = request.query_params.get("start") or request.query_params.get("startDate")
    end = request.query_params.get("end") or request.query_params.get("endDate")
    category = request.query_params.get("category")
    search = request.query_params.get("search")
    product = request.query_params.get("product")
    logistics_center = request.query_params.get(
        "logisticsCenter"
    ) or request.query_params.get("logistics_center")
    try:
        limit = int(request.query_params.get("limit") or 100)
    except Exception:
        limit = 100
    limit = max(1, min(limit, 500))

    queryset = FCInboundRecord.objects.all()
    filters = {}
    if start:
        filters["inbound_date__gte"] = start
    if end:
        filters["inbound_date__lte"] = end

    if filters:
        queryset = queryset.filter(**filters)

    if category and category != "all":
        if category == "__others__":
            top_cats = list(
                queryset.values("category")
                .annotate(totalQty=Sum("quantity"))
                .order_by("-totalQty")
                .values_list("category", flat=True)[:10]
            )
            if top_cats:
                queryset = queryset.exclude(category__in=top_cats)
        else:
            queryset = queryset.filter(category=category)

    if logistics_center:
        # 쉼표로 구분된 여러 물류 센터 처리
        centers = [c.strip() for c in logistics_center.split(",") if c.strip()]
        if centers:
            queryset = queryset.filter(logistics_center__in=centers)
    else:
        # 파라미터 없을 때 VF67 제외 (기본 동작)
        queryset = queryset.exclude(logistics_center="VF67")

    if search:
        queryset = queryset.filter(product_name__icontains=search)

    if product:
        queryset = queryset.filter(product_name=product)

    # Get top products by quantity
    rows = (
        queryset.values("product_name")
        .annotate(
            quantity=Coalesce(Sum("quantity"), 0),
        )
        .order_by("-quantity")[:limit]
    )

    # Add supply amount for each product
    result = []
    for r in rows:
        product_name = r.get("product_name") or "-"
        qty = r.get("quantity") or 0
        # Get supply amount for this product
        supply = queryset.filter(product_name=product_name).aggregate(
            total=Coalesce(
                Sum("supply_amount"), Value(Decimal("0"), output_field=DecimalField())
            )
        )["total"]
        result.append(
            {
                "name": product_name,
                "quantity": qty,
                "salesAmount": float(supply or 0),
                "supplyAmount": float(supply or 0),
            }
        )

    return Response(result)


@api_view(["GET"])
def get_fc_inbound_pivot(request):
    start = request.query_params.get("start") or request.query_params.get("startDate")
    end = request.query_params.get("end") or request.query_params.get("endDate")
    row = request.query_params.get("row", "category")
    group_by = request.query_params.get("groupBy", "day")
    category = request.query_params.get("category")
    search = request.query_params.get("search")
    product = request.query_params.get("product")
    logistics_center = request.query_params.get(
        "logisticsCenter"
    ) or request.query_params.get("logistics_center")
    try:
        limit = int(request.query_params.get("limit") or 100)
    except Exception:
        limit = 100
    limit = max(1, min(limit, 500))

    if row not in ["category", "product"]:
        return Response(
            {"message": "row must be category or product"},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if group_by not in ["day", "week", "month"]:
        return Response(
            {"message": "groupBy must be day, week, or month"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    queryset = FCInboundRecord.objects.all()
    filters = {}
    if start:
        filters["inbound_date__gte"] = start
    if end:
        filters["inbound_date__lte"] = end

    if filters:
        queryset = queryset.filter(**filters)

    if category and category != "all":
        if category == "__others__":
            top_cats = list(
                queryset.values("category")
                .annotate(totalQty=Sum("quantity"))
                .order_by("-totalQty")
                .values_list("category", flat=True)[:10]
            )
            if top_cats:
                queryset = queryset.exclude(category__in=top_cats)
        else:
            queryset = queryset.filter(category=category)

    if logistics_center:
        # 쉼표로 구분된 여러 물류 센터 처리
        centers = [c.strip() for c in logistics_center.split(",") if c.strip()]
        if centers:
            queryset = queryset.filter(logistics_center__in=centers)
    else:
        # 파라미터 없을 때 VF67 제외 (기본 동작)
        queryset = queryset.exclude(logistics_center="VF67")

    if search:
        queryset = queryset.filter(product_name__icontains=search)

    if product:
        queryset = queryset.filter(product_name=product)

    trunc_func = TruncDay
    if group_by == "week":
        trunc_func = TruncWeek
    elif group_by == "month":
        trunc_func = TruncMonth

    if row == "category":
        rows = queryset.values("category")
    else:
        rows = queryset.values("product_name")

    row_field = "category" if row == "category" else "product_name"

    rows_data = []
    for r in rows.annotate(total=Coalesce(Sum("quantity"), 0)).order_by("-total")[
        :limit
    ]:
        row_key = r.get(row_field) or "-"

        # Get total supply amount for this row
        row_queryset = queryset.filter(**{row_field: row_key})
        total_supply = row_queryset.aggregate(
            total=Coalesce(
                Sum("supply_amount"), Value(Decimal("0"), output_field=DecimalField())
            )
        )["total"]

        row_data = {
            "key": row_key,
            "values": {},
            "total": {
                "quantity": r["total"] or 0,
                "salesAmount": float(total_supply or 0),
            },
        }

        # Get daily breakdown with supply amount
        if group_by == "month":
            # For monthly grouping, get the actual date range
            daily_data = (
                row_queryset.annotate(date=trunc_func("inbound_date"))
                .values("date")
                .annotate(quantity=Coalesce(Sum("quantity"), 0))
                .order_by("date")
            )

            for d in daily_data:
                if d["date"]:
                    # Use YYYY-MM format for monthly grouping
                    date_key = d["date"].strftime("%Y-%m")
                    # For month grouping, get the date range
                    year = d["date"].year
                    month = d["date"].month
                    from datetime import datetime

                    start_date = datetime(year, month, 1).date()
                    if month == 12:
                        end_date = datetime(year + 1, 1, 1).date()
                    else:
                        end_date = datetime(year, month + 1, 1).date()
                    day_supply = row_queryset.filter(
                        inbound_date__gte=start_date, inbound_date__lt=end_date
                    ).aggregate(
                        total=Coalesce(
                            Sum("supply_amount"),
                            Value(Decimal("0"), output_field=DecimalField()),
                        )
                    )["total"]

                    row_data["values"][date_key] = {
                        "quantity": d["quantity"] or 0,
                        "salesAmount": float(day_supply or 0),
                    }
        else:
            # For day/week grouping
            daily_data = (
                row_queryset.annotate(date=trunc_func("inbound_date"))
                .values("date")
                .annotate(quantity=Coalesce(Sum("quantity"), 0))
                .order_by("date")
            )

            for d in daily_data:
                if d["date"]:
                    date_key = d["date"].strftime("%Y-%m-%d")
                    day_supply = row_queryset.filter(inbound_date=d["date"]).aggregate(
                        total=Coalesce(
                            Sum("supply_amount"),
                            Value(Decimal("0"), output_field=DecimalField()),
                        )
                    )["total"]
                    row_data["values"][date_key] = {
                        "quantity": d["quantity"] or 0,
                        "salesAmount": float(day_supply or 0),
                    }

        rows_data.append(row_data)

    return Response(rows_data)


@api_view(["POST"])
def fc_inbound_upload(request):
    """
    FC 입고 엑셀 업로드 (Coupang_Stocked_Data_List 양식).
    - 입고 실적 저장
    - 단가 컬럼 → 제품 마스터(MasterSpec.price) 자동 동기화 (단가 단일 기준)
    """
    if "file" not in request.FILES:
        return Response(
            {"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST
        )

    file = request.FILES["file"]
    if not file.name.endswith((".xlsx", ".xls")):
        return Response(
            {"error": "Only Excel files are supported"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        import io
        import hashlib
        from django.db.models import Q

        # 파일 해시 계산 (중복 체크용)
        file_content = file.read()
        file_hash = hashlib.sha256(file_content).hexdigest()

        # DataFrame 로드 (해시 중복이어도 단가 재동기화 가능하도록 먼저 파싱)
        df = pd.read_excel(io.BytesIO(file_content))
        df.columns = [str(c).strip() for c in df.columns]

        required_columns = ["SKU번호", "SKU명", "입고/반출시각", "물류센터", "수량"]
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            return Response(
                {"error": f"Missing required columns: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 단가 단일 기준: FC 파일 '단가' → MasterSpec 반영 (입고 중복과 무관하게 항상 실행)
        price_sync = _sync_master_prices_from_fc_stocked_df(df)

        already_uploaded = FCInboundFileUpload.objects.filter(file_hash=file_hash).exists()
        existing_upload = (
            FCInboundFileUpload.objects.filter(file_hash=file_hash).first()
            if already_uploaded
            else None
        )

        records_created = 0
        records_skipped = 0
        records_duplicate = 0
        records_processed = 0
        processed_skus = set()
        processed_barcodes = set()

        if not already_uploaded:
            for _, row in df.iterrows():
                records_processed += 1
                try:
                    date_str = str(row.get("입고/반출시각", ""))
                    if not date_str or date_str == "nan":
                        records_skipped += 1
                        continue

                    try:
                        date_obj = pd.to_datetime(date_str, errors="coerce")
                        if pd.isna(date_obj):
                            records_skipped += 1
                            continue
                        inbound_date = date_obj.date()
                    except Exception:
                        records_skipped += 1
                        continue

                    sku_id = str(row.get("SKU번호", "") or "").strip()
                    product_name = str(row.get("SKU명", "") or "").strip()

                    try:
                        quantity = int(
                            float(str(row.get("수량", 0)).replace(",", ""))
                        )
                    except Exception:
                        quantity = 0

                    try:
                        supply_amount = float(
                            str(row.get("공급가액", 0)).replace(",", "")
                        )
                    except Exception:
                        supply_amount = 0

                    logistics_center = str(row.get("물류센터", "") or "").strip()

                    if not sku_id or not product_name or quantity <= 0:
                        records_skipped += 1
                        continue

                    # 중복 체크
                    existing_record = FCInboundRecord.objects.filter(
                        inbound_date=inbound_date,
                        sku_id=sku_id,
                        product_name=product_name,
                        logistics_center=logistics_center,
                        quantity=quantity,
                    ).first()

                    if existing_record:
                        records_duplicate += 1
                        continue

                    # MasterSpec에서 바코드·카테고리 조회 (SKU번호 ≠ 바코드)
                    category = ""
                    barcode = ""
                    spec = MasterSpec.objects.filter(sku_id=sku_id).first()
                    if spec:
                        barcode = (spec.barcode or "").strip()
                        if spec.category_lg:
                            category = spec.category_lg

                    FCInboundRecord.objects.create(
                        inbound_date=inbound_date,
                        sku_id=sku_id,
                        barcode=barcode,
                        product_name=product_name,
                        category=category,
                        subcategory="",
                        color="",
                        quantity=quantity,
                        supply_amount=supply_amount,
                        logistics_center=logistics_center,
                    )
                    records_created += 1

                    if sku_id:
                        processed_skus.add(sku_id)
                    if barcode:
                        processed_barcodes.add(barcode)

                except Exception as e:
                    logger.error(f"Error processing row: {e}")
                    records_skipped += 1
                    continue

            # 신규 입고 품목 자동 복구: 3개월 미출고만 해제 (단종은 수동 유지)
            if processed_skus or processed_barcodes:
                _clear_no_outbound_3m_on_activity(
                    barcodes=processed_barcodes, sku_ids=processed_skus
                )

            # 파일 업로드 이력 저장
            file_upload = FCInboundFileUpload.objects.create(
                file_name=file.name,
                file_hash=file_hash,
                records_processed=records_processed,
                records_created=records_created,
                records_skipped=records_skipped,
                records_duplicate=records_duplicate,
                status="completed" if records_created > 0 else "partial",
            )
            upload_id = str(file_upload.id)
            file_name = file_upload.file_name
        else:
            # 동일 파일 재업로드: 입고 실적은 건너뛰고 단가만 재반영
            upload_id = str(existing_upload.id) if existing_upload else None
            file_name = existing_upload.file_name if existing_upload else file.name
            records_processed = len(df)
            # 단가 동기화에 쓰인 SKU도 미출고 해제
            if price_sync.get("success"):
                skus_in_file = set()
                for _, row in df.iterrows():
                    s = str(row.get("SKU번호", "") or "").strip()
                    if s and s.lower() != "nan":
                        skus_in_file.add(s)
                if skus_in_file:
                    _clear_no_outbound_3m_on_activity(sku_ids=skus_in_file)

        return Response(
            {
                "success": True,
                "uploadId": upload_id,
                "fileName": file_name,
                "alreadyUploaded": already_uploaded,
                "recordsCreated": records_created,
                "recordsSkipped": records_skipped,
                "recordsDuplicate": records_duplicate,
                "totalRows": len(df),
                "priceSync": {
                    "success": bool(price_sync.get("success")),
                    "priceColumn": price_sync.get("price_column"),
                    "newCreated": price_sync.get("new_created", 0),
                    "updated": price_sync.get("updated", 0),
                    "unchanged": price_sync.get("unchanged", 0),
                    "priceChanged": price_sync.get("price_changed", 0),
                    "totalItems": price_sync.get("total_items", 0),
                    "message": price_sync.get("message"),
                },
            }
        )

    except Exception as e:
        logger.error(f"FC Inbound upload error: {e}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
def get_fc_inbound_uploads(request):
    """FC 입고 파일 업로드 이력 조회"""
    try:
        limit = int(request.query_params.get("limit") or 50)
    except Exception:
        limit = 50

    uploads = FCInboundFileUpload.objects.all().order_by("-upload_date")[:limit]

    serializer = FCInboundFileUploadSerializer(uploads, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def delete_fc_inbound_upload(request, upload_id):
    """FC 입고 파일 업로드 이력 및 관련 레코드 삭제"""
    try:
        upload = FCInboundFileUpload.objects.get(id=upload_id)
        file_name = upload.file_name

        # 업로드 이력 삭제
        upload.delete()

        return Response(
            {
                "success": True,
                "message": f'File upload record "{file_name}" has been deleted',
            }
        )
    except FCInboundFileUpload.DoesNotExist:
        return Response(
            {"error": "Upload record not found"}, status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error deleting upload: {e}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def sync_master_specs_from_sheet(request):
    """구글 시트에서 마스터 데이터 동기화 (FC 카테고리 매핑)"""
    import requests

    sheet_url = os.environ.get("MASTER_DATA_CSV_URL")
    if not sheet_url:
        return Response(
            {"error": "MASTER_DATA_CSV_URL 환경변수가 설정되지 않았습니다."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    try:
        # CSV 다운로드
        response = requests.get(sheet_url, timeout=30)
        response.raise_for_status()

        # CSV 파싱 (BOM 제거를 위해 utf-8-sig 사용)
        try:
            decoded_content = response.content.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded_content = response.content.decode("cp949")

        df = pd.read_csv(io.StringIO(decoded_content), dtype=str)

        # 헤더 정규화
        df.columns = [str(c).strip().lstrip("\ufeff") for c in df.columns]

        # 필수 컬럼 확인
        required_cols = ["SKU ID", "바코드", "대분류"]
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            return Response(
                {"error": f"구글 시트에 필수 컬럼이 없습니다: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        added = 0
        updated = 0
        errors = 0

        for _, row in df.iterrows():
            try:
                sku_id = (
                    str(int(row.get("SKU ID", 0)))
                    if pd.notna(row.get("SKU ID"))
                    else ""
                )
                barcode = (
                    str(row.get("바코드", "")).strip()
                    if pd.notna(row.get("바코드"))
                    else ""
                )
                category_lg = str(row.get("대분류", "")).strip()
                category_md = str(row.get("중분류", "")).strip()
                product_name = str(row.get("상품명", "")).strip()

                if not sku_id and not barcode:
                    errors += 1
                    continue

                # sku_id 또는 barcode로 기존 레코드 찾기
                if sku_id:
                    spec = MasterSpec.objects.filter(sku_id=sku_id).first()
                elif barcode:
                    spec = MasterSpec.objects.filter(barcode=barcode).first()
                else:
                    spec = None

                if spec:
                    # 업데이트
                    changed = False
                    if category_lg and spec.category_lg != category_lg:
                        spec.category_lg = category_lg
                        changed = True
                    if category_md and spec.category_md != category_md:
                        spec.category_md = category_md
                        changed = True
                    if sku_id and spec.sku_id != sku_id:
                        spec.sku_id = sku_id
                        changed = True
                    if barcode and spec.barcode != barcode:
                        spec.barcode = barcode
                        changed = True
                    if changed:
                        spec.save()
                        updated += 1
                else:
                    # 새로 추가 (product_name이 없으면 sku_id 사용)
                    MasterSpec.objects.create(
                        product_name=product_name or f"SKU_{sku_id}",
                        sku_id=sku_id,
                        barcode=barcode,
                        category_lg=category_lg,
                        category_md=category_md,
                    )
                    added += 1

            except Exception as e:
                logger.error(f"Error processing row: {e}")
                errors += 1
                continue

        return Response(
            {
                "success": True,
                "added": added,
                "updated": updated,
                "errors": errors,
                "total": len(df),
            }
        )

    except Exception as e:
        logger.error(f"Master spec sync error: {e}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==================== FC 입고 구글 시트 연동 ====================


def fetch_category_mapping():
    """
    마스터 데이터에서 SKU별 대분류 매핑 테이블 생성
    Returns: dict {sku_id: category}
    """
    import requests
    import csv
    from io import StringIO

    try:
        response = requests.get(MASTER_DATA_CSV_URL, timeout=30)
        response.raise_for_status()

        # UTF-8 BOM 처리 및 인코딩
        csv_text = response.content.decode("utf-8-sig")
        csv_reader = csv.DictReader(StringIO(csv_text))

        # CSV 헤더의 실제 키 확인 (인코딩 문제 방지)
        headers = None
        category_map = {}

        for row in csv_reader:
            if headers is None:
                headers = list(row.keys())
                logger.info(f"Master data CSV headers: {headers}")

            # 여러 가능한 키 이름 시도
            sku_id = (
                row.get("SKU ID")
                or row.get("SKU_ID")
                or row.get("sku_id")
                or row.get("SKU번호")
                or row.get("SKU 번호")
                or ""
            )

            category = (
                row.get("대분류")
                or row.get("분류")
                or row.get("category")
                or row.get("Category")
                or ""
            )

            if sku_id and category:
                sku_id = str(sku_id).strip()
                category = str(category).strip()
                category_map[sku_id] = category

        logger.info(
            f"Loaded {len(category_map)} SKU-category mappings from master data"
        )
        return category_map

    except Exception as e:
        logger.error(f"Failed to fetch master data: {e}")
        return {}


@api_view(["POST"])
def sync_fc_inbound_from_sheet(request):
    """
    구글 시트 CSV에서 FC 입고 데이터를 가져와서 DB에 저장/업데이트
    중복 체크: sku_id + inbound_date + logistics_center 조합
    덮어쓰기 방식
    마스터 데이터에서 대분류를 매핑
    최적화: 벌크 연산 사용 (bulk_create, bulk_update)
    """
    import requests
    import csv
    from io import StringIO
    from datetime import datetime
    import decimal

    try:
        # 마스터 데이터에서 대분류 매핑 로드
        category_mapping = fetch_category_mapping()

        # 구글 시트 CSV 가져오기
        csv_url = os.environ.get("FC_GOOGLE_SHEET_CSV_URL")
        if not csv_url:
            return Response(
                {"error": "FC_GOOGLE_SHEET_CSV_URL 환경변수가 설정되지 않았습니다."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        response = requests.get(csv_url, timeout=30)
        response.raise_for_status()

        # CSV 텍스트 디코딩 (BOM 제거 처리)
        try:
            csv_text = response.content.decode("utf-8-sig")
        except UnicodeDecodeError:
            # 윈도우 엑셀 저장 CSV일 경우 cp949 시도
            csv_text = response.content.decode("cp949")

        # DictReader의 fieldnames에 공백/BOM이 들어가는 문제 해결을 위해 iterator 사용
        # 혹은 첫 줄(헤더)를 미리 정규화
        f = StringIO(csv_text)
        reader = csv.reader(f)
        headers = next(reader, None)

        if not headers:
            return Response(
                {"error": "CSV 파일이 비어있습니다."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 헤더 정규화 (BOM 제거, 공백 제거)
        normalized_headers = [h.strip().lstrip("\ufeff") for h in headers]

        csv_reader = csv.DictReader(f, fieldnames=normalized_headers)

        # 기존 데이터 전체 조회 (단일 쿼리) - 메모리에 맵핑
        all_existing = list(
            FCInboundRecord.objects.all().values(
                "id",
                "sku_id",
                "inbound_date",
                "logistics_center",
                "product_name",
                "quantity",
                "supply_amount",
                "category",
            )
        )

        # (sku_id, inbound_date, logistics_center) → record 맵핑
        existing_map = {
            (r["sku_id"], str(r["inbound_date"]), r["logistics_center"]): r
            for r in all_existing
        }

        # 파싱된 데이터 수집
        to_create = []
        to_update = []
        skipped = 0
        errors = 0

        for row in csv_reader:
            try:
                # CSV 데이터 파싱
                sku_id = row.get("SKU번호", "").strip()
                product_name = row.get("SKU명", "").strip()
                inbound_datetime = row.get("입고/반출시각", "").strip()
                logistics_center = row.get("물류센터", "").strip()
                quantity_str = row.get("수량", "0").replace(",", "").strip()
                supply_amount_str = row.get("총공급가액", "0").replace(",", "").strip()

                # 필수 필드 확인
                if not sku_id or not inbound_datetime:
                    skipped += 1
                    continue

                # 날짜 파싱 (YYYY/MM/DD HH:MM:SS → YYYY-MM-DD)
                try:
                    dt = datetime.strptime(inbound_datetime.split()[0], "%Y/%m/%d")
                    inbound_date = dt.date()
                except ValueError:
                    try:
                        dt = datetime.strptime(inbound_datetime, "%Y-%m-%d")
                        inbound_date = dt.date()
                    except ValueError:
                        skipped += 1
                        continue

                # 수량 파싱
                try:
                    quantity = int(quantity_str) if quantity_str else 0
                except ValueError:
                    quantity = 0

                # 공급가액 파싱
                try:
                    supply_amount = float(supply_amount_str) if supply_amount_str else 0
                    supply_amount = (
                        str(int(supply_amount))
                        if supply_amount == int(supply_amount)
                        else str(supply_amount)
                    )
                    supply_amount = Decimal(supply_amount)
                except (ValueError, decimal.InvalidOperation):
                    supply_amount = Decimal("0")

                # 대분류 매핑 (마스터 데이터에서 가져옴)
                category = category_mapping.get(sku_id, "")

                # 중복 체크: sku_id + inbound_date + logistics_center
                key = (sku_id, str(inbound_date), logistics_center)

                if key in existing_map:
                    # 기존 레코드 - 변경사항 확인 후 업데이트 목록에 추가
                    existing_record = existing_map[key]
                    needs_update = (
                        existing_record["product_name"] != product_name
                        or existing_record["quantity"] != quantity
                        or existing_record["supply_amount"] != str(supply_amount)
                        or existing_record["category"] != category
                    )

                    if needs_update:
                        # DB 객체 조회 (나중에 bulk_update용)
                        obj = FCInboundRecord.objects.get(id=existing_record["id"])
                        obj.product_name = product_name
                        obj.quantity = quantity
                        obj.supply_amount = supply_amount
                        obj.category = category
                        to_update.append(obj)
                else:
                    # 새 레코드 - 생성 목록에 추가
                    to_create.append(
                        FCInboundRecord(
                            sku_id=sku_id,
                            barcode=sku_id,
                            product_name=product_name,
                            inbound_date=inbound_date,
                            logistics_center=logistics_center,
                            quantity=quantity,
                            supply_amount=supply_amount,
                            category=category,
                        )
                    )

            except Exception as e:
                logger.error(f"Error processing FC inbound row: {e}, row: {row}")
                errors += 1
                continue

        # 벌크 연산 실행
        created = 0
        updated = 0

        with transaction.atomic():
            # 벌크 생성 (batch_size=500)
            if to_create:
                FCInboundRecord.objects.bulk_create(to_create, batch_size=500)
                created = len(to_create)

            # 벌크 업데이트 (batch_size=500)
            if to_update:
                FCInboundRecord.objects.bulk_update(
                    to_update,
                    ["product_name", "quantity", "supply_amount", "category"],
                    batch_size=500,
                )
                updated = len(to_update)

            # 신규 입고 품목 자동 복구: 3개월 미출고만 해제 (단종은 수동 유지)
            inbound_skus = {item.sku_id for item in to_create if item.sku_id} | {item.sku_id for item in to_update if item.sku_id}
            inbound_barcodes = {item.barcode for item in to_create if item.barcode} | {item.barcode for item in to_update if item.barcode}
            if inbound_skus or inbound_barcodes:
                _clear_no_outbound_3m_on_activity(
                    barcodes=inbound_barcodes, sku_ids=inbound_skus
                )

        return Response(
            {
                "success": True,
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
                "total": created + updated + skipped + errors,
            }
        )

    except requests.RequestException as e:
        logger.error(f"Failed to fetch Google Sheet: {e}")
        return Response(
            {"error": f"구글 시트 가져오기 실패: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
    except Exception as e:
        logger.error(f"FC inbound sync error: {e}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["DELETE"])
def delete_fc_inbound_uploaded_data(request):
    """
    업로드된 FC 입고 데이터를 모두 삭제.
    DESTRUCTIVE_API_KEY 설정 시 키 필수.
    """
    from .api_guards import destructive_guard_response

    blocked = destructive_guard_response(request)
    if blocked is not None:
        return blocked

    try:
        logger.warning("delete_fc_inbound_uploaded_data FULL WIPE")
        deleted_count = FCInboundRecord.objects.all().delete()[0]

        # 업로드 이력도 삭제
        FCInboundFileUpload.objects.all().delete()

        return Response({"success": True, "deleted": deleted_count})

    except Exception as e:
        logger.error(f"Failed to delete FC inbound data: {e}")
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _get_project_context():
    """
    프로젝트의 전반적인 데이터 상태를 요약하여 문자열로 반환
    (출고, 재고, 입고, 생산, 특이사항)
    """
    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Sum, Count, Q

    now = timezone.now()
    today = now.date()
    yesterday = today - timedelta(days=1)

    lines = []
    lines.append(f"=== System Info ===")
    lines.append(f"Current Time: {now.strftime('%Y-%m-%d %H:%M:%S')}")

    # 1. 출고 (Sales/Outbound)
    sales_today = OutboundRecord.objects.filter(outbound_date=today).aggregate(
        count=Count("id"), qty=Sum("quantity")
    )
    sales_yesterday = OutboundRecord.objects.filter(outbound_date=yesterday).aggregate(
        count=Count("id"), qty=Sum("quantity")
    )

    lines.append(f"\n=== Sales (Outbound) ===")
    lines.append(
        f"Today ({today}): {sales_today['qty'] or 0} items ({sales_today['count']} records)"
    )
    lines.append(
        f"Yesterday ({yesterday}): {sales_yesterday['qty'] or 0} items ({sales_yesterday['count']} records)"
    )

    # Top 5 products usually (Recent 3 days)
    recent_start = today - timedelta(days=3)
    top_products = list(
        OutboundRecord.objects.filter(outbound_date__gte=recent_start)
        .values("product_name")
        .annotate(total_qty=Sum("quantity"))
        .order_by("-total_qty")[:5]
    )

    if top_products:
        top_str = ", ".join(
            [f"{p['product_name']}({p['total_qty']})" for p in top_products]
        )
        lines.append(f"Top 5 Products (3 days): {top_str}")

    # 2. 재고 (Inventory)
    # Low stock items (current < minimum)
    low_stock_items = InventoryItem.objects.filter(
        current_stock__lt=models.F("minimum_stock")
    )
    low_stock_count = low_stock_items.count()

    lines.append(f"\n=== Inventory ===")
    lines.append(f"Total Items: {InventoryItem.objects.count()}")
    if low_stock_count > 0:
        lines.append(f"WARNING: {low_stock_count} items are below minimum stock.")
        # List up to 5 critical items
        critical_list = [
            f"{item.name} (Curr:{item.current_stock}/Min:{item.minimum_stock})"
            for item in low_stock_items[:5]
        ]
        lines.append(f"Critical Items: {', '.join(critical_list)}")
    else:
        lines.append("All items are above minimum stock levels.")

    # 3. 생산 (Production) - ProductionLog (실제 생산 실적)
    prod_today = ProductionLog.objects.filter(
        date=today, status__in=["running", "completed"]
    )
    prod_lines = prod_today.values("machine_number", "product_name", "total")

    lines.append(f"\n=== Production (Today) ===")
    if prod_lines:
        for p in prod_lines:
            lines.append(
                f"Machine {p['machine_number']}: {p['product_name']} ({p['total']} produced)"
            )
    else:
        lines.append("No active production logs for today.")

    # 3b. 생산 계획 (MachinePlan - AI 추천/적용 계획)
    # 현재 표시 중인 생산 계획 데이터 (모든 상태)
    machine_plans = MachinePlan.objects.exclude(status='cancelled').order_by('date', 'machine_number')
    lines.append(f"\n=== Production Plans (MachinePlan) ===")
    lines.append(f"Total active plans: {machine_plans.count()}")
    if machine_plans.exists():
        for i, plan in enumerate(machine_plans[:20], 1):  # 최대 20개만 표시
            lines.append(
                f"{i}. [{plan.date}] Machine {plan.machine_number}: {plan.product_name} "
                f"Qty:{plan.quantity} Status:{plan.status}"
            )
    else:
        lines.append("No production plans found.")

    # 4. 입고 (Inbound - FC)
    # Recent 3 days
    inbound_recent = FCInboundRecord.objects.filter(
        inbound_date__gte=recent_start
    ).aggregate(total_qty=Sum("quantity"))
    lines.append(f"\n=== Inbound (FC) ===")
    lines.append(f"Total Inbound (Last 3 days): {inbound_recent['total_qty'] or 0}")

    # 5. 특이사항/이슈 (Issues)
    # Recent 7 days of DeliverySpecialNote
    issue_start = today - timedelta(days=7)
    issues = DeliverySpecialNote.objects.filter(date__gte=issue_start).order_by(
        "-date"
    )[:5]

    lines.append(f"\n=== Special Notes / Issues (Last 7 days) ===")
    if issues.exists():
        for issue in issues:
            lines.append(f"[{issue.date}] {issue.product_name}: {issue.memo}")
    else:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
def ai_backtest_log(request):
    """
    백테스트 결과를 서버에 저장하는 API
    Request: { logs: [{ date, day_of_week, is_month_start, is_month_end, predicted_value, actual_value, error_rate, variant_id }...] }
    Response: { success: true, count: N }
    """
    payload = request.data if isinstance(request.data, dict) else {}
    logs = payload.get("logs", [])
    variant_id = payload.get("variantId", "unknown")

    if not logs:
        return Response(
            {"success": False, "message": "No logs provided"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Save to file (simpler than DB for now)
    import os
    from datetime import datetime as dt

    log_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "logs",
    )
    os.makedirs(log_dir, exist_ok=True)

    timestamp = dt.now().strftime("%Y%m%d_%H%M%S")
    filename = f"backtest_{variant_id}_{timestamp}.json"

    try:
        import json

        filepath = os.path.join(log_dir, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(
                {"variant_id": variant_id, "timestamp": timestamp, "logs": logs},
                f,
                ensure_ascii=False,
                indent=2,
            )

        # Calculate summary
        total_error = 0
        valid_count = 0
        over_count = 0
        under_count = 0

        for log in logs:
            pred = log.get("predicted_value", 0)
            actual = log.get("actual_value", 0)
            if actual > 0:
                total_error += abs(pred - actual) / actual
                valid_count += 1
                if pred > actual:
                    over_count += 1
                else:
                    under_count += 1

        avg_error = (total_error / valid_count * 100) if valid_count > 0 else 0

        return Response(
            {
                "success": True,
                "count": len(logs),
                "summary": {
                    "variant_id": variant_id,
                    "avg_error_percent": round(avg_error, 2),
                    "valid_days": valid_count,
                    "over_estimation": over_count,
                    "under_estimation": under_count,
                },
                "file": filename,
            }
        )

    except Exception as e:
        logger.error(f"Backtest log save error: {e}")
        return Response(
            {"success": False, "error": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
def ai_accuracy_stats(request):
    """
    백테스트 정확도 통계 API
    - 요일별, 기간별, 시간대별 정확도 분석
    Response: { stats: { total: {...}, day: {...}, period: {...} } }
    """
    import os
    import glob

    log_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "logs",
    )
    pattern = os.path.join(log_dir, "backtest_*.json")

    all_logs = []

    # 모든 백테스트 로그 파일 읽기
    for filepath in glob.glob(pattern):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
                all_logs.extend(data.get("logs", []))
        except Exception as e:
            logger.warning(f"Failed to read {filepath}: {e}")

    if not all_logs:
        return Response(
            {
                "stats": {
                    "total": {
                        "count": 0,
                        "avg_error": None,
                        "over_rate": None,
                        "under_rate": None,
                    },
                    "day": {},
                    "period": {},
                }
            }
        )

    # 요일별 통계
    day_stats = {
        i: {"count": 0, "total_error": 0, "over": 0, "under": 0} for i in range(7)
    }
    day_names = ["일", "월", "화", "수", "목", "금", "토"]

    # 기간별 통계 (month_start, month_mid, month_end)
    period_stats = {
        "month_start": {"count": 0, "total_error": 0, "over": 0, "under": 0},
        "month_mid": {"count": 0, "total_error": 0, "over": 0, "under": 0},
        "month_end": {"count": 0, "total_error": 0, "over": 0, "under": 0},
    }

    # 전체 통계
    total_count = 0
    total_error = 0
    total_over = 0
    total_under = 0

    for log in all_logs:
        pred = log.get("predicted_value", 0)
        actual = log.get("actual_value", 0)
        day_of_week = log.get("day_of_week", 0)
        is_month_start = log.get("is_month_start", 0)
        is_month_end = log.get("is_month_end", 0)

        if actual <= 0:
            continue

        error = abs(pred - actual) / actual
        is_over = 1 if pred > actual else 0

        # 전체
        total_count += 1
        total_error += error
        total_over += is_over
        total_under += 1 - is_over

        # 요일별
        if day_of_week in day_stats:
            day_stats[day_of_week]["count"] += 1
            day_stats[day_of_week]["total_error"] += error
            day_stats[day_of_week]["over"] += is_over
            day_stats[day_of_week]["under"] += 1 - is_over

        # 기간별
        if is_month_start:
            period_stats["month_start"]["count"] += 1
            period_stats["month_start"]["total_error"] += error
            period_stats["month_start"]["over"] += is_over
            period_stats["month_start"]["under"] += 1 - is_over
        elif is_month_end:
            period_stats["month_end"]["count"] += 1
            period_stats["month_end"]["total_error"] += error
            period_stats["month_end"]["over"] += is_over
            period_stats["month_end"]["under"] += 1 - is_over
        else:
            period_stats["month_mid"]["count"] += 1
            period_stats["month_mid"]["total_error"] += error
            period_stats["month_mid"]["over"] += is_over
            period_stats["month_mid"]["under"] += 1 - is_over

    # 결과 정리
    def calc_stats(s):
        if s["count"] == 0:
            return {
                "count": 0,
                "avg_error": None,
                "over_rate": None,
                "under_rate": None,
            }
        return {
            "count": s["count"],
            "avg_error": round(s["total_error"] / s["count"], 4),
            "over_rate": round(s["over"] / s["count"] * 100, 1),
            "under_rate": round(s["under"] / s["count"] * 100, 1),
        }

    # 요일별 결과
    day_result = {}
    for i, name in enumerate(day_names):
        day_result[name] = calc_stats(day_stats[i])

    # 기간별 결과
    period_result = {}
    for p, name in [
        ("month_start", "월초"),
        ("month_mid", "월중"),
        ("month_end", "월말"),
    ]:
        period_result[name] = calc_stats(period_stats[p])

    # 전체 결과
    total_result = calc_stats(
        {
            "count": total_count,
            "total_error": total_error,
            "over": total_over,
            "under": total_under,
        }
    )

    return Response(
        {"stats": {"total": total_result, "day": day_result, "period": period_result}}
    )


@api_view(["POST"])
def outbound_upload_excel(request):
    """
    바코드 통계 엑셀 파일 상의 [바코드, 제품명, 대분류, 수량] 데이터를 OutboundRecord 로 저장
    파일명(예: 바코드통계_20260401.xlsx)에서 날짜를 추출하여 처리
    """
    file_obj = request.FILES.get("file")
    if not file_obj:
        return Response(
            {"success": False, "message": "파일이 없습니다."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    filename = file_obj.name
    target_date_str = None

    # 1. 파일명에서 날짜 추출 (YYYYMMDD 형식)
    match = re.search(r"(\d{8})", filename)
    if match:
        raw_date = match.group(1)
        try:
            # 유효성 확인을 겸한 변환
            dt = datetime.strptime(raw_date, "%Y%m%d")
            target_date_str = dt.strftime("%Y-%m-%d")
        except:
            pass

    # 2. 파일명에서 추출 실패 시 프론트엔드 전송 값 사용
    if not target_date_str:
        target_date_str = request.data.get("date")

    if not target_date_str:
        # 최후의 보루: 어제 날짜
        target_date_str = (timezone.now() - timedelta(days=1)).strftime("%Y-%m-%d")

    try:
        df = pd.read_excel(file_obj)
        df = df.fillna("")

        # 컬럼 인덱스 찾기
        cols = _normalize_cols(df.columns)
        bc_idx = _find_col_index(cols, ["바코드", "barcode", "상품바코드"])
        name_idx = _find_col_index(cols, ["제품명", "상품명", "product"])
        cat_idx = _find_col_index(cols, ["대분류", "분류", "category"])
        qty_idx = _find_col_index(cols, ["수량", "quantity", "qty"])

        if bc_idx is None or qty_idx is None:
            return Response(
                {
                    "success": False,
                    "message": "엑셀 파일에 필수 열(바코드, 수량)이 없습니다.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 기존 해당 날짜 데이터 삭제
        with transaction.atomic():
            OutboundRecord.objects.filter(outbound_date=target_date_str).delete()

            records = []
            for _, row in df.iterrows():
                bc = str(row.iloc[bc_idx]).strip()
                qty = _parse_int(row.iloc[qty_idx])
                if not bc or qty <= 0:
                    continue

                records.append(
                    OutboundRecord(
                        outbound_date=target_date_str,
                        barcode=bc,
                        product_name=(
                            str(row.iloc[name_idx]).strip()
                            if name_idx is not None
                            else ""
                        )[:255],
                        category=(
                            str(row.iloc[cat_idx]).strip()
                            if cat_idx is not None
                            else "기타"
                        )[:100],
                        quantity=qty,
                        sales_amount=0,  # 파일엔 금액 정보 없음
                    )
                )

            if records:
                OutboundRecord.objects.bulk_create(records, batch_size=2000)

                # 신규 출고 품목 자동 복구: 3개월 미출고만 해제 (단종은 수동 유지)
                synced_barcodes = {r.barcode for r in records if r.barcode}
                if synced_barcodes:
                    _clear_no_outbound_3m_on_activity(barcodes=synced_barcodes)

        logger.info(
            f"Outbound Excel Upload Success: {target_date_str}, {len(records)} rows"
        )

        return Response(
            {
                "success": True,
                "created": len(records),
                "date": target_date_str,
                "message": f"{target_date_str} 데이터 {len(records)}건 업로드 완료",
            }
        )

    except Exception as e:
        logger.error(f"Outbound Excel Upload Error: {str(e)}\n{traceback.format_exc()}")
        return Response(
            {"success": False, "message": f"데이터 처리 중 오류 발생: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
def health_check(request):
    """Health check endpoint for rollback system monitoring"""
    return Response(
        {
            "status": "ok",
            "errorRate": 0,
            "uptime": 100,
            "memory": 50,
            "timestamp": timezone.now().isoformat(),
        }
    )


# ========== ProductionLog 자정 자동 이동 ==========
def _get_next_workday(date_obj):
    """다음 작업일 반환 (주말이면 월요일)"""
    from datetime import timedelta

    next_day = date_obj + timedelta(days=1)
    while next_day.weekday() >= 5:  # 5=Saturday, 6=Sunday
        next_day += timedelta(days=1)
    return next_day


def move_incomplete_production_logs():
    """
    자정 또는 수동 실행용: 이전 날짜 미완료 생산 계획을 다음 작업일로 이동
    - 이전 날짜(date < today) pending/start/stopped → 다음 작업일로
    - 오늘 미완료는 이동하지 않음 (작업 중인 계획 보존)
    """
    from datetime import datetime

    today = timezone.localdate()
    next_workday = _get_next_workday(today)

    # 이전 날짜 미완료만 조회 (오늘 것은 제외)
    incomplete_logs = ProductionLog.objects.filter(
        date__lt=today, status__in=["pending", "started", "stopped"]
    )

    moved_count = 0
    skipped_count = 0
    skipped_details = []
    errors = []

    for log in incomplete_logs:
        # 충돌 확인: 같은 조합이 이미 대상 날짜에 있는지
        conflict = ProductionLog.objects.filter(
            date=next_workday,
            machine_number=log.machine_number,
            mold_number=log.mold_number,
            product_name=log.product_name,
            color1=log.color1,
            color2=log.color2,
            unit=log.unit,
        ).exists()

        if conflict:
            skipped_count += 1
            skipped_details.append(
                {
                    "id": log.id,
                    "original_date": log.date.isoformat(),
                    "machine": log.machine_number,
                    "product": log.product_name,
                    "reason": "Already exists on target date",
                }
            )
            logger.warning(
                f"ProductionLog 충돌 스킵: ID={log.id}, date={log.date}→{next_workday}, machine={log.machine_number}, product={log.product_name}"
            )
            continue

        # 이동 시도
        try:
            log.date = next_workday
            log.save(update_fields=["date"])
            moved_count += 1
        except Exception as e:
            errors.append({"id": log.id, "date": log.date.isoformat(), "error": str(e)})
            logger.error(f"ProductionLog 이동 실패: ID={log.id}, error={e}")

    logger.info(
        f"ProductionLog 자동 이동 완료: {moved_count}개 이동됨, {skipped_count}개 스킵, {len(errors)}개 에러 → {next_workday}"
    )

    return {
        "moved_count": moved_count,
        "skipped_count": skipped_count,
        "skipped_details": skipped_details,
        "error_count": len(errors),
        "errors": errors,
        "next_workday": next_workday.isoformat(),
    }


@api_view(["POST"])
def production_move_incomplete(request):
    """미완료 생산 계획 다음 작업일로 이동 (API 호출용)"""
    result = move_incomplete_production_logs()
    return Response({"success": True, **result})


@api_view(["GET"])
def google_sheets_proxy(request):
    """Fetch a public Google Sheet CSV URL, convert to JSON.
    Query param: url=... (required)
    Returns JSON array of rows (list of dicts using header row as keys).
    CORS allowed for all origins.
    """
    url = request.GET.get("url", "").strip()
    if not url:
        return JsonResponse({"error": "url parameter is required"}, status=400)
    try:
        with urllib.request.urlopen(url) as resp:
            # Assume CSV content
            text = resp.read().decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            data = list(reader)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)
    response = JsonResponse(data, safe=False)
    response["Access-Control-Allow-Origin"] = "*"
    return response
    """미완료 생산 계획 다음 작업일로 이동 (API 호출용)"""
    result = move_incomplete_production_logs()
    return Response({"success": True, **result})


# ============================================================
# NotebookLM 분석 결과 API
# ============================================================
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_date
import json

@csrf_exempt
def outbound_analysis_list(request):
    """분석 결과 목록 조회/생성"""
    if request.method == 'GET':
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        period = request.GET.get('period')
        
        queryset = OutboundAnalysis.objects.all()
        if period:
            queryset = queryset.filter(period=period)
        if date_from:
            queryset = queryset.filter(date__gte=parse_date(date_from))
        if date_to:
            queryset = queryset.filter(date__lte=parse_date(date_to))
        
        data = [{
            'id': a.id,
            'date': str(a.date),
            'period': a.period,
            'summary': a.summary,
            'chart_data': a.chart_data,
            'table_data': a.table_data,
            'insights': a.insights,
            'recommendations': a.recommendations,
            'created_at': a.created_at.isoformat(),
        } for a in queryset[:100]]
        
        return JsonResponse({'success': True, 'data': data, 'count': len(data)})
    
    elif request.method == 'POST':
        try:
            body = json.loads(request.body)
            analysis = OutboundAnalysis.objects.create(
                date=parse_date(body.get('date', '')),
                period=body.get('period', 'daily'),
                summary=body.get('summary', {}),
                chart_data=body.get('chart_data', {}),
                table_data=body.get('table_data', {}),
                insights=body.get('insights', []),
                recommendations=body.get('recommendations', []),
                source_ids=body.get('source_ids', []),
            )
            return JsonResponse({'success': True, 'id': analysis.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

@csrf_exempt
def outbound_analysis_detail(request, pk):
    """분석 결과 상세 조회/수정/삭제"""
    try:
        analysis = OutboundAnalysis.objects.get(pk=pk)
    except OutboundAnalysis.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)
    
    if request.method == 'GET':
        return JsonResponse({
            'success': True,
            'data': {
                'id': analysis.id,
                'date': str(analysis.date),
                'period': analysis.period,
                'summary': analysis.summary,
                'chart_data': analysis.chart_data,
                'table_data': analysis.table_data,
                'insights': analysis.insights,
                'recommendations': analysis.recommendations,
                'source_ids': analysis.source_ids,
                'created_at': analysis.created_at.isoformat(),
            }
        })
    
    elif request.method == 'PUT':
        body = json.loads(request.body)
        for field in ['summary', 'chart_data', 'table_data', 'insights', 'recommendations']:
            if field in body:
                setattr(analysis, field, body[field])
        analysis.save()
        return JsonResponse({'success': True})
    
    elif request.method == 'DELETE':
        analysis.delete()
        return JsonResponse({'success': True})

@csrf_exempt
def outbound_analytics_summary(request):
    """전체 분석 요약 (VF 대시보드용)"""
    latest = OutboundAnalysis.objects.first()
    if not latest:
        return JsonResponse({'success': True, 'data': None})
    
    # 최근 7개 일별 분석에서 요약
    daily_list = OutboundAnalysis.objects.filter(period='daily').order_by('-date')[:7]
    
    return JsonResponse({
        'success': True,
        'data': {
            'latest_date': str(latest.date),
            'latest_insights': latest.insights[:3] if latest.insights else [],
            'daily_count': daily_list.count(),
            'trend': 'up' if len(daily_list) >= 2 and daily_list[0].summary.get('total', 0) > daily_list[1].summary.get('total', 0) else 'down',
        }
    })


# =============================================================================
# Prophet 예측 API
# =============================================================================

@csrf_exempt
def ai_production_forecast(request):
    """
    Prophet 기반 생산 예측 API
    
    POST /api/ai/production-forecast
    Body: {"product_name": "보노하우스 칵투스 커버형 비누 받침대", "horizon": 7}
    
    Response:
    {
        "success": true,
        "product_name": "보노하우스 칵투스 커버형 비누 받침대",
        "source": "prophet",
        "forecast": [{"date": "2026-05-20", "predicted": 40, "lower": 30, "upper": 55}, ...],
        "confidence_interval": {"lower_95": 30, "upper_95": 55, "avg_range": 20},
        "trend": "stable",
        "trend_percentage": 2.5,
        "safety_stock": 10,
        "weekly_seasonality": "confirmed"
    }
    """
    from sales_api.services.prophet_service import get_prophet_service
    
    if request.method == 'GET':
        # GET: 제품명 목록 또는 기본 예측
        cursor = connection.cursor()
        cursor.execute("""
            SELECT product_name, SUM(box_quantity) as total
            FROM outbound_records
            WHERE outbound_date > '2026-01-01'
            GROUP BY product_name
            ORDER BY total DESC
            LIMIT 20
        """)
        products = [{'name': row[0], 'total': row[1]} for row in cursor.fetchall()]
        return JsonResponse({'success': True, 'products': products})
    
    if request.method == 'POST':
        import json
        try:
            body = json.loads(request.body)
        except:
            body = request.data
        
        product_name = body.get('product_name', '')
        horizon = body.get('horizon', 7)
        
        if not product_name:
            return JsonResponse({
                'success': False,
                'message': 'product_name이 필요합니다.'
            }, status=400)
        
        service = get_prophet_service()
        result = service.get_prediction_summary(product_name, horizon=horizon)
        
        return JsonResponse(result)


@csrf_exempt
def ai_production_forecast_by_product(request, product_name):
    """
    특정 제품의 Prophet 예측 API
    
    GET /api/ai/production-forecast/<product_name>
    Query params: ?horizon=7
    """
    from sales_api.services.prophet_service import get_prophet_service
    
    horizon = request.GET.get('horizon', 7)
    
    service = get_prophet_service()
    result = service.get_prediction_summary(product_name, horizon=int(horizon))
    
    return JsonResponse(result)
