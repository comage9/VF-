import json
import re
from collections import defaultdict
from datetime import datetime, date, timedelta

from django.utils import timezone

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

from .models import TruckFreight

# 엑셀 수량 문자열 파싱용 정규식
_QTY_PAT = re.compile(r"\s*(\d+)\s*(파렛트|팔렛트|박스)?")


def _photo_url(obj: TruckFreight) -> str | None:
    """저장된 사진 URL (없으면 None)."""
    if not obj.photo:
        return None
    try:
        return obj.photo.url
    except Exception:
        return None


def _to_dict(obj: TruckFreight) -> dict:
    """모델 객체를 JSON 응답용 dict로 변환 (Serializer 대신 수동 변환)."""
    return {
        "id": obj.id,
        "date": obj.date.isoformat() if obj.date else None,
        "destination": obj.destination,
        "quantity": obj.quantity,
        "unit": obj.unit,
        "freight_fee": obj.freight_fee,
        "driver_name": obj.driver_name,
        "phone": obj.phone,
        "invoice_type": obj.invoice_type,
        "account_number": obj.account_number,
        "payment_status": obj.payment_status,
        "note": obj.note,
        "photo_url": _photo_url(obj),
        "has_photo": bool(obj.photo),
        "created_at": obj.created_at.isoformat() if obj.created_at else None,
        "updated_at": obj.updated_at.isoformat() if obj.updated_at else None,
    }


def _parse_date(value):
    """문자열/날짜를 date로 변환. 실패 시 None."""
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_int(value, default=0):
    if value is None or value == "":
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


@csrf_exempt
def api_list(request):
    """GET: 전체 운송비 목록, POST: 신규 생성"""
    if request.method == "GET":
        qs = TruckFreight.objects.all().order_by("-date", "-id")
        year = request.GET.get("year")
        if year:
            qs = qs.filter(date__year=int(year))
        data = [_to_dict(o) for o in qs]
        return JsonResponse({"data": data, "count": len(data)})

    if request.method == "POST":
        try:
            body = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({"error": "잘못된 JSON 형식입니다."}, status=400)

        obj = TruckFreight.objects.create(
            date=_parse_date(body.get("date")),
            destination=(body.get("destination") or "").strip(),
            quantity=_parse_int(body.get("quantity")),
            unit=body.get("unit") or "파렛트",
            freight_fee=_parse_int(body.get("freight_fee")),
            driver_name=(body.get("driver_name") or "").strip(),
            phone=(body.get("phone") or "").strip(),
            invoice_type=(body.get("invoice_type") or "").strip(),
            account_number=(body.get("account_number") or "").strip(),
            payment_status=body.get("payment_status") or "",
            note=(body.get("note") or "").strip(),
        )
        return JsonResponse({"data": _to_dict(obj), "ok": True}, status=201)

    return JsonResponse({"error": "지원하지 않는 메서드입니다."}, status=405)


@csrf_exempt
def api_detail(request, pk):
    """PATCH: 수정, DELETE: 삭제"""
    try:
        obj = TruckFreight.objects.get(pk=pk)
    except TruckFreight.DoesNotExist:
        return JsonResponse({"error": "데이터를 찾을 수 없습니다."}, status=404)

    if request.method == "PATCH":
        try:
            body = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({"error": "잘못된 JSON 형식입니다."}, status=400)

        if "date" in body:
            obj.date = _parse_date(body.get("date"))
        for field in ("destination", "driver_name", "phone", "invoice_type",
                      "account_number", "payment_status", "note", "unit"):
            if field in body:
                setattr(obj, field, (body.get(field) or "").strip() if field != "unit" else body.get(field))
        if "quantity" in body:
            obj.quantity = _parse_int(body.get("quantity"))
        if "freight_fee" in body:
            obj.freight_fee = _parse_int(body.get("freight_fee"))
        obj.save()
        return JsonResponse({"data": _to_dict(obj), "ok": True})

    if request.method == "DELETE":
        # 파일 정리
        if obj.photo:
            try:
                obj.photo.delete(save=False)
            except Exception:
                pass
        obj.delete()
        return JsonResponse({"ok": True})

    return JsonResponse({"error": "지원하지 않는 메서드입니다."}, status=405)


@csrf_exempt
def api_photo(request, pk):
    """POST: 사진 업로드 · DELETE: 사진 삭제

    multipart/form-data: photo=<image file>
    """
    try:
        obj = TruckFreight.objects.get(pk=pk)
    except TruckFreight.DoesNotExist:
        return JsonResponse({"error": "데이터를 찾을 수 없습니다."}, status=404)

    if request.method == "POST":
        f = request.FILES.get("photo") or request.FILES.get("file")
        if not f:
            return JsonResponse({"error": "photo 파일이 없습니다."}, status=400)
        # 크기 제한 10MB
        if f.size and f.size > 10 * 1024 * 1024:
            return JsonResponse({"error": "파일 크기는 10MB 이하여야 합니다."}, status=400)
        name = (getattr(f, "name", "") or "").lower()
        if name and not any(
            name.endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic")
        ):
            # content-type 완화: 확장자 없으면 image/* 허용
            ctype = (getattr(f, "content_type", "") or "").lower()
            if not ctype.startswith("image/"):
                return JsonResponse(
                    {"error": "이미지 파일만 업로드할 수 있습니다. (jpg/png/webp)"},
                    status=400,
                )
        # 기존 파일 교체
        if obj.photo:
            try:
                obj.photo.delete(save=False)
            except Exception:
                pass
        obj.photo = f
        obj.save(update_fields=["photo", "updated_at"])
        return JsonResponse({"ok": True, "data": _to_dict(obj)})

    if request.method == "DELETE":
        if obj.photo:
            try:
                obj.photo.delete(save=False)
            except Exception:
                pass
            obj.photo = None
            obj.save(update_fields=["photo", "updated_at"])
        return JsonResponse({"ok": True, "data": _to_dict(obj)})

    return JsonResponse({"error": "지원하지 않는 메서드입니다."}, status=405)


def api_summary(request):
    """GET: 월별 추이 + 계산서 종류별 통계"""
    qs = TruckFreight.objects.all()

    # 월별 추이
    monthly = defaultdict(lambda: {"count": 0, "fee": 0})
    dates = [o.date for o in qs if o.date]
    for o in qs:
        if not o.date:
            continue
        key = f"{o.date.year}-{o.date.month:02d}"
        monthly[key]["count"] += 1
        monthly[key]["fee"] += o.freight_fee

    # 데이터 기간 내 모든 월을 0건도 포함
    monthly_list = []
    if dates:
        min_d, max_d = min(dates), max(dates)
        y, m = min_d.year, min_d.month
        while (y, m) <= (max_d.year, max_d.month):
            key = f"{y}-{m:02d}"
            d = monthly[key]
            monthly_list.append({"month": key, "count": d["count"], "fee": d["fee"]})
            m += 1
            if m > 12:
                m = 1
                y += 1

    # 전체 합계 (모든 데이터 기준)
    total_fee = sum(o.freight_fee for o in qs)
    total_count = qs.count()

    # 최근 7일(일주일) 이내 등록 건수 — 사이드바/목록 NEW 표시용
    recent_cutoff = timezone.now() - timedelta(days=7)
    recent_new_count = qs.filter(created_at__gte=recent_cutoff).count()

    # 계산서 종류별 (빈값/미입력은 통계에서 제외)
    inv_map = defaultdict(lambda: {"count": 0, "fee": 0})
    for o in qs:
        if not o.invoice_type:
            continue
        inv_map[o.invoice_type]["count"] += 1
        inv_map[o.invoice_type]["fee"] += o.freight_fee

    # 비중 계산은 유효한(빈값 제외) 데이터 기준
    inv_total_fee = sum(d["fee"] for d in inv_map.values())
    inv_list = [
        {
            "invoice_type": inv,
            "count": d["count"],
            "fee": d["fee"],
            "ratio": round(d["fee"] / inv_total_fee * 100, 1) if inv_total_fee else 0,
        }
        for inv, d in sorted(inv_map.items(), key=lambda x: -x[1]["fee"])
    ]

    return JsonResponse({
        "monthly": monthly_list,
        "by_invoice": inv_list,
        "total_count": total_count,
        "total_fee": total_fee,
        "recent_new_count": recent_new_count,
    })


@csrf_exempt
def api_import(request):
    """POST: 엑셀 파일 업로드하여 일괄 import

    multipart/form-data: file=<엑셀 파일>
    또는 JSON: {"rows": [{"date":..., "destination":..., ...}, ...]}
    """
    if request.method != "POST":
        return JsonResponse({"error": "지원하지 않는 메서드입니다."}, status=405)

    # 엑셀 파일 업로드 처리
    uploaded = request.FILES.get("file")
    if uploaded:
        try:
            import openpyxl
        except ImportError:
            return JsonResponse({"error": "openpyxl이 설치되어 있지 않습니다."}, status=500)

        try:
            wb = openpyxl.load_workbook(uploaded, data_only=True)
        except Exception as e:
            return JsonResponse({"error": f"엑셀 파일을 읽을 수 없습니다: {e}"}, status=400)

        # "일자" 헤더가 있는 시트를 찾는다 (여러 시트 중 데이터 시트 자동 감지)
        ws = None
        header = {}
        date_col = None
        for candidate in wb.worksheets:
            cand_header = {}
            for c in range(1, min(candidate.max_column, 30) + 1):
                v = candidate.cell(1, c).value
                if v is not None:
                    cand_header[str(v).strip()] = c
            if "일자" in cand_header:
                ws = candidate
                header = cand_header
                date_col = cand_header["일자"]
                break
        if not ws or not date_col:
            return JsonResponse({"error": "엑셀에 '일자' 컬럼이 있는 시트를 찾을 수 없습니다."}, status=400)

        rows = []

        # 컬럼 매핑 (수량/단위 분리형과 통합형 모두 지원)
        def col(name):
            return header.get(name)

        for r in range(2, ws.max_row + 1):
            d = ws.cell(r, date_col).value
            if d is None:
                continue
            row = {"date": d}

            dest_c = col("납품처(입고처)") or col("납품처")
            if dest_c:
                row["destination"] = ws.cell(r, dest_c).value

            # 수량/단위: 분리형 우선, 없으면 통합형 파싱
            qty_c = col("수량")
            unit_c = col("단위")
            fee_c = col("운송비")
            if unit_c and fee_c:
                # 분리형: 수량(숫자), 단위, 운송비
                row["quantity"] = ws.cell(r, qty_c).value if qty_c else 0
                row["unit"] = ws.cell(r, unit_c).value or "파렛트"
                row["freight_fee"] = ws.cell(r, fee_c).value or 0
            else:
                # 통합형 또는 운송비가 수량 바로 다음
                qty_val = ws.cell(r, qty_c).value if qty_c else None
                fee_val = ws.cell(r, (fee_c or (qty_c + 1) if qty_c else None)).value
                qty, unit = _parse_qty_str(qty_val)
                row["quantity"] = qty or 0
                row["unit"] = unit or "파렛트"
                row["freight_fee"] = _parse_int(fee_val)

            driver_c = col("기사명")
            if driver_c:
                row["driver_name"] = ws.cell(r, driver_c).value
            phone_c = col("연락처")
            if phone_c:
                row["phone"] = ws.cell(r, phone_c).value
            inv_c = col("계산서 종류")
            if inv_c:
                row["invoice_type"] = ws.cell(r, inv_c).value
            acct_c = col("계좌번호")
            if acct_c:
                row["account_number"] = ws.cell(r, acct_c).value
            pay_c = col("입금확인")
            if pay_c:
                row["payment_status"] = ws.cell(r, pay_c).value or ""
            note_c = col("비고")
            if note_c:
                row["note"] = ws.cell(r, note_c).value

            rows.append(row)
    else:
        # JSON body
        try:
            body = json.loads(request.body)
            rows = body.get("rows", [])
        except (json.JSONDecodeError, ValueError):
            return JsonResponse({"error": "잘못된 요청 형식입니다."}, status=400)

    # DB 저장
    created = 0
    skipped = 0
    for row in rows:
        d = _parse_date(row.get("date"))
        if not d:
            skipped += 1
            continue
        qty, unit = _parse_qty_str(row.get("quantity")) if isinstance(row.get("quantity"), str) else (row.get("quantity"), row.get("unit"))
        TruckFreight.objects.create(
            date=d,
            destination=(str(row.get("destination") or "")).strip(),
            quantity=_parse_int(qty),
            unit=unit or "파렛트",
            freight_fee=_parse_int(row.get("freight_fee")),
            driver_name=(str(row.get("driver_name") or "")).strip(),
            phone=(str(row.get("phone") or "")).strip(),
            invoice_type=(str(row.get("invoice_type") or "")).strip(),
            account_number=(str(row.get("account_number") or "")).strip(),
            payment_status=row.get("payment_status") or "",
            note=(str(row.get("note") or "")).strip(),
        )
        created += 1

    return JsonResponse({"ok": True, "created": created, "skipped": skipped})


def _parse_qty_str(raw):
    """수량 문자열을 (숫자, 단위)로 분리. 이미 숫자면 (숫자, None)."""
    if raw is None:
        return (None, None)
    if isinstance(raw, (int, float)):
        return (int(raw), None)
    s = str(raw).strip().rstrip("`").strip()
    m = _QTY_PAT.match(s)
    if not m:
        return (None, None)
    qty = int(m.group(1))
    unit = m.group(2) or ""
    if unit == "팔렛트":
        unit = "파렛트"
    return (qty, unit)
